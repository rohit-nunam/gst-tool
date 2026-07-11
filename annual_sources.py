#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANNUAL SOURCES  --  Ledgers (Cash/Credit/Liability), GST-Prime TPST, and the
portal's own "Tax liability and ITC comparison" report.

These are FY-wide, month-agnostic files (one file covers the whole year),
unlike GSTR-1/3B/2B/E-Inv/EWB which are one-file-per-period. All parsers here
key their output by month label 'Mon-YY' (e.g. 'Apr-22') so they can be
cross-joined against each other and, later, against the monthly GSTR set.
"""

import csv
import re
import openpyxl

MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
MONTH_NUM = {v: k for k, v in MONTH_ABBR.items()}


def _num(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").replace("₹", "").strip()
    if s in ("", "-", "–", "NA"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _period_key(datestr):
    """'21-05-2022' -> 'May-22'. Returns None if unparseable."""
    if not datestr or datestr == "-":
        return None
    m = re.match(r"(\d{2})-(\d{2})-(\d{4})", datestr.strip())
    if not m:
        return None
    _, mm, yyyy = m.groups()
    mm = int(mm)
    return f"{MONTH_ABBR.get(mm, '?')}-{yyyy[2:]}"


def _tax_period_key(tp):
    """'Mar-22' (as already given in ledger 'Tax Period' column) -> normalise to 'Mar-22'."""
    if not tp or tp == "-":
        return None
    tp = tp.strip()
    m = re.match(r"([A-Za-z]{3})-(\d{2,4})", tp)
    if not m:
        return None
    mon, yy = m.groups()
    yy = yy[-2:]
    return f"{mon[:3].title()}-{yy}"


# ======================================================================
# CASH LEDGER  /  LIABILITY REGISTER  (same 8-group-of-6 layout)
# ======================================================================
_HEAD_GROUPS_8 = ["IGST", "CGST", "SGST", "CESS"]  # first 4 groups = Debited/Credited
# groups 5-8 = running balances (same order)


def parse_cash_or_liability_ledger(path, kind):
    """kind: 'cash' or 'liability'. Returns dict(opening={...}, transactions=[...],
    monthly_by_tax_period={period: {...totals...}}, monthly_by_txn_date={period: {...}})."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    if kind == "cash":
        # cols: 0 Sr.No,1 Date,2 Time,3 Reporting date,4 Reference No.,5 Tax Period,
        #       6 Description,7 Transaction Type, 8-31 Debited/Credited (4x6), 32-55 Balance (4x6)
        col_date, col_ref, col_taxperiod, col_desc, col_ttype = 1, 4, 5, 6, 7
        deb_start = 8
    else:  # liability register
        # cols: 0 Sr.No,1 Date,2 Reference No.,3 Ledger Used,4 Description,5 Transaction Type,
        #       6-29 Debited/Credited (4x6), 30-53 Balance (4x6)
        col_date, col_ref, col_taxperiod, col_desc, col_ttype = 1, 2, None, 4, 5
        deb_start = 6

    opening = None
    transactions = []
    for r in rows:
        if not r or len(r) < deb_start + 24:
            continue
        first = (r[0] or "").strip()
        desc = (r[col_desc] or "").strip() if col_desc < len(r) else ""
        if not first and desc != "Opening Balance":
            continue
        if not (first.isdigit() or first in ("-",) or desc == "Opening Balance"):
            continue
        date = (r[col_date] or "").strip() if col_date < len(r) else "-"
        ref = (r[col_ref] or "").strip() if col_ref < len(r) else "-"
        taxp = (r[col_taxperiod] or "").strip() if col_taxperiod is not None and col_taxperiod < len(r) else "-"
        ttype = (r[col_ttype] or "").strip() if col_ttype < len(r) else "-"
        extra = (r[3] or "").strip() if kind == "liability" and len(r) > 3 else None  # Ledger Used

        heads = {}
        bal_start = deb_start + 24
        for gi, head in enumerate(_HEAD_GROUPS_8):
            base = deb_start + gi * 6
            bbase = bal_start + gi * 6
            tax = _num(r[base]) if base < len(r) else 0.0
            total = _num(r[base + 5]) if base + 5 < len(r) else 0.0
            bal_total = _num(r[bbase + 5]) if bbase + 5 < len(r) else 0.0
            heads[head] = dict(tax=tax, total=total, balance=bal_total)
        total_debited_or_credited = sum(h["total"] for h in heads.values())

        rec = dict(date=date, ref=ref, tax_period=taxp, description=desc, ttype=ttype,
                    heads=heads, total=total_debited_or_credited,
                    balance_total=sum(h["balance"] for h in heads.values()),
                    ledger_used=extra if kind == "liability" else None)

        if desc == "Opening Balance":
            opening = rec
            continue
        if desc == "Closing Balance":
            continue
        transactions.append(rec)

    monthly_by_taxperiod = {}
    monthly_by_txndate = {}
    for t in transactions:
        sign = 1 if t["ttype"].lower() == "credit" else -1
        amt = t["total"]
        tpk = _tax_period_key(t["tax_period"]) if kind == "cash" else None
        dpk = _period_key(t["date"])
        if tpk:
            monthly_by_taxperiod.setdefault(tpk, dict(credited=0.0, debited=0.0, net=0.0))
            if t["ttype"].lower() == "credit":
                monthly_by_taxperiod[tpk]["credited"] += amt
            else:
                monthly_by_taxperiod[tpk]["debited"] += amt
            monthly_by_taxperiod[tpk]["net"] += sign * amt
        if dpk:
            monthly_by_txndate.setdefault(dpk, dict(credited=0.0, debited=0.0, net=0.0))
            if t["ttype"].lower() == "credit":
                monthly_by_txndate[dpk]["credited"] += amt
            else:
                monthly_by_txndate[dpk]["debited"] += amt
            monthly_by_txndate[dpk]["net"] += sign * amt

    return dict(opening=opening, transactions=transactions,
                monthly_by_tax_period=monthly_by_taxperiod,
                monthly_by_txn_date=monthly_by_txndate)


# ======================================================================
# CREDIT LEDGER  (different layout: 1 Credit/Debit block + 1 Balance block)
# ======================================================================
def parse_credit_ledger(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    # cols: 0 Sr.No,1 Date,2 Reference No.,3 Tax Period,4 Description,5 Transaction Type,
    #       6-10 Credit/Debit[IGST,CGST,SGST,CESS,Total], 11-15 Balance[IGST,CGST,SGST,CESS,Total]
    opening = None
    transactions = []
    for r in rows:
        if not r or len(r) < 16:
            continue
        desc = (r[4] or "").strip()
        first = (r[0] or "").strip()
        if not (first.isdigit() or desc == "Opening Balance"):
            continue
        rec = dict(date=(r[1] or "").strip(), ref=(r[2] or "").strip(),
                    tax_period=(r[3] or "").strip(), description=desc,
                    ttype=(r[5] or "").strip(),
                    igst=_num(r[6]), cgst=_num(r[7]), sgst=_num(r[8]), cess=_num(r[9]),
                    total=_num(r[10]),
                    bal_igst=_num(r[11]), bal_cgst=_num(r[12]), bal_sgst=_num(r[13]),
                    bal_cess=_num(r[14]), bal_total=_num(r[15]))
        if desc == "Opening Balance":
            opening = rec
            continue
        transactions.append(rec)

    monthly_by_taxperiod = {}
    for t in transactions:
        tpk = _tax_period_key(t["tax_period"])
        if not tpk:
            continue
        m = monthly_by_taxperiod.setdefault(
            tpk, dict(credited=0.0, debited=0.0, net=0.0, accrued_desc=set()))
        sign = 1 if t["ttype"].lower() == "credit" else -1
        if t["ttype"].lower() == "credit":
            m["credited"] += t["total"]
            m["accrued_desc"].add(t["description"])
        else:
            m["debited"] += t["total"]
        m["net"] += sign * t["total"]
    for m in monthly_by_taxperiod.values():
        m["accrued_desc"] = sorted(m["accrued_desc"])

    return dict(opening=opening, transactions=transactions,
                monthly_by_tax_period=monthly_by_taxperiod)


# ======================================================================
# GST-PRIME TPST  (12-month self-filing summary, Excel)
# ======================================================================
def parse_tpst(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = None
    for i, r in enumerate(rows):
        if r and r[0] == "SNo":
            hdr_i = i
            break
    if hdr_i is None:
        return {}
    hdr = [str(c).strip() if c else "" for c in rows[hdr_i]]
    H = {h: i for i, h in enumerate(hdr)}
    # This report variant's real columns: SNo, Return Period, No Of Sellers,
    # No Of Invoices, Taxable Value, SGST, CGST, IGST, CESS, Total GST.
    # It carries NO filing-date / net-ITC-claimed / cash-paid columns at all
    # (an older assumed layout had those; this file genuinely doesn't) -- so
    # those three come through as None rather than a guessed number. The
    # existing downstream cross-checks in build_annual_workbook.py already
    # turn a None into a visible "N/A" rather than a false PASS/REVIEW, so
    # nothing is silently hidden.
    required = ("Taxable Value", "Total GST")
    missing = [c for c in required if c not in H]
    if missing:
        raise ValueError(f"TPST file {path!r} is missing expected column(s) {missing} -- "
                          f"columns found: {hdr}")
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or r[H.get("Return Period", 1)] is None:
            continue
        period = str(r[H.get("Return Period", 1)]).strip()  # e.g. 'Mar-2023'
        m = re.match(r"([A-Za-z]{3})-(\d{4})", period)
        key = f"{m.group(1)}-{m.group(2)[2:]}" if m else period
        g = lambda name: r[H[name]] if name in H and H[name] < len(r) else None
        out[key] = dict(
            filing_date=None,   # not present in this report variant
            outward_taxable=_num(g("Taxable Value")),
            net_itc_claimed=None,    # not present in this report variant
            total_cash_paid=None,    # not present in this report variant
            total_tax_liability=_num(g("Total GST")),  # = SGST+CGST+IGST+CESS for the period
            no_of_sellers=_num(g("No Of Sellers")), no_of_invoices=_num(g("No Of Invoices")),
            sgst=_num(g("SGST")), cgst=_num(g("CGST")), igst=_num(g("IGST")), cess=_num(g("CESS")),
        )
    return out


# ======================================================================
# PORTAL "Tax liability and ITC comparison" report (Excel, Comparison Summary sheet)
# ======================================================================
def parse_portal_comparison(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Comparison Summary" not in wb.sheetnames:
        return {}
    ws = wb["Comparison Summary"]
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for r in rows:
        if not r or not r[0]:
            continue
        period = str(r[0]).strip()
        m = re.match(r"([A-Za-z]{3})-(\d{2})$", period)
        if not m:
            continue
        key = f"{m.group(1)}-{m.group(2)}"
        def _blank(v):
            return v is None or (isinstance(v, str) and v.strip() == "")

        out[key] = dict(
            gstr1_liability=_num(r[1]), gstr3b_liability=_num(r[2]),
            diff_liability=_num(r[3]), cum_diff_liability=_num(r[4]),
            itc_3b_unadj=_num(r[6]), itc_2b=_num(r[7]), diff_itc_unadj=_num(r[8]),
            cum_diff_itc_unadj=_num(r[9]),
            itc_3b_adj=None if (len(r) <= 11 or _blank(r[11])) else _num(r[11]),
            diff_itc_adj=None if (len(r) <= 12 or _blank(r[12])) else _num(r[12]),
            cum_diff_itc_adj=None if (len(r) <= 13 or _blank(r[13])) else _num(r[13]),
        )
    return out


if __name__ == "__main__":
    cash = parse_cash_or_liability_ledger("Electronic_Cash_Ledger__2___1_.csv", "cash")
    credit = parse_credit_ledger("ElectronicCreditLedger__2___1_.csv")
    liab = parse_cash_or_liability_ledger(
        "Electronic_Liability_Register_Electronic_Liability_Register_2__1_.csv", "liability")
    tpst = parse_tpst("GST-Prime_TPST_Taxpayer_Profile-_Statements__1_3B_22-23.xlsx")
    comp = parse_portal_comparison("2022-23_05AAECM6380J1ZA_Tax_liability_and_ITC_comparison__1_.xlsx")

    print("CASH opening:", cash["opening"])
    print("CASH txns:", len(cash["transactions"]))
    print("CASH monthly (by tax period):", dict(list(cash["monthly_by_tax_period"].items())[:3]))

    print("\nCREDIT opening:", credit["opening"])
    print("CREDIT txns:", len(credit["transactions"]))
    print("CREDIT monthly (by tax period) Jul-22:", credit["monthly_by_tax_period"].get("Jul-22"))

    print("\nLIABILITY opening:", liab["opening"])
    print("LIABILITY txns:", len(liab["transactions"]))

    print("\nTPST months:", list(tpst.keys()))
    print("TPST Jan-23:", tpst.get("Jan-23"))

    print("\nPORTAL COMPARISON months:", list(comp.keys()))
    print("PORTAL Jan-23:", comp.get("Jan-23"))
