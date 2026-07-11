#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANNUAL E-WAY-BILL PARSER
=========================
The whole-FY Inward/Outward EWB workbooks (one file each, not one-per-month)
don't follow one consistent sheet-name or header-text convention:
  - Outward file: real data on sheet 'OUT EWB', doc-no/doc-date columns are
    literally named 'INVOICE'/'DATE'. A second sheet 'R1' is a bonus annual
    B2B invoice register (not EWB data) -- kept separately if useful later.
  - Inward file: sheet 'Sheet1' is empty; real data is on 'merged_sheet_1',
    where the doc-no/doc-date columns are mislabelled 'Doc'/'&' (a broken
    merged-cell header from the source export).

Despite the text differences, both files share the SAME column order for the
first 13 columns. So this parser finds the right sheet by header CONTENT
(must contain 'EWB No.' + 'From GSTIN & Name' + 'To GSTIN & Name'), then
reads the doc-number/doc-date pair by POSITION (the two columns immediately
after 'EWB No. & Dt.'), not by their (unreliable) header text.

Every row also gets a `month` tag ('Mon-YY') derived from the EWB date, so a
per-month engine can filter this annual list down to one period.
"""

import re
import datetime as _dt
import openpyxl

GSTIN_RE = re.compile(r"(\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d])")
MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def _num(v):
    if v is None:
        return 0.0
    s = str(v).replace(",", "").strip()
    if s in ("", "-", "#N/A", "NA"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _gstin_of(cell):
    if not cell:
        return ""
    m = GSTIN_RE.search(str(cell))
    return m.group(1) if m else str(cell).split("/")[0].strip()


def _as_date(v):
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return _dt.datetime.strptime(v.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _split_ewb_no_dt(v):
    """'351569103816 - 04/03/2023 11:43:00' -> (ewb_no, date)."""
    if not v:
        return "", None
    s = str(v)
    parts = s.split(" - ", 1)
    ewbno = parts[0].strip()
    date = None
    if len(parts) > 1:
        m = re.search(r"(\d{2})/(\d{2})/(\d{4})", parts[1])
        if m:
            dd, mm, yyyy = m.groups()
            date = _dt.date(int(yyyy), int(mm), int(dd))
    return ewbno, date


def _find_data_sheet(wb):
    """Return worksheet whose header row matches the EWB column signature."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            hdr = [str(c).strip() if c else "" for c in row]
            if "EWB No." in hdr and "From GSTIN & Name" in hdr and "To GSTIN & Name" in hdr:
                return ws, hdr, row
    return None, None, None


def parse_annual_ewb(path):
    """Return list of dicts: ewbno, ewbdate, month, docno, docdate, from_gstin,
    to_gstin, assess, taxval, hsn, vehicle, rate."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, hdr, _ = _find_data_sheet(wb)
    if ws is None:
        return []
    H = {h: i for i, h in enumerate(hdr) if h}
    idx_ewbdt = H.get("EWB No. & Dt.")
    doc_no_col = idx_ewbdt + 1 if idx_ewbdt is not None else None
    doc_dt_col = idx_ewbdt + 2 if idx_ewbdt is not None else None

    def g(r, name):
        i = H.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = []
    rows = ws.iter_rows(min_row=1, values_only=True)
    next(rows)  # header
    for r in rows:
        if not any(r):
            continue
        if str(r[0] or "").strip() in ("EWB No.", ""):
            continue
        ewbno, ewbdate = _split_ewb_no_dt(g(r, "EWB No. & Dt."))
        docno = str(r[doc_no_col]).strip() if doc_no_col is not None and doc_no_col < len(r) and r[doc_no_col] else ""
        docdate = _as_date(r[doc_dt_col]) if doc_dt_col is not None and doc_dt_col < len(r) else None
        month = f"{MONTH_ABBR.get(ewbdate.month,'?')}-{str(ewbdate.year)[2:]}" if ewbdate else None
        out.append(dict(
            ewbno=str(g(r, "EWB No.") or ewbno).strip(), ewbdate=ewbdate, month=month,
            docno=docno, docdate=docdate,
            from_gstin=_gstin_of(g(r, "From GSTIN & Name")),
            from_name=str(g(r, "From GSTIN & Name") or "").split("/", 1)[-1].strip(),
            to_gstin=_gstin_of(g(r, "To GSTIN & Name")),
            to_name=str(g(r, "To GSTIN & Name") or "").split("/", 1)[-1].strip(),
            from_place=str(g(r, "From Place & Pin") or "").strip(),
            to_place=str(g(r, "To Place & Pin") or "").strip(),
            assess=_num(g(r, "Assess Val.")), taxval=_num(g(r, "Tax Val.")),
            hsn=str(g(r, "HSN Code") or "").strip(),
            vehicle=str(g(r, "Latest Vehicle No.") or "").strip(),
            rate=_num(g(r, "TAX RATE")),
        ))
    return out


def filter_by_month(ewb_rows, month_key):
    """month_key e.g. 'Jan-23' -- matches on EWB date's month (not doc date)."""
    return [r for r in ewb_rows if r["month"] == month_key]


if __name__ == "__main__":
    import sys
    inw = parse_annual_ewb(sys.argv[1] if len(sys.argv) > 1 else "EWB_IN_annual.xlsx")
    out = parse_annual_ewb(sys.argv[2] if len(sys.argv) > 2 else "EWB_OUT_annual.xlsx")
    print("Inward EWB rows:", len(inw))
    print("Outward EWB rows:", len(out))
    print("Sample inward:", inw[0] if inw else None)
    print("Sample outward:", out[0] if out else None)
    from collections import Counter
    print("Inward by month:", Counter(r["month"] for r in inw))
    print("Outward by month:", Counter(r["month"] for r in out))
    jan_out = filter_by_month(out, "Jan-23")
    print("Outward EWB in Jan-23:", len(jan_out))
