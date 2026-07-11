#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FOLDER CLASSIFIER  (merged-file model)
========================================
Scans a folder and identifies ONE merged workbook per document type (whole
FY, however many months it currently covers), plus the whole-FY annual
sources -- all by CONTENT signature, never by filename.

Returns:
  gstr1_merged, gstr3b_merged, einv_merged, gstr2b_merged : paths (or None)
  gstr1_months, gstr3b_months, einv_months, gstr2b_months : sets of 'Mon-YY'
      labels actually found inside each merged file (for coverage reporting --
      discovering a month here does NOT mean every sub-sheet has data for it,
      just that the file's period markers include it)
  ewb_out_annual, ewb_in_annual : whole-FY EWB workbook paths
  cash_ledger, credit_ledger, liab_ledger : CSV paths
  tpst, portal_comparison : xlsx paths
  bo_profile    : pdf path
  self_gstin, company_name
"""

import os
import re
import glob
import csv
import openpyxl
import merged_period_utils as mpu


def _sheetnames(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = list(wb.sheetnames)
    wb.close()
    return sn


def _looks_like_gstr3b_merged(path):
    """Content signature for the merged GSTR-3B workbook: at least one sheet
    contains the literal 'Form GSTR-3B' banner text. Sheet NAMES (e.g.
    'Jan_2022-23') are never consulted, per instruction."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            for row in wb[sn].iter_rows(min_row=1, max_row=3, values_only=True):
                if any(c and "Form GSTR-3B" in str(c) for c in row):
                    return True
    finally:
        wb.close()
    return False


def _gstr1_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["b2b, sez, de_inv"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 3)


def _einv_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["b2b, sez, de"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 3)


def _gstr2b_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["ITC Available"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return mpu.months_present(rows, 0)


def _gstr3b_months(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    months = set()
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            fy = tp = None
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c not in (None, "")]
                if not cells:
                    continue
                key = cells[0].upper()
                if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
                    fy = cells[1]
                elif key == "TAX PERIOD" and len(cells) >= 2:
                    tp = cells[1]
                if fy and tp:
                    break
            if fy and tp:
                try:
                    months.update(mpu.months_for_tax_period(fy, tp))
                except mpu.PeriodParseError:
                    pass  # a stray non-data sheet with unrelated Year/Tax Period-looking text
    finally:
        wb.close()
    return months


def _csv_first_line(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            line = " ".join(str(c) for c in row if c).strip()
            if line:
                return line
    return ""


def _read_me_gstin_and_name(gstr1_path):
    """Content-based GSTIN + Legal Name from the merged GSTR-1's 'Read me' sheet."""
    wb = openpyxl.load_workbook(gstr1_path, read_only=True, data_only=True)
    try:
        sn = "Read me" if "Read me" in wb.sheetnames else wb.sheetnames[0]
        gstin = name = None
        for row in wb[sn].iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue
            key = cells[0].upper()
            if key == "GSTIN" and len(cells) >= 2:
                gstin = cells[1]
            elif key == "LEGAL NAME" and len(cells) >= 2:
                name = cells[1]
        return gstin, name
    finally:
        wb.close()


def _build_month_file_map(files, months_fn, label):
    """files: list of candidate paths for ONE document type (e.g. every merged
    GSTR-1 workbook found in the folder, however many FYs that spans).
    months_fn: the _gstr1_months/_gstr3b_months/etc function for that type.
    Returns ({month_label: filepath}, warnings). If two files both claim the
    SAME month, that is a genuine ambiguity (e.g. two overlapping exports)
    and is NOT silently resolved by picking one -- it's reported as a
    warning and the LATER file (by mtime) wins, so at least the behaviour is
    deterministic and visible, not silently arbitrary."""
    month_map = {}
    warnings = []
    # newest-mtime-last, so a later duplicate legitimately overrides an
    # earlier one (e.g. a corrected re-export) rather than the reverse
    for f in sorted(files, key=lambda p: os.path.getmtime(p)):
        try:
            months = months_fn(f)
        except Exception as ex:
            warnings.append(f"{label} file {f!r} could not be read for month coverage: {ex}")
            continue
        for m in months:
            if m in month_map and month_map[m] != f:
                warnings.append(f"{label}: month {m!r} found in BOTH {month_map[m]!r} and {f!r} -- "
                                 f"using {f!r} (newer file). Verify these aren't two different FYs "
                                 f"that happen to reuse the same 'Mon-YY' label by mistake.")
            month_map[m] = f
    return month_map, warnings


def classify_folder(folder="."):
    xlsx = sorted(glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xlsm")))
    csvs = sorted(glob.glob(os.path.join(folder, "*.csv")))
    pdfs = sorted(glob.glob(os.path.join(folder, "*.pdf")))

    # CHANGED (multi-year support): every doc type now collects a LIST of
    # matching files, not a single overwritten path -- a taxpayer with 5
    # years of data may supply 5 separate merged GSTR-1 workbooks (one per
    # FY), each internally covering up to 12 months via its own period
    # markers. Previously this loop did `gstr1_merged = f` on every match,
    # so only the LAST file (by glob/sort order) survived and every earlier
    # FY's data was silently discarded without any error or warning.
    gstr1_files, gstr3b_files, einv_files, gstr2b_files = [], [], [], []
    tpst_files, portal_comparison_files, ewb_candidates = [], [], []
    bo_profile_files = []
    gstr9_files, gstr9c_files, table8a_files, bs_pl_files = [], [], [], []

    for f in xlsx:
        sn = set(_sheetnames(f))
        if not sn:
            continue
        if "b2b, sez, de_inv" in sn and "hsn" in sn:
            gstr1_files.append(f); continue
        if "ITC Available" in sn and "B2B" in sn:
            gstr2b_files.append(f); continue
        if "b2b, sez, de" in sn and "b2b, sez, de_inv" not in sn:
            einv_files.append(f); continue
        if "Comparison Summary" in sn:
            portal_comparison_files.append(f); continue
        # Table 8A: government-standard export, always has this exact sheet set
        if {"B2B", "B2BA", "CDNR", "CDNRA"}.issubset(sn) and "Read me" in sn:
            table8a_files.append(f); continue
        if _looks_like_gstr3b_merged(f):
            gstr3b_files.append(f); continue
        # TPST: single sheet, first cell mentions 'Taxpayer Profile'
        if len(sn) <= 2:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            top = " ".join(str(c) for row in ws.iter_rows(min_row=1, max_row=2, values_only=True)
                            for c in row if c)
            wb.close()
            if "Taxpayer Profile" in top or "TP:ST" in top or "TP: ST" in top:
                tpst_files.append(f); continue
        # Annual EWB: has 'EWB No.' + 'From GSTIN & Name' header on some sheet
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        found_ewb = False
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                hdr = [str(c).strip() if c else "" for c in row]
                if "EWB No." in hdr and "From GSTIN & Name" in hdr and "To GSTIN & Name" in hdr:
                    ewb_candidates.append(f)
                    found_ewb = True
                    break
            if found_ewb:
                break
        wb.close()

    for f in pdfs:
        try:
            head = ""
            import pdfplumber
            with pdfplumber.open(f) as pdf:
                head = (pdf.pages[0].extract_text() or "")[:400] if pdf.pages else ""
        except Exception:
            head = ""
        hl = head.lower()
        if "gstr-9c" in hl or "reconciliation statement" in hl:
            gstr9c_files.append(f)
        elif "gstr-9" in hl or "annual return" in hl:
            gstr9_files.append(f)
        elif not head.strip():
            # No extractable text -- likely a scanned financial statement or the BO Profile
            # (BO Profile is text-based and caught separately below by its own signature scan
            # in the original single-PDF slot; a genuinely blank/scanned PDF here is presumed
            # to be the Balance Sheet/P&L export unless proven otherwise).
            bs_pl_files.append(f)
        else:
            bo_profile_files.append(f)

    # ---- direction for the annual EWB files: self-GSTIN mostly in From -> outward ----
    # (unchanged logic; still works across multiple years' EWB files pooled together)
    self_gstin = None
    ewb_out_files, ewb_in_files = [], []
    if ewb_candidates:
        import ewb_annual_parser as ewbp
        from collections import Counter
        parsed = {f: ewbp.parse_annual_ewb(f) for f in ewb_candidates}
        freq = Counter()
        gstin_re = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d]$")
        for f, rows in parsed.items():
            for r in rows:
                for g in (r["from_gstin"], r["to_gstin"]):
                    if gstin_re.match(g or ""):
                        freq[g] += 1
        if freq:
            self_gstin = freq.most_common(1)[0][0]
        for f, rows in parsed.items():
            fr = sum(1 for r in rows if r["from_gstin"] == self_gstin)
            to = sum(1 for r in rows if r["to_gstin"] == self_gstin)
            (ewb_out_files if fr >= to else ewb_in_files).append(f)

    cash_ledgers, credit_ledgers, liab_ledgers = [], [], []
    for c in csvs:
        line = _csv_first_line(c).lower()
        if "cash ledger" in line:
            cash_ledgers.append(c)
        elif "credit ledger" in line:
            credit_ledgers.append(c)
        elif "liability register" in line or "liability ledger" in line:
            liab_ledgers.append(c)

    # self_gstin / company_name refinement from the FIRST merged GSTR-1's Read me sheet
    company_name = None
    if gstr1_files:
        g1_gstin, g1_name = _read_me_gstin_and_name(gstr1_files[0])
        self_gstin = self_gstin or g1_gstin
        company_name = g1_name

    gstr1_month_map, w1 = _build_month_file_map(gstr1_files, _gstr1_months, "GSTR-1")
    gstr3b_month_map, w2 = _build_month_file_map(gstr3b_files, _gstr3b_months, "GSTR-3B")
    einv_month_map, w3 = _build_month_file_map(einv_files, _einv_months, "E-Invoice")
    gstr2b_month_map, w4 = _build_month_file_map(gstr2b_files, _gstr2b_months, "GSTR-2B")
    warnings = w1 + w2 + w3 + w4

    return dict(
        # NEW multi-year keys (month-level file resolution -- use these)
        gstr1_month_map=gstr1_month_map, gstr3b_month_map=gstr3b_month_map,
        einv_month_map=einv_month_map, gstr2b_month_map=gstr2b_month_map,
        gstr1_files=gstr1_files, gstr3b_files=gstr3b_files,
        einv_files=einv_files, gstr2b_files=gstr2b_files,
        classify_warnings=warnings,
        # BACKWARD-COMPAT single-path keys (first file found; old single-FY code
        # that doesn't know about *_month_map still works unchanged)
        gstr1_merged=gstr1_files[0] if gstr1_files else None,
        gstr3b_merged=gstr3b_files[0] if gstr3b_files else None,
        einv_merged=einv_files[0] if einv_files else None,
        gstr2b_merged=gstr2b_files[0] if gstr2b_files else None,
        gstr1_months=set(gstr1_month_map), gstr3b_months=set(gstr3b_month_map),
        einv_months=set(einv_month_map), gstr2b_months=set(gstr2b_month_map),
        # EWB -- now lists (one set of annual workbooks per FY supplied)
        ewb_out_files=ewb_out_files, ewb_in_files=ewb_in_files,
        ewb_out_annual=ewb_out_files[0] if ewb_out_files else None,
        ewb_in_annual=ewb_in_files[0] if ewb_in_files else None,
        # Annual-level sources -- now lists (one set per FY)
        cash_ledgers=cash_ledgers, credit_ledgers=credit_ledgers, liab_ledgers=liab_ledgers,
        cash_ledger=cash_ledgers[0] if cash_ledgers else None,
        credit_ledger=credit_ledgers[0] if credit_ledgers else None,
        liab_ledger=liab_ledgers[0] if liab_ledgers else None,
        tpst_files=tpst_files, portal_comparison_files=portal_comparison_files,
        bo_profile_files=bo_profile_files,
        tpst=tpst_files[0] if tpst_files else None,
        portal_comparison=portal_comparison_files[0] if portal_comparison_files else None,
        bo_profile=bo_profile_files[0] if bo_profile_files else None,
        # NEW optional annual-return-side documents
        gstr9_files=gstr9_files, gstr9c_files=gstr9c_files,
        table8a_files=table8a_files, bs_pl_files=bs_pl_files,
        self_gstin=self_gstin, company_name=company_name,
    )


if __name__ == "__main__":
    import sys
    res = classify_folder(sys.argv[1] if len(sys.argv) > 1 else ".")
    print("Self GSTIN:", res["self_gstin"], "|", res["company_name"])
    print("\nGSTR-1 files:", res["gstr1_files"], "-> months:", sorted(res["gstr1_months"]))
    print("GSTR-3B files:", res["gstr3b_files"], "-> months:", sorted(res["gstr3b_months"]))
    print("E-Invoice files:", res["einv_files"], "-> months:", sorted(res["einv_months"]))
    print("GSTR-2B files:", res["gstr2b_files"], "-> months:", sorted(res["gstr2b_months"]))
    if res["classify_warnings"]:
        print("\nWARNINGS:")
        for w in res["classify_warnings"]:
            print(" -", w)
    print("\nEWB Out files:", res["ewb_out_files"])
    print("EWB In files:", res["ewb_in_files"])
    print("Cash/Credit/Liability ledgers:", res["cash_ledgers"], res["credit_ledgers"], res["liab_ledgers"])
    print("TPST files:", res["tpst_files"])
    print("Portal comparison files:", res["portal_comparison_files"])
    print("BO profile files:", res["bo_profile_files"])
    print("GSTR-9 files:", res["gstr9_files"])
    print("GSTR-9C files:", res["gstr9c_files"])
    print("Table 8A files:", res["table8a_files"])
    print("BS/P&L files:", res["bs_pl_files"])
