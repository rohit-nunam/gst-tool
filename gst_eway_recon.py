#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST SCRUTINY  --  E-WAY BILL RECONCILIATION LAYER  (27-check matrix)
====================================================================
Sixth-source layer: brings E-Way Bill OUTWARD and INWARD into the
GSTR-1 / 2B / 3B / E-Invoice reconciliation.

Reuses CONFIG + parsers from gst_scrutiny_tool.py.

Honest scope (decided with the user):
  BUILD   : #1-9, #15-18, #23, #25-27   (full, file-driven)
  PARTIAL : #10-13 (2B is a PDF summary -> aggregate only, no line list)
            #22 (no validity/expiry col), #24 (no cancel-status col)
  SKIP    : #14 (no books), #19/#20 (no purchase e-invoice / books),
            #21 (no filing date)  -- #7/#16 still run on dates inside the files

NO safety nets, NO invented data. Every PARTIAL/SKIP is labelled, not faked.

USAGE:
    Put next to gst_scrutiny_tool.py + gst_analysis_checks.py, set the two
    EWB filenames in CONFIG below, then:
        python gst_eway_recon.py
    -> GST_Scrutiny_EWayBill.xlsx
"""

import os, re, datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import gst_scrutiny_tool as raw
import gstr2b_parser as g2b
import merged_period_utils as mpu

num = raw.num
TOL = 1.0                 # rupee tolerance for value matches
VALUE_TOL_PCT = 0.01      # 1% tolerance for EWB-vs-return value compares

# ----------------------------------------------------------------------
# CONFIG  --  set by run_monthly_pipeline.py per month (merged-file model)
# ----------------------------------------------------------------------
GSTR2B_FILE  = None                                   # merged GSTR-2B workbook path
SELF_GSTIN   = ""
COMPANY_NAME = ""
EWB_THRESHOLD = 50000.0   # Rule 138 inter-state EWB threshold (consignment value)

OUTPUT_FILE = "GST_Scrutiny_EWayBill.xlsx"
_LAST_2B_SRC = "pdf"    # set in main(): 'excel' enables line-level checks
_LAST_2B_FILE = None    # actual 2B Excel filename used (for the header stamp)
_LAST_EINV_FILE = None  # actual e-invoice filename used


def find_file(configured, patterns, search_dir=".", exclude=("SCRUTINY", "EWAYBILL", "EWAY", "COMPARISON", "ANALYSIS")):
    """Return a usable file path. 1) use `configured` if it exists; 2) else scan search_dir for
    a file whose name matches any regex in `patterns` (case-insensitive) and isn't one of our outputs."""
    import glob as _glob, re as _re
    if configured and os.path.exists(configured):
        return configured
    cands = []
    for f in _glob.glob(os.path.join(search_dir, "*.xlsx")) + _glob.glob(os.path.join(search_dir, "*.xlsm")):
        name = os.path.basename(f).upper()
        if any(x in name for x in exclude):
            continue
        if any(_re.search(p, name, _re.I) for p in patterns):
            cands.append(f)
    if not cands:
        return None
    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cands[0]


# ----------------------------------------------------------------------
# State code <-> name (needed: GSTR-1 POS is a name, EWB carries GSTIN codes)
# ----------------------------------------------------------------------
STATE = {
 "01":"Jammu and Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
 "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
 "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
 "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
 "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh","24":"Gujarat",
 "25":"Daman and Diu","26":"Dadra and Nagar Haveli","27":"Maharashtra","28":"Andhra Pradesh",
 "29":"Karnataka","30":"Goa","31":"Lakshadweep","32":"Kerala","33":"Tamil Nadu",
 "34":"Puducherry","35":"Andaman and Nicobar Islands","36":"Telangana","37":"Andhra Pradesh",
 "38":"Ladakh","97":"Other Territory",
}
NAME2CODE = {v.lower(): k for k, v in STATE.items()}
NAME2CODE["andhra pradesh"] = "37"

def state_code(s):
    """Pull a 2-digit state code from a GSTIN or a leading-coded string."""
    m = re.match(r"\s*(\d{2})", str(s or ""))
    return m.group(1) if m else None

def name_to_code(name):
    return NAME2CODE.get(str(name or "").strip().lower())


# ----------------------------------------------------------------------
# Severity model (shared style with analysis layer)
# ----------------------------------------------------------------------
FLAG, REVW, INFO, PASS, SKIP = "FLAG", "REVIEW", "INFO", "PASS", "SKIPPED"
SEV_ORDER = {FLAG: 0, REVW: 1, INFO: 2, PASS: 3, SKIP: 4}

class F:
    __slots__ = ("ref", "title", "sev", "detail", "rows")
    def __init__(self, ref, title, sev, detail, rows=None):
        self.ref, self.title, self.sev, self.detail, self.rows = ref, title, sev, detail, rows or []


# ----------------------------------------------------------------------
# EWB parser
# ----------------------------------------------------------------------
def _split_doc(v):
    """'MR22-23/0519 - 10/01/2023' -> ('MR22-23/0519', date(2023,1,10))."""
    s = str(v or "").strip()
    if not s:
        return "", None
    parts = re.split(r"\s+-\s+", s, maxsplit=1)
    docno = parts[0].strip()
    dt = None
    if len(parts) > 1:
        dt = _parse_dt(parts[1])
    return docno, dt

def _split_ewb(v):
    """'301546430758 - 10/01/2023 17:01:00' -> ('301546430758', datetime)."""
    s = str(v or "").strip()
    if not s:
        return "", None
    parts = re.split(r"\s+-\s+", s, maxsplit=1)
    ewbno = parts[0].strip()
    dt = _parse_dt(parts[1]) if len(parts) > 1 else None
    return ewbno, dt

def _parse_dt(s):
    s = str(s or "").strip()
    if not s:
        return None
    s = s.split()[0]  # drop time part
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def _gstin_of(combined):
    """'05AAECM6380J1ZA / M.R...' -> '05AAECM6380J1ZA'."""
    return str(combined or "").split("/")[0].strip()

def parse_ewb(path):
    """Return list of dicts for one EWB file (outward or inward)."""
    out = []
    if not path or not os.path.exists(path):
        return out
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    H = {h: i for i, h in enumerate(hdr)}
    def g(r, *names):
        for n in names:
            if n in H and H[n] < len(r):
                return r[H[n]]
        return None
    for r in rows[1:]:
        if not any(r):
            continue
        docno, docdt = _split_doc(g(r, "Doc No. & Dt."))
        ewbno, ewbdt = _split_ewb(g(r, "EWB No. & Dt."))
        out.append(dict(
            ewbno=ewbno or str(g(r, "EWB No.") or "").strip(),
            ewbdate=ewbdt,
            docno=docno, docdate=docdt,
            from_gstin=_gstin_of(g(r, "From GSTIN & Name")),
            to_gstin=_gstin_of(g(r, "To GSTIN & Name")),
            assess=num(g(r, "Assess Val.")),
            taxval=num(g(r, "Tax Val.")),
            hsn=str(g(r, "HSN Code") or "").strip(),
            vehicle=str(g(r, "Latest Vehicle No.") or "").strip(),
        ))
    return out


# ----------------------------------------------------------------------
# Pull GSTR-1 & e-invoice line detail (reuse analysis layer's reader if present)
# ----------------------------------------------------------------------
def read_gstr1_invoices(path, month):
    """invno -> dict(taxable, igst, cgst, sgst, pos, gstin, rate_lines, consignment).
    Scoped to ONE month's block out of the merged GSTR-1 workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["b2b, sez, de_inv"]; rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate([str(c).strip() if c else "" for c in rows[3]])}
    def g(r, k): return r[H[k]] if k in H and H[k] < len(r) else None
    inv = {}
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        no = str(g(r, "Invoice Number") or "").strip()
        d = inv.setdefault(no, dict(taxable=0.0, igst=0.0, cgst=0.0, sgst=0.0,
                                    invval=0.0, pos=str(g(r, "Place Of Supply") or "").strip(),
                                    gstin=str(g(r, "GSTIN/UIN of Recipient") or "").strip(),
                                    rates=set()))
        d["taxable"] += num(g(r, "Taxable Value"))
        d["igst"] += num(g(r, "Integrated Tax"))
        d["cgst"] += num(g(r, "Central Tax"))
        d["sgst"] += num(g(r, "State/UT Tax"))
        d["invval"] += num(g(r, "Invoice Value"))
        d["rates"].add(num(g(r, "Rate")))
    return inv

def read_einv_invoices(path, month):
    out = {}
    if not path or not os.path.exists(path):
        return out   # E-Invoice legitimately not supplied at all -- graceful
    wb = openpyxl.load_workbook(path, data_only=True)
    if "b2b, sez, de" not in wb.sheetnames:
        return out
    ws = wb["b2b, sez, de"]; rows = list(ws.iter_rows(values_only=True))
    if month not in mpu.months_present(rows, 3):
        return out   # E-Invoice doesn't cover this month -- same graceful state
    H = {h: i for i, h in enumerate([str(c).strip() if c else "" for c in rows[3]])}
    invcol = "Invoice number" if "Invoice number" in H else "Invoice Number"
    def g(r, k): return r[H[k]] if k in H and H[k] < len(r) else None
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        no = str(g(r, invcol) or "").strip()
        d = out.setdefault(no, dict(taxable=0.0, igst=0.0, cgst=0.0, sgst=0.0, invval=0.0))
        d["taxable"] += num(g(r, "Taxable Value"))
        d["igst"] += num(g(r, "Integrated Tax"))
        d["cgst"] += num(g(r, "Central Tax"))
        d["sgst"] += num(g(r, "State/UT Tax"))
        d["invval"] += num(g(r, "Invoice Value"))
    return out


def _val_mismatch(a, b):
    if max(abs(a), abs(b)) == 0:
        return False
    return abs(a - b) > max(TOL, VALUE_TOL_PCT * max(abs(a), abs(b)))


# ----------------------------------------------------------------------
# THE 27 CHECKS
# ----------------------------------------------------------------------
def run(ewb_out, ewb_in, g1inv, einv, g3b, b2b, ewb_out_file_supplied=True, ewb_in_file_supplied=True):
    """ewb_out_file_supplied / ewb_in_file_supplied: whether the ANNUAL EWB
    workbook for that direction was supplied AT ALL this run (independent of
    whether this particular month happens to have zero rows in it, which is
    a legitimate business state, not a data gap). Defaults to True so any
    existing caller that doesn't pass these keeps its old behaviour.

    GRACEFUL DEGRADATION (fixed): previously, when an entire EWB direction
    was never supplied for this taxpayer (common for smaller taxpayers below
    the Rule-138 threshold, or service businesses with no goods movement),
    `ewb_out`/`ewb_in` arrived here as an empty list indistinguishable from
    "file present, zero rows this month" -- and several checks silently
    produced a MISLEADING result instead of an honest 'no data' state:
      #1  showed PASS ("0/0 EWB-Out doc-numbers found in GSTR-1") -- looked
          like a clean reconciliation; it was actually zero data to check.
      #4  flagged EVERY inter-state >Rs50k invoice as REVIEW ("no EWB found"),
          since the empty ewbset made every such invoice look EWB-less --
          flooding the dashboard with noise that isn't a real finding.
      #10 showed REVIEW ("0 of 0 inward-EWB documents matched") -- a false
          alarm in the opposite direction from #1's false PASS.
      #12 showed PASS ("every inward EWB matches a 2B invoice") -- true only
          because there were zero inward EWBs to fail to match.
    All four (plus #3, #13) now check the file-supplied flag FIRST and emit
    an explicit SKIP with the reason, before running their normal logic."""
    R = []
    def gv(k, i, d=0.0):
        v = g3b.get(k); return v[i] if v and i < len(v) else d

    b2b_available = b2b.get("available", True)

    # Computed unconditionally (safe: sums to 0 on an empty list) so later
    # sections that reference these can do so regardless of which branch
    # ran above -- only the FINDINGS that report them are gated behind
    # ewb_out_file_supplied, not the sums themselves.
    ewb_out_total = sum(e["assess"] for e in ewb_out)
    ewb_out_tax = sum(e["taxval"] for e in ewb_out)

    # EWB outward keyed by doc number (invoice number)
    out_by_doc = {}
    for e in ewb_out:
        out_by_doc.setdefault(e["docno"], []).append(e)
    in_by_doc = {}
    for e in ewb_in:
        in_by_doc.setdefault(e["docno"], []).append(e)

    if not ewb_out_file_supplied:
        for ref, title in [("#1", "EWB-Out invoice present in GSTR-1"), ("#3", "EWB-Out with NO matching GSTR-1 invoice"),
                            ("#4", "GSTR-1 inter-state >\u20b950k with NO EWB-Out")]:
            R.append(F(ref, title, SKIP,
                       "No outward EWB workbook was supplied for this taxpayer/FY at all -- this is a "
                       "data-availability gap (common for taxpayers under the Rule-138 threshold or with "
                       "no goods movement), not evidence of a missing EWB on any specific invoice. "
                       "Skipped rather than shown as PASS or REVIEW to avoid a false-clean or "
                       "false-flagged result."))
    if not ewb_in_file_supplied:
        for ref, title in [("#10", "EWB-In invoice matched to GSTR-2B"), ("#12", "EWB-In with NO matching GSTR-2B invoice"),
                            ("#13", "GSTR-2B inter-state >\u20b950k with NO inward EWB")]:
            R.append(F(ref, title, SKIP,
                       "No inward EWB workbook was supplied for this taxpayer/FY at all -- data-availability "
                       "gap, not a finding. Skipped rather than shown as PASS or REVIEW."))

    # ===== A. EWB-Out vs GSTR-1 =====
    if ewb_out_file_supplied:
        # #1 invoice matching
        g1set = set(g1inv) - {"None", ""}
        ewbset = set(out_by_doc) - {""}
        matched = g1set & ewbset
        unmatched = sorted(ewbset - g1set)
        R.append(F("#1", "EWB-Out invoice present in GSTR-1", PASS if not unmatched else REVW,
                   f"{len(matched)}/{len(ewbset)} EWB-Out doc-numbers found in GSTR-1 "
                   f"(GSTR-1 invoices: {len(g1set)}). {len(unmatched)} EWB doc(s) not in GSTR-1 — "
                   "classified by tax in #3 (zero-tax movements expected; tax-bearing = flag).",
                   [("EWB doc not in GSTR-1",)] + [(d,) for d in unmatched]))

        # #2 value mismatch (EWB consignment 'assess' is post-discount taxable; compare to GSTR-1 taxable)
        vm = []
        for doc in sorted(matched):
            ewb_assess = sum(x["assess"] for x in out_by_doc[doc])
            g1_tax = g1inv[doc]["taxable"]
            if _val_mismatch(ewb_assess, g1_tax):
                vm.append((doc, ewb_assess, g1_tax, round(ewb_assess - g1_tax, 2)))
        R.append(F("#2", "EWB-Out value vs GSTR-1 taxable", PASS if not vm else REVW,
                   f"{len(vm)} invoice(s) with >1% value gap (EWB assessable can differ from GSTR-1 "
                   "taxable due to freight/discount; review, not auto-flag).",
                   [("Invoice", "EWB assess", "GSTR-1 taxable", "diff")] + vm))

        # #3 EWB-Out exists but NO GSTR-1.
        # Three buckets, not one:
        #   (a) tax-bearing AND matches a 2B supplier credit note (GSTIN+value) -> PURCHASE RETURN,
        #       correctly absent from GSTR-1 outward (it's a debit-note-out against a purchase). Expected.
        #   (b) tax-bearing, NO 2B match -> genuine concern (suppressed sale / booked elsewhere). FLAG.
        #   (c) zero-tax -> approval / stock-transfer / job-work challan. Expected.
        no_g1 = sorted(ewbset - g1set)
        # build 2B credit-note index by (supplier_gstin, taxable, tax) when 2B Excel present
        cdnr_idx = {}
        twob_lines = b2b.get("_lines") if b2b_available else None
        if twob_lines:
            for c in twob_lines["cdnr"]:
                ct = c["igst"] + c["cgst"] + c["sgst"]
                cdnr_idx.setdefault((state_code(c["gstin"]) or c["gstin"][:2],
                                     round(c["taxable"], 2), round(ct, 2)), []).append(c)
        pur_return, taxed, zero = [], [], []
        for d in no_g1:
            t = sum(x["taxval"] for x in out_by_doc[d])
            a = sum(x["assess"] for x in out_by_doc[d])
            to = out_by_doc[d][0]["to_gstin"]
            if t <= TOL:
                zero.append((d, to, round(a, 2), round(t, 2), "zero-tax movement (approval/stock-transfer)"))
                continue
            key = (state_code(to) or to[:2], round(a, 2), round(t, 2))
            match = cdnr_idx.get(key)
            if match:
                note = match[0]["note"]
                pur_return.append((d, to, round(a, 2), round(t, 2), f"matches 2B credit note {note} (purchase return)"))
            else:
                taxed.append((d, to, round(a, 2), round(t, 2), "tax-bearing, NO 2B credit-note match — investigate"))
        sev3 = FLAG if taxed else (INFO if (pur_return or zero) else PASS)
        detail3 = []
        if taxed:
            detail3.append(f"{len(taxed)} TAX-BEARING outward EWB(s) with NO GSTR-1 invoice AND no matching 2B "
                           "credit note -> goods moved with tax but not in GSTR-1; investigate. ")
        if pur_return:
            detail3.append(f"{len(pur_return)} outward EWB(s) RECONCILED as PURCHASE RETURNS — each matches a "
                           "supplier credit note in GSTR-2B (GSTIN + taxable + tax), so the goods went back to the "
                           "supplier and correctly do NOT appear as a GSTR-1 outward sale. No action. ")
        if zero:
            detail3.append(f"{len(zero)} ZERO-TAX movement(s) (approval / stock-transfer / job-work challan) "
                           "correctly absent from GSTR-1; verify invoiced on sale/return. ")
        if not no_g1:
            detail3 = ["Every outward EWB has a GSTR-1 invoice."]
        R.append(F("#3", "EWB-Out with NO matching GSTR-1 invoice", sev3, "".join(detail3),
                   [("EWB doc-no", "To GSTIN", "assess", "tax", "classification")] + taxed + pur_return + zero))

        # #4 GSTR-1 inter-state >50k with NO EWB-Out.
        # NOT an automatic Rule 138 violation: EWB is required only if goods physically move.
        # Job-work (JWI) / pure-service / delivery-challan cases may legitimately have no EWB.
        # Classify by document-number prefix and route to REVIEW with the right question, not FLAG.
        miss_ewb = []
        own = SELF_GSTIN[:2]
        for no, d in g1inv.items():
            if no in ("None", ""):
                continue
            pos_code = state_code(d["pos"]) or name_to_code(d["pos"])
            inter = pos_code and pos_code != own
            consign = d["taxable"] + d["igst"] + d["cgst"] + d["sgst"]
            if inter and consign > EWB_THRESHOLD and no not in ewbset:
                pfx = re.match(r"([A-Za-z]+)", no)
                pfx = pfx.group(1) if pfx else ""
                jobwork = pfx.upper() in ("JWI", "JW")
                miss_ewb.append((no, d["pos"], round(consign, 2),
                                 "job-work invoice — verify goods moved" if jobwork
                                 else "verify goods movement"))
        # severity: REVIEW (a missing EWB on an inter-state >50k supply is a question, not a proven breach)
        sev4 = REVW if miss_ewb else PASS
        R.append(F("#4", "GSTR-1 inter-state >₹50k with NO EWB-Out", sev4,
                   ("EWB is mandatory only where goods physically move. These inter-state >₹50k supplies have "
                    "no outward EWB — Rule 138 applies ONLY if goods actually moved. Job-work (JWI) / pure-service "
                    "/ delivery-challan movements may legitimately have no EWB under this GSTIN. VERIFY physical "
                    "movement per invoice before treating as a violation; this is a review item, not a proven breach.")
                   if miss_ewb else "All inter-state >₹50k GSTR-1 invoices have an EWB.",
                   [("Invoice", "POS", "consignment ₹", "action")] + miss_ewb))
    else:
        # SKIP findings for #1/#3/#4 already appended above; ewbset must still exist
        # (as an empty set) for later sections (#5-9, #15-18) that reference it.
        ewbset = set()
        matched = set()



    # ===== B. EWB-Out vs E-Invoice =====
    if einv:
        einvset = set(einv) - {"None", ""}
        # #5 match
        em = ewbset & einvset
        R.append(F("#5", "EWB-Out invoice present in E-Invoice", PASS if ewbset <= einvset else REVW,
                   f"{len(em)}/{len(ewbset)} EWB-Out doc-numbers found in e-invoice. "
                   "(B2C / sub-threshold EWBs may legitimately have no e-invoice.)",
                   [("EWB doc not in e-inv",)] + [(d,) for d in sorted(ewbset - einvset)]))
        # #6 value mismatch
        vm6 = []
        for doc in sorted(em):
            ea = sum(x["assess"] for x in out_by_doc[doc]); et = einv[doc]["taxable"]
            if _val_mismatch(ea, et):
                vm6.append((doc, ea, et, round(ea - et, 2)))
        R.append(F("#6", "EWB-Out value vs E-Invoice taxable", PASS if not vm6 else REVW,
                   f"{len(vm6)} invoice(s) with >1% gap.",
                   [("Invoice", "EWB assess", "E-inv taxable", "diff")] + vm6))
        # #7 EWB date vs e-invoice... e-invoice file has no per-invoice date kept here -> use doc date vs ewb date
        gap7 = []
        for doc in sorted(em):
            for x in out_by_doc[doc]:
                if x["ewbdate"] and x["docdate"]:
                    g = (x["ewbdate"] - x["docdate"]).days
                    if g > 1:
                        gap7.append((doc, x["docdate"], x["ewbdate"], g))
        R.append(F("#7", "EWB-date vs invoice(doc)-date gap (>1 day)",
                   PASS if not gap7 else REVW,
                   "EWB generated well after invoice date -> delayed generation; verify movement timing. "
                   "(Return-filing-date checks #10/#21 skipped: no filing date supplied.)",
                   [("Invoice", "Doc date", "EWB date", "gap days")] + gap7))
    else:
        for ref, t in [("#5", "EWB-Out vs E-Invoice match"), ("#6", "EWB-Out vs E-Invoice value"),
                       ("#7", "EWB-date vs invoice-date gap")]:
            R.append(F(ref, t, INFO, "E-invoice file not supplied."))

    # ===== C. EWB-Out vs GSTR-3B =====
    # #8 aggregate outward EWB vs 3B 3.1(a): BOTH assessable and tax.
    if ewb_out_file_supplied:
        b3b_tax = gv("3.1a", 0)                                  # taxable value
        b3b_outtax = gv("3.1a", 1) + gv("3.1a", 2) + gv("3.1a", 3)   # IGST+CGST+SGST
        ratio8 = ewb_out_total / b3b_tax if b3b_tax else 0
        tax_gap = ewb_out_tax - b3b_outtax
        sev8 = REVW if tax_gap > TOL else INFO
        R.append(F("#8", "EWB-Out aggregate vs GSTR-3B 3.1(a)", sev8,
                   f"Assessable: EWB-Out {ewb_out_total:,.2f} vs 3B 3.1(a) taxable {b3b_tax:,.2f} "
                   f"(ratio {ratio8:.2f}). TAX: EWB-Out tax {ewb_out_tax:,.2f} vs 3B 3.1(a) output tax "
                   f"{b3b_outtax:,.2f} -> EWB tax higher by {abs(tax_gap):,.2f}. "
                   + ("NOT A NEW FINDING: the outward EWB carries more tax than 3B output tax because some outward "
                      "EWBs are PURCHASE RETURNS (goods sent back to suppliers, matching supplier credit notes in "
                      "2B — see #3), not outward sales. Those returns inflate outward EWB tax but correctly never "
                      "hit 3B output liability. The gap is explained by #3's reclassification, not a separate "
                      "suppression item. " if tax_gap > TOL else "")
                   + "EWB covers only goods-movement supplies (not services / B2C sub-threshold), so EWB < 3B "
                     "is normal; a true EWB-tax > 3B-output-tax gap that ISN'T purchase-returns is the signal.",
                   []))

        # #9 tax type consistency: EWB inter/intra (from->to state) vs GSTR-1 head
        tt = []
        for doc in sorted(matched):
            e = out_by_doc[doc][0]
            fr, to = state_code(e["from_gstin"]), state_code(e["to_gstin"])
            if not (fr and to):
                continue
            inter = fr != to
            d = g1inv[doc]
            has_igst = d["igst"] > TOL
            has_local = d["cgst"] > TOL or d["sgst"] > TOL
            if inter and has_local and not has_igst:
                tt.append((doc, f"{fr}->{to} inter", "GSTR-1 has CGST/SGST"))
            if (not inter) and has_igst and not has_local:
                tt.append((doc, f"{fr}->{to} intra", "GSTR-1 has IGST"))
        R.append(F("#9", "Tax-type (inter/intra) EWB vs GSTR-1 head", PASS if not tt else FLAG,
                   "EWB movement direction contradicts the tax head charged in GSTR-1."
                   if tt else "EWB direction matches GSTR-1 tax head on all matched invoices.",
                   [("Invoice", "EWB direction", "GSTR-1")] + tt))
    else:
        R.append(F("#8", "EWB-Out aggregate vs GSTR-3B 3.1(a)", SKIP,
                   "No outward EWB workbook supplied -- cannot aggregate."))
        R.append(F("#9", "Tax-type (inter/intra) EWB vs GSTR-1 head", SKIP,
                   "No outward EWB workbook supplied -- cannot test."))

    # ===== D. EWB-In vs GSTR-2B =====
    ewb_in_assess = sum(e["assess"] for e in ewb_in)
    ewb_in_tax = sum(e["taxval"] for e in ewb_in)
    b2b_itc = ((b2b["ITC_all_other_IGST"] + b2b["ITC_all_other_CGST"] + b2b["ITC_all_other_SGST"])
               if b2b_available else None)
    twob_lines = b2b.get("_lines") if b2b_available else None   # set when 2B came from Excel

    if not ewb_in_file_supplied:
        pass  # SKIP findings for #10/#12/#13 already appended at top; #11 handled just below.
    elif twob_lines:
        # ---- LINE-LEVEL (2B Excel invoice list available) ----
        b2b_inv = twob_lines["b2b"]
        def nkey(g, n): return (str(g).strip().upper(), str(n).strip().upper())
        # primary index: (supplier, invoice-no).  value index: (supplier_state, taxable, tax)
        b2b_map, b2b_val = {}, {}
        for x in b2b_inv:
            b2b_map.setdefault(nkey(x["gstin"], x["invno"]), []).append(x)
            vk = (str(x["gstin"]).strip().upper(), round(x["taxable"], 2),
                  round(x["igst"]+x["cgst"]+x["sgst"], 2))
            b2b_val.setdefault(vk, []).append(x)
        ewbin_map = {}
        for e in ewb_in:
            ewbin_map.setdefault(nkey(e["from_gstin"], e["docno"]), []).append(e)

        def matches_2b(k):
            """True if this EWB key matches a 2B invoice by number OR (same supplier + value).
            The value fallback prevents false 'not filed' flags when only the invoice-number
            FORMAT differs (e.g. EWB '631' vs 2B 'HFPL/631', or '002037' vs '2037')."""
            if k in b2b_map:
                return ("invoice-no", b2b_map[k])
            es = ewbin_map[k]
            ea = round(sum(x["assess"] for x in es), 2); et = round(sum(x["taxval"] for x in es), 2)
            vk = (k[0], ea, et)
            if vk in b2b_val:
                return ("value", b2b_val[vk])
            return (None, None)

        matched_in, by_value = {}, 0
        for k in ewbin_map:
            how, hit = matches_2b(k)
            if how:
                matched_in[k] = (how, hit)
                if how == "value":
                    by_value += 1

        # #10 invoice matching
        R.append(F("#10", "EWB-In invoice matched to GSTR-2B (no. or value)",
                   PASS if matched_in else REVW,
                   f"{len(matched_in)} of {len(ewbin_map)} inward-EWB documents matched a 2B invoice "
                   f"(by invoice-number, or by supplier+value when the number format differed — "
                   f"{by_value} matched on value). 2B B2B invoices: {len(b2b_map)}. Unmatched in #12.",
                   []))

        # #11 value match on matched-by-number set (value-matched are equal by construction)
        vmism = []
        for k, (how, hit) in matched_in.items():
            if how != "invoice-no":
                continue
            ea = sum(x["assess"] for x in ewbin_map[k]); ba = sum(x["taxable"] for x in hit)
            if _val_mismatch(ea, ba):
                vmism.append((k[1], k[0], round(ea, 2), round(ba, 2), round(ea-ba, 2)))
        R.append(F("#11", "EWB-In vs GSTR-2B value (matched invoices)",
                   PASS if not vmism else REVW,
                   f"{len(vmism)} matched invoice(s) with >1% taxable gap (under/over-invoicing signal). "
                   "EWB assessable can differ from 2B taxable for freight/discount; review." if vmism
                   else "All number-matched inward-EWB/2B invoices agree on value within tolerance.",
                   [("Invoice", "Supplier", "EWB assess", "2B taxable", "diff")] + vmism))

        # #12 EWB-In with NO 2B match (neither number nor value) -> supplier hasn't filed
        only_ewb = sorted(k for k in ewbin_map if k not in matched_in)
        rows12 = []
        for k in only_ewb:
            rows12.append((k[1], k[0], round(sum(x["assess"] for x in ewbin_map[k]), 2),
                           round(sum(x["taxval"] for x in ewbin_map[k]), 2)))
        R.append(F("#12", "EWB-In with NO matching GSTR-2B invoice",
                   PASS if not only_ewb else FLAG,
                   f"{len(only_ewb)} inward EWB(s) with no 2B match by invoice-number OR by supplier+value -> "
                   "supplier hasn't filed (or filed under a materially different invoice); ITC not yet "
                   "available — chase these suppliers." if only_ewb
                   else "Every inward EWB matches a 2B invoice.",
                   [("Doc no", "Supplier GSTIN", "EWB assess", "EWB tax")] + rows12))

        # #13 2B invoice >50k inter-state with NO inward EWB
        own = SELF_GSTIN[:2]
        matched_2b_ids = set()
        for k, (how, hit) in matched_in.items():
            for x in hit:
                matched_2b_ids.add((x["gstin"].upper(), x["invno"].upper()))
        only_2b = []
        seen2b = {}
        for x in b2b_inv:
            seen2b.setdefault((x["gstin"].upper(), x["invno"].upper()), []).append(x)
        for key, xs in seen2b.items():
            if key in matched_2b_ids:
                continue
            sup_state = state_code(xs[0]["gstin"])
            taxable = sum(x["taxable"] for x in xs)
            consign = taxable + sum(x["igst"]+x["cgst"]+x["sgst"] for x in xs)
            inter = sup_state and sup_state != own
            if inter and consign > EWB_THRESHOLD:
                only_2b.append((xs[0]["invno"], xs[0]["gstin"], round(consign, 2)))
        R.append(F("#13", "GSTR-2B inter-state >₹50k with NO inward EWB",
                   PASS if not only_2b else REVW,
                   f"{len(only_2b)} inter-state 2B invoice(s) >₹50k with no matching inward EWB -> supplier "
                   "may not have generated an EWB (their Rule 138 issue) or goods moved on a challan. Verify; "
                   "affects your defensibility if questioned." if only_2b
                   else "All inter-state >₹50k 2B invoices have a matching inward EWB.",
                   [("Invoice", "Supplier GSTIN", "consignment ₹")] + only_2b))
    else:
        # ---- AGGREGATE ONLY: either 2B wasn't supplied at all, or it was supplied but only
        # as a PDF summary (no line-level invoice list) -- these are different situations and
        # get different messages rather than one generic "insufficient" note. ----
        if not b2b_available:
            reason = (f"GSTR-2B not supplied for this month ({b2b.get('_reason', 'no reason recorded')})"
                      " -- line-level matching not possible.")
        else:
            reason = ("GSTR-2B supplied as PDF summary (no invoice list) -> line-level matching not possible. "
                      "Supply the GSTR-2B Excel download to enable #10-#13.")
        R.append(F("#10", "EWB-In invoice match GSTR-2B (line-level)", INFO if b2b_available else SKIP,
                   reason + " Aggregate compare in #11/#26.", []))
        if b2b_itc is not None:
            R.append(F("#11", "EWB-In aggregate vs GSTR-2B ITC", INFO,
                       f"Inward EWB assessable {ewb_in_assess:,.2f}, EWB tax {ewb_in_tax:,.2f}. "
                       f"2B 'all other ITC' tax {b2b_itc:,.2f}. CAUTION — DIFFERENT BASES, NOT A LIKE-FOR-LIKE "
                       "RATIO: EWB-In is goods-only, 2B ITC includes services+goods+RCM. Scale context only.", []))
        else:
            R.append(F("#11", "EWB-In aggregate vs GSTR-2B ITC", SKIP,
                       f"GSTR-2B not supplied -- cannot compare. Inward EWB assessable {ewb_in_assess:,.2f}, "
                       f"EWB tax {ewb_in_tax:,.2f} shown for reference only.", []))
        R.append(F("#12", "EWB-In exists but no 2B entry", INFO if b2b_available else SKIP,
                   "Needs 2B invoice list (PDF summary insufficient)." if b2b_available
                   else "GSTR-2B not supplied.", []))
        R.append(F("#13", "2B entry but no EWB-In (>50k inter)", INFO if b2b_available else SKIP,
                   "Needs 2B invoice list (PDF summary insufficient)." if b2b_available
                   else "GSTR-2B not supplied.", []))

    # ===== E. skipped (no books) =====
    R.append(F("#14", "Unaccounted purchases (EWB-In vs books)", SKIP, "No purchase register supplied."))

    # ===== F. EWB-Out vs EWB-In (same transaction) =====
    # For a single GSTIN, outward and inward EWBs are different counterparties; overlap only if
    # the same doc-no appears on both sides (rare). Report overlaps, else N/A.
    overlap = set(out_by_doc) & set(in_by_doc) - {""}
    R.append(F("#15", "EWB-Out vs EWB-In value (same doc-no)",
               PASS if not overlap else REVW,
               "No document number appears on both outward and inward EWB (expected for a single GSTIN: "
               "your outward = others' inward, not in your own download)."
               if not overlap else f"{len(overlap)} doc-no on both sides; verify.",
               [("doc-no",)] + [(d,) for d in sorted(overlap)]))

    # #16 EWB time gap generation vs document date (both sides)
    gap16 = []
    for tag, lst in (("OUT", ewb_out), ("IN", ewb_in)):
        for e in lst:
            if e["ewbdate"] and e["docdate"]:
                g = (e["ewbdate"] - e["docdate"]).days
                if abs(g) > 2:
                    gap16.append((tag, e["docno"], e["docdate"], e["ewbdate"], g))
    R.append(F("#16", "EWB generation vs document date gap (>2 days)",
               PASS if not gap16 else REVW,
               "Large gap between document date and EWB generation -> verify genuine movement timing."
               if gap16 else "All EWBs generated within 2 days of document date.",
               [("Side", "doc-no", "doc date", "EWB date", "gap")] + gap16[:50]))

    # ===== G. Triangulation GSTR-1 + E-Inv + EWB-Out =====
    if einv:
        tri = []
        allinv = (set(g1inv) | set(einv) | ewbset) - {"None", ""}
        for no in sorted(allinv):
            in_g1 = no in g1inv; in_ei = no in einv; in_ew = no in ewbset
            if not (in_g1 and in_ei and in_ew):
                tri.append((no, "Y" if in_g1 else "-", "Y" if in_ei else "-", "Y" if in_ew else "-"))
        R.append(F("#17", "Triangulation: GSTR-1 / E-Invoice / EWB-Out", PASS if not tri else REVW,
                   "Invoices not present in all three sources. EWB-absent (goods) can be legitimate for "
                   "services/B2C; GSTR-1-absent but EWB-present = suppression signal (see #3)."
                   if tri else "Every invoice appears in all three sources.",
                   [("Invoice", "GSTR-1", "E-Inv", "EWB-Out")] + tri[:80]))
    else:
        R.append(F("#17", "Triangulation GSTR-1/E-Inv/EWB-Out", INFO, "E-invoice file not supplied."))

    # #18 HSN rate-wise across sources -- EWB has HSN code, GSTR-1 HSN sheet has rate-wise.
    hsn_out = {}
    for e in ewb_out:
        h = e["hsn"][:4] if e["hsn"] else "?"
        hsn_out[h] = hsn_out.get(h, 0.0) + e["assess"]
    R.append(F("#18", "HSN rate-wise across sources", INFO,
               "EWB-Out HSN-wise assessable (4-digit): "
               + "; ".join(f"{k}:{v:,.0f}" for k, v in sorted(hsn_out.items()))
               + ". Cross-check against GSTR-1 HSN summary rate rows for misclassification "
               "(HSN-summary rate split must be read from the GSTR-1 'hsn' sheet).",
               []))

    # ===== H. skipped =====
    R.append(F("#19", "3B ITC (4A5) vs EWB-In aggregate", SKIP,
               "Skipped per scope (purchase/books side). Aggregate context shown in #11/#26."))
    R.append(F("#20", "E-Invoice (purchase) vs EWB-In", SKIP, "No purchase-side e-invoice data."))

    # ===== I. timing =====
    R.append(F("#21", "EWB gen date vs GSTR-1 filing date", SKIP, "No GSTR-1 filing date supplied."))
    R.append(F("#22", "EWB validity expiry vs supply date", INFO,
               "EWB export has no validity/expiry column -> cannot test. Download detailed EWB with "
               "'Valid Upto' to enable."))

    # #23 multiple EWBs per invoice (partial dispatch / threshold-splitting)
    multi = [(doc, len(lst), round(sum(x["assess"] for x in lst), 2))
             for doc, lst in out_by_doc.items() if doc and len(lst) > 1]
    R.append(F("#23", "Multiple EWBs per invoice (partial dispatch)",
               PASS if not multi else REVW,
               "Allowed for partial dispatch, but multiple sub-threshold EWBs on one invoice can be "
               "used to dodge e-invoice/EWB limits -> verify." if multi else
               "No invoice has more than one outward EWB.",
               [("Invoice", "#EWBs", "total assess")] + multi))

    R.append(F("#24", "EWB cancelled after return filing", INFO,
               "EWB export has no cancellation-status column -> cannot test. Download with status to enable."))

    # ===== J. risk ratios =====
    # #25 outward EWB value / GSTR-1 taxable
    g1_total_tax_val = sum(d["taxable"] for no, d in g1inv.items())
    if ewb_out_file_supplied:
        ratio25 = ewb_out_total / g1_total_tax_val if g1_total_tax_val else 0
        R.append(F("#25", "Ratio: EWB-Out assessable / GSTR-1 B2B taxable",
                   FLAG if ratio25 < 0.9 else PASS,
                   f"EWB-Out {ewb_out_total:,.2f} / GSTR-1 B2B {g1_total_tax_val:,.2f} = {ratio25:.3f}. "
                   + ("Ratio <0.9 -> sizeable B2B supply with no goods movement; check accommodation bills."
                      if ratio25 < 0.9 else
                      "Ratio ≥0.9 -> most B2B supply backed by goods movement. (B2C/services not in EWB-Out.)"),
                   []))
    else:
        R.append(F("#25", "Ratio: EWB-Out assessable / GSTR-1 B2B taxable", SKIP,
                   "No outward EWB workbook supplied -- cannot compute.", []))

    # #26 inward EWB value / 2B taxable value (Excel) or ITC tax (PDF fallback)
    if twob_lines:
        b2b_taxable = sum(x["taxable"] for x in twob_lines["b2b"])
        ratio26 = ewb_in_assess / b2b_taxable if b2b_taxable else 0
        R.append(F("#26", "Ratio: EWB-In assessable / 2B B2B taxable", INFO,
                   f"EWB-In assessable {ewb_in_assess:,.2f} / 2B B2B taxable {b2b_taxable:,.2f} = {ratio26:.2f}. "
                   "Now a like-for-like VALUE ratio (both taxable value) from the 2B Excel. EWB-In covers only "
                   "goods movement while 2B B2B also includes goods received without an e-way bill (sub-"
                   "threshold / local), so EWB-In < 2B is normal; EWB-In >> 2B would suggest goods received "
                   "without matching ITC documents.",
                   []))
    elif b2b_itc is not None:
        R.append(F("#26", "Ratio: EWB-In assessable / 2B ITC", INFO,
                   f"EWB-In assessable {ewb_in_assess:,.2f}; 2B ITC tax {b2b_itc:,.2f}; EWB-In tax "
                   f"{ewb_in_tax:,.2f}. DO NOT OVER-WEIGHT: numerator is goods-movement value, denominator is "
                   "ITC tax (goods+services+RCM) — different bases/units. A true value ratio needs the 2B "
                   "Excel (taxable value), which isn't in the PDF summary. Scale context only.",
                   []))
    else:
        R.append(F("#26", "Ratio: EWB-In assessable / 2B", SKIP,
                   f"GSTR-2B not supplied -- cannot compute. EWB-In assessable {ewb_in_assess:,.2f} shown "
                   "for reference only.", []))

    # #27 same vehicle multiple trips between same GSTIN pair (circular-trading indicator)
    veh = {}
    for tag, lst in (("OUT", ewb_out), ("IN", ewb_in)):
        for e in lst:
            if not e["vehicle"]:
                continue
            key = (e["vehicle"], e["from_gstin"], e["to_gstin"])
            veh.setdefault(key, []).append((tag, e["docno"], e["ewbdate"]))
    rep = [(v, fr, to, len(trips)) for (v, fr, to), trips in veh.items() if len(trips) > 1]
    rep.sort(key=lambda x: -x[3])
    R.append(F("#27", "Same vehicle, repeated trips on same GSTIN pair",
               PASS if not rep else REVW,
               "Repeated vehicle between the same GSTIN pair can be normal (regular supplier) but is also "
               "a circular-trading indicator -> eyeball the high-frequency pairs." if rep else
               "No vehicle repeats on the same GSTIN pair.",
               [("Vehicle", "From", "To", "#trips")] + rep[:30]))

    R.sort(key=lambda x: (SEV_ORDER[x.sev], x.ref))
    return R


# ----------------------------------------------------------------------
# WRITE
# ----------------------------------------------------------------------
FILL = {FLAG: PatternFill("solid", fgColor="FFC7CE"), REVW: PatternFill("solid", fgColor="FFEB9C"),
        INFO: PatternFill("solid", fgColor="DDEBF7"), PASS: PatternFill("solid", fgColor="C6EFCE"),
        SKIP: PatternFill("solid", fgColor="E7E6E6")}
FONTC = {FLAG: Font(bold=True, color="9C0006"), REVW: Font(bold=True, color="9C6500"),
         INFO: Font(bold=True, color="2F5496"), PASS: Font(bold=True, color="006100"),
         SKIP: Font(bold=True, color="808080")}
HEAD = PatternFill("solid", fgColor="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def write(R, outpath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "EWB Findings"
    ws.cell(1, 1, f"E-WAY BILL RECONCILIATION (27-check matrix) — {raw.PERIOD_LABEL}").font = Font(bold=True, size=13, color="1F3864")
    ws.cell(2, 1, f"GSTIN {SELF_GSTIN}  |  {COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    import datetime as _d
    _src = f"2B Excel: {_LAST_2B_FILE}" if (_LAST_2B_SRC == "excel") else "2B PDF summary (no Excel found)"
    ws.cell(4, 1, f"BUILD v3 (line-level 2B + e-invoice + #3 purchase-return reclass)  |  "
                  f"generated {_d.datetime.now():%Y-%m-%d %H:%M:%S}  |  source: {_src}").font = \
        Font(size=9, italic=True, color="C00000")
    counts = {s: sum(1 for x in R if x.sev == s) for s in (FLAG, REVW, INFO, PASS, SKIP)}
    ws.cell(3, 1, "  ".join(f"{s}: {c}" for s, c in counts.items())).font = Font(bold=True, size=10)
    hdr = ["Ref", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(5, i, h); c.fill = HEAD; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    r = 6
    detail_blocks = []
    for f in R:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.sev); cv.fill = FILL[f.sev]; cv.font = FONTC[f.sev]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, f.detail)
        for c in range(1, 5):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
        if len(f.rows) > 1:
            detail_blocks.append(f)
    for col, w in zip("ABCD", [6, 42, 10, 110]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    # Detail sheet: the actual offending rows for each check that has them
    ws2 = wb.create_sheet("EWB Detail")
    rr = 1
    ws2.cell(rr, 1, "PER-CHECK DETAIL ROWS").font = Font(bold=True, size=12, color="1F3864"); rr += 2
    for f in detail_blocks:
        ws2.cell(rr, 1, f"{f.ref}  {f.title}  [{f.sev}]").font = Font(bold=True, color="1F3864"); rr += 1
        head = f.rows[0]
        for j, h in enumerate(head, 1):
            c = ws2.cell(rr, j, h); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
        rr += 1
        for row in f.rows[1:]:
            for j, v in enumerate(row, 1):
                # guard: a string starting with = + - @ is treated by Excel as a formula; prefix a space
                if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                    v = " " + v
                c = ws2.cell(rr, j, v)
                c.border = BORDER; c.font = Font(size=10)
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            rr += 1
        rr += 1
    for col, w in zip("ABCDE", [22, 18, 18, 14, 10]):
        ws2.column_dimensions[col].width = w

    wb.save(outpath)
    return counts


def main():
    """Legacy/manual-testing entry point -- master_build.py does NOT call this;
    it drives read_gstr1_invoices/read_einv_invoices/run() directly per month
    via run_monthly_pipeline.py. Kept working for standalone spot-checks:
        python gst_eway_recon.py <month, e.g. Jan-23>
    (still requires raw.GSTR1_FILE / GSTR3B_FILE / EINV_FILE / GSTR2B_FILE and
    ewb_annual_parser-based ewb_out/ewb_in lists to be set up by the caller.)
    """
    import sys as _sys, os as _os
    global _LAST_2B_SRC, _LAST_2B_FILE, _LAST_EINV_FILE
    if len(_sys.argv) < 2:
        raise SystemExit("Usage: python gst_eway_recon.py <month, e.g. Jan-23>")
    month = _sys.argv[1]

    print("="*70)
    print(f"GST E-WAY BILL RECONCILIATION — input check ({month})")
    print("="*70)
    for label, fn in [("GSTR-1", raw.GSTR1_FILE), ("E-Invoice", raw.EINV_FILE),
                      ("GSTR-3B", raw.GSTR3B_FILE), ("GSTR-2B xlsx", GSTR2B_FILE)]:
        ok = fn and _os.path.exists(fn)
        print(f"  {'OK ' if ok else 'MISS'} {label:13} {fn if fn else '(not found)'}")
    print("="*70)

    g1inv = read_gstr1_invoices(raw.GSTR1_FILE, month)
    einv = read_einv_invoices(raw.EINV_FILE, month) if raw.EINV_FILE else {}
    _LAST_EINV_FILE = _os.path.basename(raw.EINV_FILE) if raw.EINV_FILE else None
    g3b = raw.parse_gstr3b(raw.GSTR3B_FILE, month)
    b2b = g2b.summary_for_month(GSTR2B_FILE, month)

    R = run([], [], g1inv, einv, g3b, b2b)
    _LAST_2B_SRC = b2b.get("_source", "pdf-hardcoded").split("-")[0]
    _LAST_2B_FILE = b2b.get("_file")
    counts = write(R, OUTPUT_FILE)
    src = b2b.get("_source", "pdf-hardcoded")
    print(f"Saved: {OUTPUT_FILE}")
    print(f"EWB-Out/In: not loaded in this standalone stub (use master_build.py for the full run)  |  2B source: {src}"
          + (f" ({len(b2b['_lines']['b2b'])} B2B invoices, {len(b2b['_lines']['cdnr'])} CDNR notes)"
             if b2b.get("_lines") else ""))
    print(f"E-invoice: {'loaded ('+str(len(einv))+' invoices)' if einv else 'NOT FOUND -> #5/#6/#7/#17 limited'}")
    if src != "excel":
        print("  >>> NOTE: 2B Excel was NOT used. #10-#13 and #3 reclass are LIMITED.")
    print("  " + "  ".join(f"{s}:{c}" for s, c in counts.items()))
    for f in R:
        print(f"  [{f.sev:8}] {f.ref:4} {f.title}")


if __name__ == "__main__":
    main()
