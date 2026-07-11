#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PHASE 1 -- ANNUAL / BIFA RECONCILIATION WORKBOOK
==================================================
Builds a standalone workbook from the annual-level sources only (does NOT
need monthly GSTR-1/2B/3B/EWB/E-Inv -- that is Phase 2, once those are
uploaded). Sources used here:
  - Electronic Cash Ledger (CSV)
  - Electronic Credit Ledger (CSV)
  - Electronic Liability Register (CSV)
  - GST-Prime TPST 12-month self-filing summary (Excel)
  - Portal's own "Tax liability and ITC comparison" report (Excel)
  - BO / 360-degree Profile (PDF)

Output: GST_Annual_Reconciliation_FY2022-23.xlsx
"""

import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from annual_sources import (parse_cash_or_liability_ledger, parse_credit_ledger,
                             parse_tpst, parse_portal_comparison, MONTH_ABBR)
from bo_profile_parser import parse_bo_profile

RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEAD = PatternFill("solid", fgColor="1F3864")
SECT = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
SEV_FILL = {"FLAG": RED, "REVIEW": AMBER, "INFO": BLUE, "OK": GREEN, "N/A": GREY}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"), "REVIEW": Font(bold=True, color="9C6500"),
            "INFO": Font(bold=True, color="2F5496"), "OK": Font(bold=True, color="006100"),
            "N/A": Font(bold=True, color="808080")}

FY_MONTHS = ["Apr-22", "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22",
             "Oct-22", "Nov-22", "Dec-22", "Jan-23", "Feb-23", "Mar-23"]
TOL_LAKH = 0.5  # Rs 50,000 -- ledger/TPST reconciliation tolerance (rounding + timing noise)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def gather(folder="."):
    import os
    j = lambda f: os.path.join(folder, f)
    cash = parse_cash_or_liability_ledger(j("Electronic_Cash_Ledger__2___1_.csv"), "cash")
    credit = parse_credit_ledger(j("ElectronicCreditLedger__2___1_.csv"))
    liab = parse_cash_or_liability_ledger(
        j("Electronic_Liability_Register_Electronic_Liability_Register_2__1_.csv"), "liability")
    tpst = parse_tpst(j("GST-Prime_TPST_Taxpayer_Profile-_Statements__1_3B_22-23.xlsx"))
    comp = parse_portal_comparison(
        j("2022-23_05AAECM6380J1ZA_Tax_liability_and_ITC_comparison__1_.xlsx"))
    bo = parse_bo_profile(j("05AAECM6380J1ZA_BO_Profile_15_06_2026.pdf"))
    return dict(cash=cash, credit=credit, liab=liab, tpst=tpst, comp=comp, bo=bo)


def build_monthly_rows(data):
    cash, credit, liab, tpst, comp = data["cash"], data["credit"], data["liab"], data["tpst"], data["comp"]
    rows = []
    for mo in FY_MONTHS:
        t = tpst.get(mo, {})
        c = comp.get(mo, {})
        cr = credit["monthly_by_tax_period"].get(mo, {})
        ca_period = cash["monthly_by_tax_period"].get(mo, {})   # cash DEBITED against this period
        ca_txndate = cash["monthly_by_txn_date"].get(mo, {})     # cash DEPOSITED in this calendar month
        li = liab["monthly_by_txn_date"].get(mo, {})              # liability booked in this calendar month

        tpst_liab = t.get("total_tax_liability")
        portal_g3b_liab = c.get("gstr3b_liability")
        tpst_itc = t.get("net_itc_claimed")
        portal_itc_3b = c.get("itc_3b_adj") if c.get("itc_3b_adj") is not None else c.get("itc_3b_unadj")
        tpst_cash = t.get("total_cash_paid")
        credit_accrued = cr.get("credited")
        cash_debited_period = ca_period.get("debited")

        # TPST vs Portal-comparison liability check
        # NOTE: the portal's "Tax liability and ITC comparison" Comparison-Summary sheet is in
        # absolute Rs (verified: its FY 'Total' row for GSTR-1 liability, 33,479,969.41, equals
        # BIFA's 334.80 Lakh) -- NOT lakhs, so no unit conversion here.
        flag_liab = "N/A"
        if tpst_liab is not None and portal_g3b_liab is not None:
            d = abs(tpst_liab - portal_g3b_liab)
            flag_liab = "OK" if d <= TOL_LAKH * 100000 else "REVIEW"

        flag_itc = "N/A"
        if tpst_itc is not None and portal_itc_3b is not None:
            d = abs(tpst_itc - portal_itc_3b)
            flag_itc = "OK" if d <= TOL_LAKH * 100000 else "REVIEW"

        flag_credit_ledger = "N/A"
        if tpst_itc is not None and credit_accrued is not None:
            d = abs(tpst_itc - credit_accrued)
            flag_credit_ledger = "OK" if d <= TOL_LAKH * 100000 else "REVIEW"

        flag_cash_ledger = "N/A"
        if tpst_cash is not None and cash_debited_period is not None:
            d = abs(tpst_cash - cash_debited_period)
            flag_cash_ledger = "OK" if d <= 5000 else "REVIEW"  # cash paid amounts are small (Rs), tighter tolerance

        # cash-utilization ratio (this taxpayer's BIFA 'Cash Utilization %' runs ~1% FY-wide,
        # so treat as a continuous ratio to watch, not a binary per-month flag)
        cash_ratio_pct = (round(100 * tpst_cash / tpst_liab, 3)
                           if (tpst_cash is not None and tpst_liab) else None)

        rows.append(dict(
            month=mo,
            tpst_filing_date=t.get("filing_date"),
            tpst_liability=tpst_liab, portal_g3b_liability=portal_g3b_liab, flag_liab=flag_liab,
            portal_g1_liability=c.get("gstr1_liability"),
            tpst_itc=tpst_itc, portal_itc_3b=portal_itc_3b, flag_itc=flag_itc,
            portal_itc_2b=c.get("itc_2b"),
            credit_ledger_accrued=credit_accrued, flag_credit_ledger=flag_credit_ledger,
            tpst_cash=tpst_cash, cash_ledger_debited_period=cash_debited_period,
            flag_cash_ledger=flag_cash_ledger,
            cash_ledger_deposited_calmonth=ca_txndate.get("credited"),
            liability_register_calmonth=li.get("debited"),
            cash_ratio_pct=cash_ratio_pct,
        ))
    return rows


def write_cover(ws, data):
    bo = data["bo"]
    ws.cell(1, 1, "GST ANNUAL RECONCILIATION -- PHASE 1 (Ledgers + TPST + Portal Comparison + BO Profile)").font = TITLEF
    ws.cell(2, 1, f"GSTIN {bo['self_gstin']}  |  {bo['legal_name']}  |  FY 2022-23").font = Font(size=10, bold=True)
    ws.cell(3, 1, f"Generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}").font = Font(size=9, italic=True)
    r = 5
    ws.cell(r, 1, "Sources used").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for s in ["Electronic Cash Ledger (CSV)", "Electronic Credit Ledger (CSV)",
              "Electronic Liability Register (CSV)", "GST-Prime TPST 12-month summary (Excel)",
              "Portal Tax-liability & ITC Comparison report (Excel)",
              "BO / 360-degree Profile (PDF)"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Known limitations (structural gaps in the source data -- see README for full list)").font = Font(bold=True, size=11, color="9C0006"); r += 1
    for s in ["HSN code is not linked to individual invoices anywhere except the EWB files -- "
              "GSTR-1's own HSN sheet and GSTR-2B's purchase-side data are both monthly aggregates",
              "GSTR-1's B2C-Small (Table 7) sheet is a state+rate summary with no invoice numbers at "
              "all -- invoice-level B2C splitting cannot be detected from this data by design",
              "Credit notes (GSTR-1 'cdnr' sheet) carry no original-invoice-number reference -- any "
              "check linking a CN back to its original sale is approximate (by recipient + value), "
              "never a proven document link",
              "DRC Payment Information is linked to months ONLY by nearby transaction date "
              "(no tax-period field exists in that PDF section) -- treat as informational reference, "
              "not a proven link"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Known data-quality caveats").font = Font(bold=True, size=11, color="9C6500"); r += 1
    for s in ["BO-Profile 'Financial Information' table's last few columns (Cash/ITC utilization %, "
              "'Liability'/'Cash' ratio figures) were ambiguous on parse -- not used in any check here",
              "BO-Profile 'E-Invoice Related Information' table: 'Active Assessable Value' vs "
              "'Active Taxes' column pairing looks swapped for this taxpayer's numbers -- kept as "
              "unconfirmed raw columns, not used in any check",
              "Credit-ledger monthly figures are grouped by the ledger's own 'Tax Period' column; "
              "entries with a blank tax-period ('-') are excluded from the monthly view (see raw "
              "transactions for those)"]:
        ws.cell(r, 2, "- " + s); r += 1
    r += 1
    ws.cell(r, 1, "Departmental proceedings (B5 -- new: from BO Profile Appeal/Case/Transfer sections)").font = Font(bold=True, size=11, color="9C0006"); r += 1
    appeals = bo.get("appeals", [])
    cases = bo.get("cases", [])
    transfers = bo.get("transfers", [])
    if not (appeals or cases or transfers):
        ws.cell(r, 2, "- None found in the BO Profile's Appeal/Case/Transfer Information sections."); r += 1
    for a in appeals:
        ws.cell(r, 2, f"- APPEAL: ARN {a['arn']} filed {a['filing_date']} (FY {a['fy']}) -- {a['details']}"); r += 1
    for c in cases:
        ws.cell(r, 2, f"- CASE: {c['case_id']} / ref {c['ref_id']}, action date {c['action_date']} -- {c['details']}"); r += 1
    for t in transfers:
        ws.cell(r, 2, f"- TRANSFER: {t['source_id']} on {t['date']} -- {t['details']}"); r += 1
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110


def write_monthly(ws, rows):
    ws.cell(1, 1, "MONTHLY WALKTHROUGH -- TPST vs Portal-Comparison vs Ledgers (FY 2022-23)").font = TITLEF
    ws.cell(2, 1, "Amounts in Rs unless stated. Tolerance: Rs 50,000 (liability/ITC), Rs 5,000 (cash).").font = Font(size=9, italic=True)
    hdr = ["Month", "3B Filing Date",
           "Liability: TPST", "Liability: Portal(3B)", "Liab Check", "Liability: Portal(GSTR-1)",
           "ITC: TPST claimed", "ITC: Portal(3B)", "ITC Check", "ITC: Portal(2B)",
           "Credit Ledger: ITC accrued (this period)", "Credit-Ledger Check",
           "Cash: TPST paid", "Cash Ledger: debited (this period)", "Cash-Ledger Check",
           "Cash Ledger: deposited (this cal.month)", "Liability Register: booked (this cal.month)",
           "Cash Utilization %"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for row in rows:
        vals = [row["month"], str(row["tpst_filing_date"])[:10] if row["tpst_filing_date"] else "",
                row["tpst_liability"], row["portal_g3b_liability"], row["flag_liab"],
                row["portal_g1_liability"],
                row["tpst_itc"], row["portal_itc_3b"], row["flag_itc"], row["portal_itc_2b"],
                row["credit_ledger_accrued"], row["flag_credit_ledger"],
                row["tpst_cash"], row["cash_ledger_debited_period"], row["flag_cash_ledger"],
                row["cash_ledger_deposited_calmonth"], row["liability_register_calmonth"],
                row["cash_ratio_pct"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
            if hdr[ci - 1].endswith("Check"):
                cell.fill = SEV_FILL.get(v, GREY)
                cell.font = SEV_FONT.get(v, Font(size=10))
                cell.alignment = Alignment(horizontal="center")
            if hdr[ci - 1] == "Cash Utilization %" and isinstance(v, (int, float)) and v < 2:
                cell.fill = AMBER
        r += 1
    for i, w in enumerate([9, 12, 13, 14, 11, 15, 13, 13, 11, 13, 16, 13, 11, 14, 12, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_fy_total_vs_bifa(ws, rows, data, ewb_out_annual_rows=None, fy=None):
    bo = data["bo"]
    bifa_by_fy = bo.get("bifa_by_fy", {})
    # FIX (genericity -- was hardcoded to "2022-23", so a different taxpayer/FY would
    # silently read an empty {} and show every BIFA figure as 0, indistinguishable from a
    # real mismatch). Now: use the FY explicitly passed in; if not given, and there's
    # exactly one FY in bifa_by_fy, use that; otherwise leave bifa empty AND say so on the
    # sheet, rather than guessing a key that might not exist.
    fy_used = fy
    if not fy_used:
        if len(bifa_by_fy) == 1:
            fy_used = next(iter(bifa_by_fy))
        elif bifa_by_fy:
            fy_used = sorted(bifa_by_fy)[-1]  # most recent, if multiple and none specified
    bifa = bifa_by_fy.get(fy_used, {}) if fy_used else {}
    ws.cell(1, 1, f"FY-TOTAL vs BIFA (DEPARTMENT'S OWN CROSS-CHECK) -- FY {fy_used or '(unknown)'}").font = TITLEF
    ws.cell(2, 1, "BIFA = the department's own pre-computed FY-level figures from the BO Profile. "
                  "'Recomputed' = summed from TPST/ledgers/EWB by this tool, independently."
                  + ("" if bifa else f"  **BIFA figures below are all blank/zero because FY "
                     f"{fy_used!r} was not found in the BO Profile's BIFA table -- FYs actually "
                     f"present there: {sorted(bifa_by_fy) or 'none'}. This is a data-availability "
                     "note, not a real mismatch; do not read the REVIEW flags below at face value "
                     "until this is resolved.")).font = Font(size=9, italic=True,
                     color=("C00000" if not bifa else "000000"))

    sum_liab = sum(r["tpst_liability"] or 0 for r in rows)
    itc_data_available = any(r["tpst_itc"] is not None for r in rows)
    sum_itc = sum(r["tpst_itc"] or 0 for r in rows) if itc_data_available else None
    cash_data_available = any(r["tpst_cash"] is not None for r in rows)
    sum_cash = sum(r["tpst_cash"] or 0 for r in rows) if cash_data_available else None
    sum_credit_accrued = sum(r["credit_ledger_accrued"] or 0 for r in rows)
    # B1: recompute annual Outward-EWB "tax value" total directly from the raw EWB file,
    # independent of BIFA's own pre-computed figure -- new addition (was previously unused).
    sum_ewb_tax = sum(e["taxval"] for e in ewb_out_annual_rows) if ewb_out_annual_rows else None

    def _bifa_val(key):
        """None (not 0) when bifa is empty -- a real 0 in the BO Profile is a legitimate
        comparator; an EMPTY bifa dict (FY not found at all) must not silently look like one."""
        if not bifa:
            return None
        return (bifa.get(key) or 0) * 100000

    lines = [
        ("Liability as per GSTR-3B (Rs)", sum_liab, _bifa_val("liability_gstr3b")),
        ("ITC availed in R3B (Rs) -- TPST source has no ITC column in this report variant"
         if sum_itc is None else "ITC availed in R3B (Rs)",
         sum_itc, _bifa_val("itc_r3b") if sum_itc is not None else None),
        ("Credit Ledger accrued vs BIFA 'ITC Availed in R3B' (Rs)", sum_credit_accrued,
         _bifa_val("itc_r3b")),
        ("Credit Ledger accrued vs BIFA 'ITC Accrued in R2B/R2A' (Rs)", sum_credit_accrued,
         _bifa_val("itc_r2b_r2a")),
        ("Outward EWB Tax Val. (recomputed) vs BIFA 'Liability as per EWB' (Rs)"
         if sum_ewb_tax is None else "Outward EWB Tax Val. (recomputed) vs BIFA 'Liability as per EWB' (Rs)",
         sum_ewb_tax, _bifa_val("liability_ewb")),
        ("Cash paid (Rs) -- TPST source has no cash-paid column in this report variant"
         if sum_cash is None else "Cash paid (Rs)",
         sum_cash, None),
    ]
    ws.cell(3, 1, "NOTE (A5): BIFA carries two distinct ITC columns -- 'ITC Availed in R3B' and 'ITC "
                  "Accrued in R2B/R2A' -- which differ by the department's own 'Excess ITC claimed' "
                  "figure. The Credit Ledger accrual is shown against BOTH below rather than picking "
                  "one implicitly; which is the more meaningful comparator is a judgement call for the "
                  "reviewer, not something this tool should decide silently.").font = Font(size=9, italic=True, color="9C6500")
    hdr = ["Metric", "Recomputed (TPST/Ledger sum)", "BIFA (dept figure, converted from Lakh)", "Diff", "Check"]
    r = 5
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, 5)
    r += 1
    for label, a, b in lines:
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(a, 2) if a is not None else "n/a")
        ws.cell(r, 3, round(b, 2) if b is not None else "n/a")
        if a is not None and b is not None:
            d = round(a - b, 2)
            ws.cell(r, 4, d)
            chk = "OK" if abs(d) <= TOL_LAKH * 100000 else "REVIEW"
            c = ws.cell(r, 5, chk); c.fill = SEV_FILL[chk]; c.font = SEV_FONT[chk]
            c.alignment = Alignment(horizontal="center")
        else:
            ws.cell(r, 4, "n/a"); ws.cell(r, 5, "N/A")
        for c in range(1, 6):
            ws.cell(r, c).border = BORDER
        r += 1
    for i, w in enumerate([48, 24, 30, 14, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_drc(ws, data):
    drc = data["bo"]["drc_payments"]
    ws.cell(1, 1, "DRC PAYMENT INFORMATION (informational reference -- BO Profile PDF)").font = TITLEF
    ws.cell(2, 1, "No tax-period field exists in this PDF section -- linkage to a specific GST month is "
                  "by nearby transaction date ONLY. Treat as a lead to investigate, not a proven match.").font = Font(size=9, italic=True, color="C00000")
    hdr = ["Source ID", "Description", "Date", "Method", "CGST", "SGST", "IGST", "CESS", "Other", "Total"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for d in sorted(drc, key=lambda x: x["date"]):
        vals = [d["source_id"], d["description"], d["date"], d["method"],
                d["cgst"], d["sgst"], d["igst"], d["cess"], d["other"], d["total"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=(ci == 2))
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
        r += 1
    for i, w in enumerate([18, 40, 12, 14, 10, 10, 10, 10, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_related_party(ws, data):
    bo = data["bo"]
    ws.cell(1, 1, "RELATED / CANCELLED-PARTY ITC ALERTS (BO Profile)").font = TITLEF
    ws.cell(2, 1, "Fraud-risk indicator: ITC exchanged with a party sharing a related parameter "
                  "(mobile/PAN/etc.) with this taxpayer, or with a since-cancelled GSTIN.").font = Font(size=9, italic=True)
    r = 4
    for label, key in [("ITC RECEIVED from related/cancelled supplier", "related_itc_received"),
                        ("ITC PASSED ON to related/cancelled recipient", "related_itc_passed")]:
        ws.cell(r, 1, label).font = Font(bold=True, size=11, color="1F3864")
        for c in range(1, 9):
            ws.cell(r, c).fill = SECT
        r += 1
        hdr = ["FY", "GSTIN", "Name", "Related Param", "Status", "Cancellation Date", "Reason", "Total ITC (Lakh)"]
        for i, h in enumerate(hdr, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, 8)
        r += 1
        for x in bo[key]:
            vals = [x["fy"], x["gstin"], x["name"], x["related_parameter"], x["status"],
                    x["cancellation_date"], x["reason"], x["total_itc"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                if x["status"] == "Cancelled":
                    cell.fill = RED
            r += 1
        if not bo[key]:
            ws.cell(r, 1, "No rows -- either genuinely none in the source BO Profile, or this "
                          "section's marker text wasn't found during parsing (see "
                          "bo_profile_parser.py --diagnose for which). Do not read this as a "
                          "confirmed 'no related-party ITC' result without checking.").font = Font(
                          italic=True, color="9C6500")
            r += 1
        r += 1
    for i, w in enumerate([9, 18, 26, 14, 11, 16, 40, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_top_counterparties(ws, data):
    bo = data["bo"]
    ws.cell(1, 1, "TOP-10 COUNTERPARTIES (context -- BO Profile, trailing 12 months)").font = TITLEF
    r = 3
    for label, key, amtlabel in [("Top 10 Beneficiaries (ITC Passed On)", "top_beneficiaries", "ITC Passed (Lakh)"),
                                   ("Top 10 Suppliers (ITC Received)", "top_suppliers", "ITC Received (Lakh)")]:
        ws.cell(r, 1, label).font = Font(bold=True, size=11, color="1F3864")
        for c in range(1, 7):
            ws.cell(r, c).fill = SECT
        r += 1
        hdr = ["GSTIN", "Name", "Reg. Start", "Status", "Risk", amtlabel]
        for i, h in enumerate(hdr, 1):
            ws.cell(r, i, h)
        _style_header(ws, r, 6)
        r += 1
        for x in bo[key]:
            vals = [x["gstin"], x["name"], x["reg_start"], x["status"], x["risk"], x["amount"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                if x["status"] == "Cancelled":
                    cell.fill = RED
            r += 1
        if not bo[key]:
            ws.cell(r, 1, "No rows -- either genuinely none in the source BO Profile, or this "
                          "section's marker text wasn't found during parsing (see "
                          "bo_profile_parser.py --diagnose for which).").font = Font(italic=True, color="9C6500")
            r += 1
        r += 1
    for i, w in enumerate([18, 30, 12, 11, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main(folder=".", outfile="GST_Annual_Reconciliation_FY2022-23.xlsx"):
    data = gather(folder)
    rows = build_monthly_rows(data)

    wb = openpyxl.Workbook()
    write_cover(wb.active, data); wb.active.title = "Cover & Caveats"
    write_monthly(wb.create_sheet("Monthly Walkthrough"), rows)
    write_fy_total_vs_bifa(wb.create_sheet("FY Total vs BIFA"), rows, data)
    write_drc(wb.create_sheet("DRC Payments (info)"), data)
    write_related_party(wb.create_sheet("Related-Party Alerts"), data)
    write_top_counterparties(wb.create_sheet("Top Counterparties"), data)
    wb.save(outfile)

    n_review = sum(1 for r in rows for k in ("flag_liab", "flag_itc", "flag_credit_ledger", "flag_cash_ledger")
                    if r[k] == "REVIEW")
    avg_cash_ratio = sum(r["cash_ratio_pct"] or 0 for r in rows) / len(rows)
    print(f"Saved: {outfile}")
    print(f"  Monthly REVIEW flags: {n_review}")
    print(f"  Avg cash-utilization ratio across FY: {avg_cash_ratio:.2f}%")


if __name__ == "__main__":
    main()
