#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MASTER BUILD  --  the single entry point for the whole tool.

    python master_build.py [folder]

Put every input file (however many months you have) in one folder next to
all the .py files listed in HOW_TO_RUN.md, then run this. It will:
  1. Classify every file in the folder by content signature (folder_classifier.py)
  2. Run the full single-month engine (Comparison + Analysis-14 + EWB-27 +
     Doc-Series-Integrity) for every month that has at least GSTR-1 + GSTR-3B
  3. Run the Phase-1 annual reconciliation (ledgers + TPST + portal comparison
     + BO Profile) -- build_annual_workbook.py's logic, reused as-is
  4. Cross-month rectification pairing: match every GSTR-1 amendment-sheet row
     (b2ba/cdnra) found in ANY month against the ORIGINAL invoice/note in
     whichever earlier month first reported it, plus a best-effort DRC-payment
     cross-reference
  5. Write ONE workbook: Master Dashboard first, then per-month sheets, then
     the annual sheets, then the rectification-pairs sheet.

Missing months are handled gracefully -- whatever you have is what gets
analysed; the Dashboard states plainly which months are covered.
"""

import os
import re
import sys
import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from folder_classifier import classify_folder
import ewb_annual_parser as ewbp
from run_monthly_pipeline import run_month
import amendments as amd
import gst_unified_scrutiny as uni   # reuse its sheet writers
import gst_scrutiny_tool as raw

from annual_sources import (parse_cash_or_liability_ledger, parse_credit_ledger,
                             parse_tpst, parse_portal_comparison)
from bo_profile_parser import parse_bo_profile
import build_annual_workbook as annualwb
import hsn_fraud_checks as hfc
import filing_compliance as fc
import annual_return_parser as arp
import forensic_checks as fchk
import merged_period_utils as mpu

RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREY = PatternFill("solid", fgColor="E7E6E6")
HEAD = PatternFill("solid", fgColor="1F3864")
SECT = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
SEV_FILL = {"FLAG": RED, "MISMATCH": RED, "REVIEW": AMBER, "INFO": BLUE,
            "PASS": GREEN, "MATCH": GREEN, "SKIPPED": GREY}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"), "MISMATCH": Font(bold=True, color="9C0006"),
            "REVIEW": Font(bold=True, color="9C6500"), "INFO": Font(bold=True, color="2F5496"),
            "PASS": Font(bold=True, color="006100"), "MATCH": Font(bold=True, color="006100"),
            "SKIPPED": Font(bold=True, color="808080")}

MONTH_ORDER = ["Apr-22", "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22",
               "Oct-22", "Nov-22", "Dec-22", "Jan-23", "Feb-23", "Mar-23"]
# ^ RETAINED for any code path that still imports this name directly, but
# main() below no longer uses it -- see _sort_months_chronologically(), which
# builds the real month order dynamically from whatever months are actually
# present in the data (any number of FYs, not just one).

_MONTH_NUM = {v: k for k, v in mpu.CAL_MONTH_ABBR.items()}


def _month_sort_key(label):
    """'Jan-23' -> a real sortable value (2023, 1). Works across any number
    of years -- this is what makes multi-year runs order correctly without
    a hardcoded 12-month list."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", label) if label else None
    if not m:
        return (9999, 99)
    mon, yy = m.group(1), int(m.group(2))
    return (2000 + yy, _MONTH_NUM.get(mon.title(), 0))


def _sort_months_chronologically(months):
    return sorted(months, key=_month_sort_key)


def _fy_label_for_month(label):
    """'Apr-22'->'2022-23', 'Jan-23'->'2022-23', 'Apr-23'->'2023-24'. Indian
    FY runs Apr-Mar. Used only for grouping/display, not for any parsing."""
    year, mon = _month_sort_key(label)
    if mon >= 4:
        return f"{year}-{str(year+1)[2:]}"
    return f"{year-1}-{str(year)[2:]}"


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def sheet_name(prefix, month, limit=31):
    n = f"{month} {prefix}"
    return n[:limit]


# ======================================================================
# CROSS-MONTH: rectification pairing
# ======================================================================
def build_rectification_pairs(month_results, month_g1_lines, months_covered):
    """month_g1_lines: {month: {(invno,rate): [taxable, igst]}} from each month's own B2B.
    months_covered: the ACTUAL chronologically-sorted list of months this run
    covers (any number of FYs) -- replaces the old hardcoded MONTH_ORDER so
    multi-year runs still correctly find "the earlier month" across an FY
    boundary, not just within one hardcoded 12-month window.
    For every amendment row in ANY month, find which earlier month first
    reported the 'original' invoice/note number, and pair them."""
    pairs = []
    for res in month_results:
        amend_month = res["month"]
        for row in res["b2ba"]:
            orig_month = None
            for m in months_covered:
                if m == amend_month:
                    break
                if m in month_g1_lines and any(k[0] == row["orig_invno"] for k in month_g1_lines[m]):
                    orig_month = m
                    break
            pairs.append(dict(
                kind="B2B Invoice Amendment", gstin=row["gstin"], recipient=row["recipient"],
                original_ref=row["orig_invno"], original_month=orig_month or "NOT FOUND in any earlier month provided",
                revised_ref=row["revised_invno"], amended_in_month=amend_month,
                taxable=row["taxable"], igst=row["igst"], cgst=row["cgst"], sgst=row["sgst"],
            ))
        for row in res["cdnra"]:
            pairs.append(dict(
                kind="Credit/Debit Note Amendment", gstin=row["gstin"], recipient="",
                original_ref=row["orig_noteno"], original_month="(note amendments not month-matched -- see original_ref)",
                revised_ref=row["revised_noteno"], amended_in_month=amend_month,
                taxable=row["taxable"], igst=row["igst"], cgst=row["cgst"], sgst=row["sgst"],
            ))
    return pairs


# ======================================================================
# WRITERS
# ======================================================================
def write_master_dashboard(ws, month_results, months_covered, months_gap, rect_pairs,
                            annual_review_count, hsn_findings=None, forensic_findings=None,
                            cancel_findings=None):
    fys = sorted(set(_fy_label_for_month(m) for m in months_covered))
    ws.cell(1, 1, "MASTER DASHBOARD -- ALL MONTHS + ANNUAL SOURCES, RANKED TOGETHER").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME}  |  "
                  f"generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}").font = Font(size=9, italic=True)
    ws.cell(3, 1, f"FY(s) covered: {', '.join(fys) or 'none'}  |  "
                  f"Months covered ({len(months_covered)}): {', '.join(months_covered) or 'none'}").font = Font(size=10, bold=True, color="006100")
    ws.cell(4, 1, f"Gaps WITHIN the covered span ({len(months_gap)}): {', '.join(months_gap) or 'none'} -- "
                  "a month between the earliest and latest month supplied that has NO GSTR-1+GSTR-3B pair; "
                  "this is NOT the same as 'nothing was supplied for that FY' if the FY itself wasn't in "
                  "scope for this run at all.").font = Font(size=10, bold=True, color="9C0006")

    items = []
    RANK = {"FLAG": 0, "MISMATCH": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "MATCH": 3, "SKIPPED": 4}
    for res in month_results:
        mo = res["month"]
        for row in uni._comp_rows_iter(res["comparisons"]):
            sect, check, ll, lv, rl, rv, diff, result, tag = row
            if result == "MISMATCH":
                items.append((RANK[result], mo, "Comparison", sect.split(".")[0], check, result,
                              f"{ll}={lv:,.2f} vs {rl}={rv:,.2f} (diff {diff:,.2f}). {tag}".strip()))
        for f in res["findings14"]:
            if f.severity in ("FLAG", "REVIEW"):
                items.append((RANK[f.severity], mo, "Analysis(14)", f.ref, f.title, f.severity, f.detail))
        for f in res["findings27"]:
            if f.sev in ("FLAG", "REVIEW"):
                items.append((RANK[f.sev], mo, "E-Way-Bill", f.ref, f.title, f.sev, f.detail))
        for dg in res["doc_gap"]:
            still_unexplained = dg.get("still_unexplained", dg.get("missing", []))
            explained_by_einv = dg.get("explained_by_cancelled_einvoice", [])
            if still_unexplained:
                items.append((0, mo, "Doc-Series", "Table13", "Missing invoice serials (unexplained)",
                               "FLAG", f"Range {dg['range']}: missing {still_unexplained}"))
            elif dg.get("missing") and dg.get("explained_by_declared_cancellation"):
                items.append((1, mo, "Doc-Series", "Table13",
                               "Missing invoice serials (explained by declared cancellation)",
                               "REVIEW", f"Range {dg['range']}: missing {dg['missing']} -- matches "
                               f"Table 13's own declared cancelled count exactly."))
            elif explained_by_einv:
                items.append((1, mo, "Doc-Series", "Table13",
                               "Missing invoice serials (explained by cancelled e-invoice)",
                               "REVIEW", f"Range {dg['range']}: {explained_by_einv} -- each of these serials "
                               f"has a CANCELLED e-invoice against it, so its absence from GSTR-1 is expected, "
                               f"not a real gap."))

    for f in (hsn_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            mo = f.numbers.get("month", "FY-wide")
            items.append((RANK[f.severity], mo, "HSN/Fraud", f.ref, f.title, f.severity, f.detail))

    # NEW: forensic (R13/R14) and cancelled-e-invoice cross-check findings now also feed the
    # ranked dashboard -- previously these only appeared on their own dedicated sheets, so a
    # FLAG here was invisible from the top-level view unless you opened that specific sheet.
    for f in (forensic_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "FY-wide", "Forensic (R13/R14)", f.ref, f.title, f.severity, f.detail))
    for f in (cancel_findings or []):
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "FY-wide", "Cancelled E-Inv", f.ref, f.title, f.severity, f.detail))

    items.sort(key=lambda x: (x[0], _month_sort_key(x[1]) if x[1] != "FY-wide" else (9998, 0), x[2]))


    FYWIDE_PIPELINES = ("HSN/Fraud", "Forensic (R13/R14)", "Cancelled E-Inv")
    monthly_items = [it for it in items if it[2] not in FYWIDE_PIPELINES]
    fywide_items = [it for it in items if it[2] in FYWIDE_PIPELINES]
    nflag_m = sum(1 for it in monthly_items if it[5] in ("FLAG", "MISMATCH"))
    nrev_m = sum(1 for it in monthly_items if it[5] == "REVIEW")
    nflag_f = sum(1 for it in fywide_items if it[5] in ("FLAG", "MISMATCH"))
    nrev_f = sum(1 for it in fywide_items if it[5] == "REVIEW")
    ws.cell(6, 1, f"Monthly actionable items: {len(monthly_items)}  (FLAG/MISMATCH: {nflag_m}  REVIEW: {nrev_m})   |   "
                  f"FY-wide items (HSN/Fraud + Forensic R13-R14 + Cancelled E-Inv): {len(fywide_items)}  "
                  f"(FLAG: {nflag_f}  REVIEW: {nrev_f})   |   "
                  f"ANNUAL-source REVIEW items: {annual_review_count}   |   "
                  f"Rectification pairs found: {len(rect_pairs)}").font = Font(bold=True, size=11, color="C00000")

    hdr = ["Month", "Pipeline", "Ref/Section", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        ws.cell(8, i, h)
    _style_header(ws, 8, 6)
    r = 9
    for (_, mo, pipeline, ref, title, result, detail) in items:
        ws.cell(r, 1, mo); ws.cell(r, 2, pipeline); ws.cell(r, 3, ref); ws.cell(r, 4, title)
        cv = ws.cell(r, 5, result); cv.fill = SEV_FILL[result]; cv.font = SEV_FONT[result]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 6, detail)
        for c in range(1, 7):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (4, 6)))
            if c != 5:
                cell.font = Font(size=10)
        r += 1
    if not items:
        ws.cell(r, 1, "No FLAG / MISMATCH / REVIEW across any supplied month.").font = Font(italic=True, color="006100")
    for col, w in zip("ABCDEF", [9, 14, 12, 40, 11, 100]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A9"


def write_rectification_sheet(ws, pairs, drc_payments):
    ws.cell(1, 1, "CROSS-MONTH RECTIFICATION PAIRS").font = TITLEF
    ws.cell(2, 1, "Every GSTR-1 amendment-sheet row (b2ba/cdnra) found in any supplied month, "
                  "linked back to the month that first reported the original document -- "
                  "so an error and its later correction both show up, together.").font = Font(size=9, italic=True)
    hdr = ["Kind", "GSTIN", "Recipient", "Original Ref", "Reported In (original month)",
           "Revised Ref", "Amended In (month)", "Taxable", "IGST", "CGST", "SGST"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not pairs:
        ws.cell(r, 1, "No amendment rows found in the month(s) supplied so far -- "
                      "this sheet fills in automatically as more months are added.").font = Font(italic=True, color="808080")
        r += 1
    for p in pairs:
        vals = [p["kind"], p["gstin"], p["recipient"], p["original_ref"], p["original_month"],
                p["revised_ref"], p["amended_in_month"], p["taxable"], p["igst"], p["cgst"], p["sgst"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
            if ci == 5 and isinstance(v, str) and "NOT FOUND" in v:
                cell.fill = AMBER
        r += 1
    r += 2
    ws.cell(r, 1, "DRC PAYMENTS -- informational reference only (no tax-period field in source; "
                  "match by nearby date, not proven)").font = Font(bold=True, size=11, color="1F3864")
    r += 1
    hdr2 = ["Source ID", "Description", "Date", "Method", "Total (Lakh)"]
    for i, h in enumerate(hdr2, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, 5)
    r += 1
    for d in sorted(drc_payments, key=lambda x: x["date"]):
        vals = [d["source_id"], d["description"], d["date"], d["method"], d["total"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDEFGHIJK", [26, 14, 26, 22, 16, 14, 16, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def write_doc_series(ws, month_results):
    ws.cell(1, 1, "DOCUMENT SERIES INTEGRITY (Table 13 vs actual B2B invoice numbers)").font = TITLEF
    ws.cell(2, 1, "'Found via fuzzy match' = same series, different prefix punctuation/segment than "
                  "Table 13's own printed range (confirmed real in this file's own JWI series, not a "
                  "code defect) -- shown separately so the difference stays visible. A missing count "
                  "that exactly equals Table 13's own declared 'Cancelled' figure is marked explained, "
                  "not left as an unexplained red flag.").font = Font(size=9, italic=True)
    hdr = ["Month", "Series Range", "Table-13 Total", "Table-13 Cancelled", "Actually Found",
           "Missing Serials", "Found via fuzzy match (diff. prefix format)", "Status"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for res in month_results:
        for dg in res["doc_gap"]:
            missing = dg.get("missing", [])
            still_unexplained = dg.get("still_unexplained", missing)
            explained_decl = dg.get("explained_by_declared_cancellation", False)
            explained_einv = dg.get("explained_by_cancelled_einvoice", [])
            if dg.get("note"):
                status = "CHECK MANUALLY"
            elif not missing:
                status = "OK"
            elif not still_unexplained and explained_einv and not explained_decl:
                status = "EXPLAINED BY CANCELLED E-INVOICE"
            elif not still_unexplained and explained_decl:
                status = "EXPLAINED BY DECLARED CANCELLATION"
            elif not still_unexplained:
                status = "EXPLAINED (CANCELLATION + CANCELLED E-INVOICE)"
            else:
                status = "UNEXPLAINED -- REVIEW"
            missing_display = ", ".join(still_unexplained) if still_unexplained else (
                ", ".join(missing) if missing else "-- none --")
            fuzzy_note = ", ".join(dg.get("found_via_fuzzy_match", [])) or "--"
            if explained_einv:
                fuzzy_note += f"  |  cancelled e-invoice: {', '.join(explained_einv)}"
            vals = [res["month"], dg.get("range"), dg.get("table13_total"), dg.get("table13_cancelled"),
                    dg.get("actual_count"), missing_display or dg.get("note", "-- none --"),
                    fuzzy_note, status]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = BORDER
                cell.font = Font(size=10)
                cell.alignment = Alignment(wrap_text=(ci in (6, 7)))
            if status == "UNEXPLAINED -- REVIEW" or status == "CHECK MANUALLY":
                ws.cell(r, 6).fill = RED
                ws.cell(r, 8).fill = RED
            elif status in ("EXPLAINED BY DECLARED CANCELLATION", "EXPLAINED BY CANCELLED E-INVOICE",
                            "EXPLAINED (CANCELLATION + CANCELLED E-INVOICE)"):
                ws.cell(r, 8).fill = AMBER
            else:
                ws.cell(r, 8).fill = GREEN
            r += 1
    for col, w in zip("ABCDEFGH", [10, 24, 12, 14, 12, 40, 32, 26]):
        ws.column_dimensions[col].width = w


def write_hsn_fraud_checks(ws, findings):
    ws.cell(1, 1, "HSN-CODE-WISE + FRAUD-PATTERN CHECKS (FY-WIDE)").font = TITLEF
    ws.cell(2, 1, "Categories A (HSN-only) / B (POS-state-code) / C (combined) plus the numbered "
                  "fraud-pattern list. Every check is grounded against this taxpayer's real files -- "
                  "see each row's detail for the exact arithmetic, and the module docstring in "
                  "hsn_fraud_checks.py for what could NOT be checked and why.").font = Font(size=9, italic=True)
    nflag = sum(1 for f in findings if f.severity == "FLAG")
    nrev = sum(1 for f in findings if f.severity == "REVIEW")
    ws.cell(3, 1, f"FLAG: {nflag}   REVIEW: {nrev}   (INFO/PASS rows included below for "
                  f"completeness/audit-trail).").font = Font(bold=True, size=11, color="C00000")
    SEV_RANK = {"FLAG": 0, "REVIEW": 1, "INFO": 2, "PASS": 3}
    ordered = sorted(findings, key=lambda f: (SEV_RANK.get(f.severity, 9), f.ref))
    hdr = ["Ref", "Check", "Result", "Detail"]
    r = 5
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    for f in ordered:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity)
        cv.fill = SEV_FILL.get(f.severity, GREY)
        cv.font = SEV_FONT.get(f.severity, Font(size=10))
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, f.detail)
        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCD", [10, 40, 11, 110]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"


def _safe_parse_ledger(path, kind):
    """Wraps annual_sources.parse_cash_or_liability_ledger -- previously an
    absent ledger CSV crashed the ENTIRE run (unguarded open(None)). Now
    returns the same empty shape the real parser would return for a file
    with a header but zero transactions, so every downstream consumer
    (build_monthly_rows, write_cover, etc.) keeps working unchanged."""
    if not path or not os.path.exists(path):
        print(f"[info] {kind.title()} ledger not supplied -> annual ledger checks show N/A for this head.")
        return dict(opening=None, transactions=[], monthly_by_tax_period={}, monthly_by_txn_date={})
    return parse_cash_or_liability_ledger(path, kind)


def _safe_parse_credit(path):
    if not path or not os.path.exists(path):
        print("[info] Credit ledger not supplied -> annual credit-ledger checks show N/A.")
        return dict(opening=None, transactions=[], monthly_by_tax_period={})
    return parse_credit_ledger(path)


def _safe_parse_tpst(path):
    if not path or not os.path.exists(path):
        print("[info] TPST not supplied -> TPST-based annual checks show N/A.")
        return {}
    return parse_tpst(path)


def _safe_parse_portal(path):
    if not path or not os.path.exists(path):
        print("[info] Portal Tax-Liability-&-ITC-Comparison report not supplied -> that comparison shows N/A.")
        return {}
    return parse_portal_comparison(path)


_EMPTY_BO_PROFILE = dict(
    self_gstin=None, legal_name=None, trade_name=None, demographic={},
    financial_by_fy={}, bifa_by_fy={}, itc_passed_by_fy={}, itc_received_by_fy={},
    ewb_by_fy={}, einv_by_fy={}, refund_by_fy={},
    top_beneficiaries=[], top_suppliers=[], related_itc_received=[], related_itc_passed=[],
    drc_payments=[], appeals=[], cases=[], transfers=[],
)


def _safe_parse_bo(path):
    """BO/360-degree Profile is a rich but genuinely optional source (not
    every taxpayer/consultant has portal access to pull it). Previously
    parse_bo_profile(None) crashed the entire run; now degrades to an empty
    profile with every key the rest of the codebase expects, matching real
    output shape exactly (see bo_profile_parser.parse_bo_profile's own
    return statement)."""
    if not path or not os.path.exists(path):
        print("[info] BO/360-degree Profile not supplied -> DRC/related-party/BIFA cross-checks show N/A.")
        return dict(_EMPTY_BO_PROFILE)
    return parse_bo_profile(path)


# ======================================================================
# NEW SHEETS: Filing Compliance & Late Fee, Forensic Checks (R13/R14/BS-PL),
# Cancelled E-Invoices
# ======================================================================
def write_filing_compliance(ws, compliance_by_month):
    ws.cell(1, 1, "FILING COMPLIANCE -- ARN DATES, DUE DATES, LATE FEE & INTEREST").font = TITLEF
    ws.cell(2, 1, "Late fee per Section 47 (Rs 50/day normal, Rs 20/day nil, capped per Notification "
                  "07/2023-CT). Interest per Section 50(1), 18% p.a. on cash-paid tax, approximated using "
                  "the filing date as the payment date -- verify against the Liability Register's own "
                  "Interest-head entry (Forensic Framework Part 1, A4).").font = Font(size=9, italic=True)
    hdr = ["Month", "GSTR-1 ARN", "GSTR-1 Filed", "GSTR-1 Due", "GSTR-1 Late Fee (Rs)",
           "GSTR-3B ARN", "GSTR-3B Filed", "GSTR-3B Due", "GSTR-3B Late Fee (Rs)",
           "Sec 50 Interest (Rs)", "GSTR1-vs-3B gap (days)"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not compliance_by_month:
        ws.cell(r, 1, "No ARN date could be extracted -- see notes on the GSTR-1/GSTR-3B parsers; "
                      "check #10 in each month's Analysis(14) sheet also explains what's missing.").font = Font(italic=True)
    for c in (compliance_by_month or []):
        g1fee = (c.get("gstr1_late_fee") or {}).get("fee_payable")
        g3fee = (c.get("gstr3b_late_fee") or {}).get("fee_payable")
        interest = (c.get("gstr3b_interest") or {}).get("interest")
        vals = [c["month"], c.get("gstr1_arn"), str(c.get("gstr1_filing_date") or ""),
                str(c.get("gstr1_due_date") or ""), g1fee,
                c.get("gstr3b_arn"), str(c.get("gstr3b_filing_date") or ""),
                str(c.get("gstr3b_due_date") or ""), g3fee,
                interest, c.get("gstr1_vs_gstr3b_gap_days")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
                if v > 0 and hdr[ci-1].endswith("(Rs)"):
                    cell.fill = AMBER
        r += 1
    for col, w in zip("ABCDEFGHIJK", [9, 18, 12, 12, 16, 18, 12, 12, 16, 14, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def write_forensic_checks(ws, findings):
    ws.cell(1, 1, "FORENSIC CHECKS -- Part 2 of GST_Forensic_Comparison_Framework_v1.md").font = TITLEF
    ws.cell(2, 1, "R13 (turnover-gap rule) and R14 (four-way ITC reconciliation), plus the generic "
                  "Balance-Sheet/P&L rule engine (R0-R12) when structured BS/PL data was supplied. "
                  "See OCR_LIMITATION.md for why a scanned BS/PL PDF is not auto-parsed here.").font = Font(size=9, italic=True)
    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail"]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    SEV_RANK = {"FLAG": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "SKIPPED": 4}
    for f in sorted(findings, key=lambda x: (SEV_RANK.get(x.severity, 9), x.ref)):
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity)
        cv.fill = SEV_FILL.get(f.severity, GREY); cv.font = SEV_FONT.get(f.severity, Font(size=10))
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in (f.numbers or {}).items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDE", [8, 40, 11, 34, 100]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def write_cancelled_einvoices(ws, all_cancelled, cross_check_findings, col_found):
    ws.cell(1, 1, "CANCELLED E-INVOICES").font = TITLEF
    note = ("The E-Invoice export's IRN-status column WAS found and read." if col_found else
            "The E-Invoice export does NOT appear to carry a recognisable IRN-status/cancellation "
            "column under any of the header names this tool knows (IRN Status, Status, Cancel Date, "
            "etc.) -- 'zero cancelled e-invoices' below may mean either 'genuinely none' or 'this "
            "export doesn't expose that field'; check the raw file if that distinction matters.")
    ws.cell(2, 1, note).font = Font(size=9, italic=True, color=("006100" if col_found else "9C6500"))
    r = 4
    for f in cross_check_findings:
        cv = ws.cell(r, 1, f"[{f.severity}] {f.ref} {f.title}")
        cv.fill = SEV_FILL.get(f.severity, GREY); cv.font = Font(bold=True, size=10)
        ws.cell(r, 2, f.detail).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    hdr = ["Month", "Invoice No.", "Rate", "Taxable", "IGST", "IRN", "Cancel Date"]
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1
    if not all_cancelled:
        ws.cell(r, 1, "No cancelled e-invoices found (see note above for what this does/doesn't confirm).").font = Font(italic=True)
    for c in all_cancelled:
        vals = [c.get("month"), c.get("invno"), c.get("rate"), c.get("taxable"), c.get("igst"),
                c.get("irn"), c.get("cancel_date")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
        r += 1
    for col, w in zip("ABCDEFG", [9, 20, 8, 14, 14, 22, 14]):
        ws.column_dimensions[col].width = w



def write_hsn_review_table(ws, hsn_rows, month_label):
    """Appends a 'HSN RATE REVIEW' section into the CURRENT worksheet
    (the same sheet write_comparison() just wrote into), below whatever
    write_comparison() left there. One row per HSN code actually used this
    month (taxable value, rate charged, tax amount -- exactly as reported
    in GSTR-1's own HSN summary table, i.e. already the per-HSN aggregate,
    no further summing needed), plus two reference-rate columns and an
    explicit 'verify' flag on every row -- per the person's own request:
    show the raw HSN-wise breakdown side by side with whatever reference
    data exists, and flag every one for manual verification rather than
    silently trusting either reference (matches this tool's own severity
    discipline throughout -- see HSN_RATE_HISTORY's and
    _load_mcp_india_stack_hsn_table's own docstrings for exactly why
    neither reference is treated as ground truth on its own)."""
    on_date = hfc._month_label_to_date(month_label)
    mcp_table = hfc._load_mcp_india_stack_hsn_table()

    r = ws.max_row + 3
    ws.cell(r, 1, f"HSN RATE REVIEW -- TAXABLE SUPPLY BY HSN CODE ({month_label})").font = TITLEF
    r += 1
    ws.cell(r, 1, "Every HSN code used this month, with the rate/taxable/tax as actually reported "
                  "in GSTR-1's HSN summary, alongside two independent reference rates where "
                  "available. EVERY row is flagged VERIFY -- neither reference column is ground "
                  "truth on its own (see HSN & Fraud Pattern Checks sheet, checks A1/A1-EXT/A7, "
                  "for the full-strength automated comparison); this table is a fast side-by-side "
                  "worksheet for manual review, not a verdict.").font = Font(size=9, italic=True)
    r += 2
    hdr = ["HSN Code", "Description (taxpayer's own)", "Rate Charged (%)", "Taxable Value (Rs)",
           "Tax Amount (Rs)", "Curated Reference Rate (%)", "mcp-india-stack Reference Rate (%)", "Status"]
    for i, h in enumerate(hdr, 1):
        ws.cell(r, i, h)
    _style_header(ws, r, len(hdr))
    r += 1

    if not hsn_rows:
        ws.cell(r, 1, "No HSN summary rows for this month.").font = Font(italic=True)
        return

    for row in sorted(hsn_rows, key=lambda x: x["hsn"]):
        hsn, desc, rate = row["hsn"], row["desc"], row["rate"]
        tax_amt = row["igst"] + row["cgst"] + row["sgst"]

        curated = hfc._hsn_rate_for_date(hsn, on_date) if on_date else None
        curated_rate = curated["rate"] if curated else None
        curated_display = (curated_rate if curated_rate is not None else
                            ("unconfirmed for this period" if curated else "not in curated list"))

        mcp_prefix = hfc._hsn_prefix_lookup(hsn, mcp_table) if mcp_table else None
        mcp_rate = mcp_table[mcp_prefix][0] if mcp_prefix else None
        mcp_display = mcp_rate if mcp_rate is not None else "not found"

        mismatch = ((curated_rate is not None and abs(rate - curated_rate) > 0.01) or
                    (mcp_rate is not None and abs(rate - mcp_rate) > 0.01))
        status = "VERIFY -- reference rate differs" if mismatch else "VERIFY"

        vals = [hsn, desc, rate, row["taxable"], tax_amt, curated_display, mcp_display, status]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = BORDER
            cell.font = Font(size=10)
            if ci in (4, 5) and isinstance(v, float):
                cell.number_format = "#,##0.00"
            if ci == 8:
                cell.fill = RED if mismatch else AMBER
                cell.alignment = Alignment(horizontal="center")
        r += 1


def main(folder="."):
    print("Classifying folder (merged files)...")
    res = classify_folder(folder)

    if res.get("classify_warnings"):
        print("Classification warnings:")
        for w in res["classify_warnings"]:
            print("  -", w)

    if not res["self_gstin"]:
        raise RuntimeError(
            "Could not determine self-GSTIN from the supplied files (no EWB annual "
            "workbooks and no merged GSTR-1 'Read me' sheet found). Stopping -- "
            "every output filename/header depends on this being correct."
        )
    if not res["gstr1_month_map"] or not res["gstr3b_month_map"]:
        raise RuntimeError(
            "No merged GSTR-1 and/or GSTR-3B workbook found in the folder -- these are the only "
            "two genuinely mandatory inputs. Everything else (E-Invoice, GSTR-2B, EWB, ledgers, "
            "TPST, portal comparison, BO Profile, GSTR-9/9C/8A, BS/PL) is optional and degrades "
            "gracefully if absent."
        )

    raw.SELF_GSTIN = res["self_gstin"]
    raw.COMPANY_NAME = res["company_name"]

    print("Parsing annual EWB (whole FY(s), filtered per month)...")
    # CHANGED (multi-year + graceful degradation): now pools EWB rows across
    # EVERY outward/inward annual workbook found (any number of FYs), and
    # separately tracks whether each DIRECTION was supplied AT ALL -- see
    # gst_eway_recon.run()'s docstring for why that distinction matters.
    ewb_out_rows = [r for f in res["ewb_out_files"] for r in ewbp.parse_annual_ewb(f)]
    ewb_in_rows = [r for f in res["ewb_in_files"] for r in ewbp.parse_annual_ewb(f)]
    ewb_out_file_supplied = bool(res["ewb_out_files"])
    ewb_in_file_supplied = bool(res["ewb_in_files"])

    # A month is "covered" only if BOTH GSTR-1 and GSTR-3B have a resolved
    # file for it (possibly from DIFFERENT merged workbooks if this is a
    # multi-year run) -- sourced from the month->file maps, no fixed-length
    # calendar assumed, so this naturally spans any number of FYs.
    months_covered = _sort_months_chronologically(
        set(res["gstr1_month_map"]) & set(res["gstr3b_month_map"]))
    # Gaps WITHIN the covered span (not "out of a fixed 12"): any calendar
    # month strictly between the earliest and latest covered month that is
    # NOT itself covered.
    months_gap = []
    if len(months_covered) >= 2:
        y0, m0 = _month_sort_key(months_covered[0])
        y1, m1 = _month_sort_key(months_covered[-1])
        covered_set = set(months_covered)
        y, m = y0, m0
        while (y, m) <= (y1, m1):
            lbl = f"{mpu.CAL_MONTH_ABBR[m]}-{str(y)[2:]}"
            if lbl not in covered_set:
                months_gap.append(lbl)
            m += 1
            if m == 13:
                m = 1; y += 1

    print(f"Months covered: {months_covered}")
    fys_covered = sorted(set(_fy_label_for_month(m) for m in months_covered))
    print(f"FY(s) covered: {fys_covered}")

    # ---- filing compliance: read ARN dates ONCE per unique file (not once per
    #      month -- these functions scan the whole workbook each time) ----
    print("Extracting ARN / filing dates for late-fee & interest computation...")
    gstr1_arn_by_month, gstr3b_arn_by_month = {}, {}
    for f in res["gstr1_files"]:
        m, warns = fc.gstr1_arn_dates_by_month(f)
        gstr1_arn_by_month.update({k: v for k, v in m.items() if k != "_readme_fallback"})
        for w in warns:
            print("  [filing_compliance]", w)
    for f in res["gstr3b_files"]:
        gstr3b_arn_by_month.update(fc.gstr3b_arn_dates_by_month(f))
    # QRMP detection: a GSTR-1 marker that fans one marker into 3 months = quarterly filer.
    # (Approximate signal: if GSTR-1 has fewer distinct ARNs than months, quarterly is likely --
    # exact detection needs the real marker text's own Tax-Period field, already used inside
    # gstr1_arn_dates_by_month(); left as monthly-default here since this taxpayer's GSTR-3B is
    # confirmed one-sheet-per-month = non-QRMP, and QRMP support is otherwise architecturally
    # ready in filing_compliance.py's due_date_gstr1()/due_date_gstr3b() is_qrmp parameter.)
    gstr1_is_qrmp = gstr3b_is_qrmp = False

    month_results = []
    month_g1_lines = {}
    compliance_records = []
    all_cancelled_einvoices = []
    einv_cancel_col_found_any = False
    run_errors = []
    for m in months_covered:
        print(f"Running month {m}...")
        files = dict(gstr1=res["gstr1_month_map"].get(m), gstr3b=res["gstr3b_month_map"].get(m),
                     einv=res["einv_month_map"].get(m), gstr2b=res["gstr2b_month_map"].get(m))
        try:
            out = run_month(m, files, ewb_out_rows, ewb_in_rows, res["self_gstin"], res["company_name"],
                             ewb_out_file_supplied=ewb_out_file_supplied,
                             ewb_in_file_supplied=ewb_in_file_supplied,
                             gstr1_arn_by_month=gstr1_arn_by_month, gstr3b_arn_by_month=gstr3b_arn_by_month,
                             gstr1_is_qrmp=gstr1_is_qrmp, gstr3b_is_qrmp=gstr3b_is_qrmp)
        except Exception as ex:
            # ROBUSTNESS (new): one month's unexpected parsing issue no longer takes down the
            # entire multi-year run. Logged loudly, and the Dashboard will show this month as
            # having zero findings WITH an explicit note -- never silently dropped.
            print(f"  *** ERROR processing {m}: {ex} -- this month is SKIPPED, all other months continue. ***")
            run_errors.append((m, str(ex)))
            continue
        month_results.append(out)
        month_g1_lines[m] = out["comp_raw"]["g1"].get("lines", {})
        if out.get("compliance"):
            compliance_records.append(out["compliance"])
        if out.get("cancelled_einvoices"):
            all_cancelled_einvoices.extend(dict(c, month=m) for c in out["cancelled_einvoices"])
        einv_cancel_col_found_any = einv_cancel_col_found_any or out.get("einv_cancel_col_found", False)

    if run_errors:
        print(f"\n*** {len(run_errors)} month(s) failed and were skipped: {[m for m, _ in run_errors]} ***\n")

    rect_pairs = build_rectification_pairs(month_results, month_g1_lines, months_covered)

    print("Building Phase-1 annual reconciliation (graceful if any source is missing)...")
    annual_data = dict(
        cash=_safe_parse_ledger(res["cash_ledger"], "cash"),
        credit=_safe_parse_credit(res["credit_ledger"]),
        liab=_safe_parse_ledger(res["liab_ledger"], "liability"),
        tpst=_safe_parse_tpst(res["tpst"]),
        comp=_safe_parse_portal(res["portal_comparison"]),
        bo=_safe_parse_bo(res["bo_profile"]),
    )
    annual_rows = annualwb.build_monthly_rows(annual_data)
    annual_review_count = sum(1 for r in annual_rows
                               for k in ("flag_liab", "flag_itc", "flag_credit_ledger", "flag_cash_ledger")
                               if r[k] == "REVIEW")

    print("Running HSN-code-wise + fraud-pattern checks...")
    files_for_hsn = dict(gstr1=res["gstr1_merged"], gstr3b=res["gstr3b_merged"],
                          einv=res["einv_merged"], gstr2b=res["gstr2b_merged"])
    hsn_findings = hfc.run_all(files_for_hsn, ewb_out_rows, ewb_in_rows, months_covered, annual_data,
                                annual_rows, res["self_gstin"],
                                hsn_sac_master_override=res.get("hsn_sac_master_file"))

    # ---- NEW: annual-return-side sources (GSTR-9, GSTR-9C, Table 8A) + forensic checks ----
    print("Parsing GSTR-9 / GSTR-9C / Table 8A (optional; graceful if absent)...")
    gstr9 = arp.parse_gstr9(res["gstr9_files"][0] if res["gstr9_files"] else None)
    gstr9c = arp.parse_gstr9c(res["gstr9c_files"][0] if res["gstr9c_files"] else None)
    table8a = arp.parse_table_8a(res["table8a_files"][0] if res["table8a_files"] else None)

    print("Running forensic checks (R13 turnover-gap, R14 four-way ITC)...")
    # R13 needs GSTR-1's own Table-8 (exemp) rows per month -- re-derive from what parse_gstr1
    # already read (nil_exempt_taxable is a SUM; R13 needs to know row-count, so re-read the
    # 'exemp' sheet's row presence per month directly here, reusing the same content-based logic).
    exemp_rows_by_month = {}
    for m, res_m in zip(months_covered, month_results):
        g1 = res_m["comp_raw"]["g1"]
        exemp_rows_by_month[m] = [1] * 0 if g1.get("nil_exempt_taxable") in (None,) else (
            [1] if (g1.get("nil_taxable", 0) or g1.get("exempt_taxable", 0) or g1.get("nongst_taxable", 0)) else [])
    r13 = fchk.check_turnover_gap(gstr9c, exemp_rows_by_month)

    gstr2b_fy_total = None  # left None unless a whole-FY GSTR-2B aggregate is separately supplied;
    # architecturally ready (check_four_way_itc accepts it) -- wiring a true FY-sum from the
    # per-month 2B summaries already computed in month_results is the natural next step once a
    # full year of GSTR-2B data is available to sum (each month's summary is already in
    # month_results[i]['comp_raw']['b2b'] when available=True).
    _2b_sums = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
    _2b_any = False
    for res_m in month_results:
        b2b = res_m["comp_raw"]["b2b"]
        if b2b.get("available"):
            _2b_any = True
            _2b_sums["igst"] += b2b.get("ITC_all_other_IGST", 0) or 0
            _2b_sums["cgst"] += b2b.get("ITC_all_other_CGST", 0) or 0
            _2b_sums["sgst"] += b2b.get("ITC_all_other_SGST", 0) or 0
            _2b_sums["cess"] += b2b.get("ITC_all_other_CESS", 0) or 0
    if _2b_any:
        gstr2b_fy_total = _2b_sums

    annual_turnover = gstr9c.get("turnover_declared_gstr9") if gstr9c.get("available") else None
    r14 = fchk.check_four_way_itc(gstr9, gstr2b_fy_total, table8a, gstr9c, annual_turnover=annual_turnover)
    forensic_findings = [r13, r14]

    # ---- NEW: BS/P&L rule engine (R0-R12) -- FIX: this was built and tested standalone
    # (bs_pl_input.py's own __main__ block) but never actually wired into master_build.py's
    # pipeline, so it never appeared in the output workbook even when filled in. Tries to
    # import BS_PL_DATA from bs_pl_input.py in the SAME folder as the running script (not the
    # data folder -- this is a hand-typed structured input, not a content-detected file, per
    # OCR_LIMITATION.md); degrades to a single INFO finding if that file/variable is absent or
    # empty, exactly like every other optional source in this tool. ----
    try:
        import bs_pl_input as bspl_mod
        bs_pl_data = getattr(bspl_mod, "BS_PL_DATA", {})
    except ImportError:
        bs_pl_data = {}
    # SAFETY CHECK (new): refuse to use bs_pl_input.py's data if it's tagged for a DIFFERENT
    # GSTIN than the one actually being processed this run -- guards against a taxpayer's old
    # BS/PL figures silently being reused for a different taxpayer's tool run. A dict with no
    # '_gstin' tag at all is also refused (forces explicit tagging rather than an implicit
    # "assume it matches").
    if bs_pl_data:
        tagged_gstin = bs_pl_data.get("_gstin")
        if tagged_gstin != res["self_gstin"]:
            print(f"[warn] bs_pl_input.py's BS_PL_DATA is tagged for GSTIN {tagged_gstin!r}, "
                  f"but this run is processing {res['self_gstin']!r} -- REFUSING to use it "
                  "(prevents a stale/wrong-taxpayer's Balance Sheet figures being silently "
                  "applied). Update bs_pl_input.py's '_gstin' tag and figures for this taxpayer.")
            bs_pl_data = {}
    if bs_pl_data:
        print("Running BS/P&L rule engine (R0-R12) against bs_pl_input.BS_PL_DATA...")
        bo_for_bspl = annual_data.get("bo") if annual_data.get("bo", {}).get("drc_payments") else None
        forensic_findings += fchk.check_bs_pl_rules(bs_pl_data, gstr9c=gstr9c, bo_profile=bo_for_bspl)
    else:
        forensic_findings.append(fchk.Finding(
            "R0-R12", "Balance Sheet / P&L rule engine", "INFO",
            "bs_pl_input.py not found next to master_build.py, or its BS_PL_DATA dict is empty -- "
            "R0-R12 not run. Fill in bs_pl_input.py (see OCR_LIMITATION.md for why this is a "
            "hand-typed template, not auto-OCR'd) to enable.", {}))

    print("Running cancelled-e-invoice cross-checks...")
    cancelled_by_month = {}
    for m, res_m in zip([r["month"] for r in month_results], month_results):
        if res_m.get("cancelled_einvoices"):
            cancelled_by_month[m] = res_m["cancelled_einvoices"]
    g1_named_by_month = {r["month"]: r.get("g1_named_invnos", set()) for r in month_results}
    all_cancelled, cancel_findings = fchk.build_cancelled_einvoice_findings(
        cancelled_by_month, g1_named_by_month, ewb_out_rows)

    # Cross-reference Doc-Series 'missing' serials against the cancelled-e-invoice list for
    # that same month -- a Table-13-declared serial that's genuinely absent from GSTR-1 AND
    # turns out to be a cancelled e-invoice is explained, not a real gap. Mutates each month's
    # doc_gap list in place, so both write_master_dashboard (below) and write_doc_series pick
    # up the enriched status.
    for res_m in month_results:
        m = res_m["month"]
        invnos_this_month = {c["invno"] for c in cancelled_by_month.get(m, [])}
        fchk.enrich_doc_gap_with_cancelled_einvoices(res_m["doc_gap"], invnos_this_month)

    print("Writing workbook...")
    wb = openpyxl.Workbook()
    ws_dash = wb.active; ws_dash.title = "Master Dashboard"
    write_master_dashboard(ws_dash, month_results, months_covered, months_gap,
                            rect_pairs, annual_review_count, hsn_findings,
                            forensic_findings=forensic_findings, cancel_findings=cancel_findings)
    if run_errors:
        r = ws_dash.max_row + 2
        ws_dash.cell(r, 1, f"NOTE: {len(run_errors)} month(s) failed to process and were skipped "
                            "(their data is NOT included above):").font = Font(bold=True, color="C00000")
        for m, err in run_errors:
            r += 1
            ws_dash.cell(r, 1, f"  {m}: {err}")

    _hsn_rows_cache = {}   # keyed by resolved GSTR-1 file path -- avoids re-parsing the same
                           # merged file once per month when several months share one file
    for res_m in month_results:
        m = res_m["month"]
        raw.PERIOD_LABEL = m
        ws_comp = wb.create_sheet(sheet_name("Comparison", m))
        uni.write_comparison(ws_comp, res_m["comparisons"], only_mismatch=False)
        g1_path_this_month = res["gstr1_month_map"].get(m)
        if g1_path_this_month:
            if g1_path_this_month not in _hsn_rows_cache:
                _hsn_rows_cache[g1_path_this_month] = hfc._hsn_rows_by_month(g1_path_this_month)
            hsn_rows_this_month = _hsn_rows_cache[g1_path_this_month].get(m, [])
        else:
            hsn_rows_this_month = []
        write_hsn_review_table(ws_comp, hsn_rows_this_month, m)
        uni.write_analysis14(wb.create_sheet(sheet_name("Analysis14", m)), res_m["findings14"])
        uni.write_eway(wb.create_sheet(sheet_name("EWB", m)), wb.create_sheet(sheet_name("EWB Detail", m)),
                        res_m["findings27"])

    write_doc_series(wb.create_sheet("Doc-Series Integrity"), month_results)
    write_rectification_sheet(wb.create_sheet("Rectification Pairs"), rect_pairs, annual_data["bo"]["drc_payments"])
    write_hsn_fraud_checks(wb.create_sheet("HSN & Fraud Pattern Checks"), hsn_findings)
    write_filing_compliance(wb.create_sheet("Filing Compliance & Late Fee"), compliance_records)
    write_forensic_checks(wb.create_sheet("Forensic Checks (R13-R14)"), forensic_findings)
    write_cancelled_einvoices(wb.create_sheet("Cancelled E-Invoices"), all_cancelled, cancel_findings,
                               einv_cancel_col_found_any)

    annualwb.write_cover(wb.create_sheet("Annual Cover & Caveats"), annual_data)
    annualwb.write_monthly(wb.create_sheet("Annual Ledger Walkthrough"), annual_rows)
    annualwb.write_fy_total_vs_bifa(wb.create_sheet("FY Total vs BIFA"), annual_rows, annual_data, ewb_out_rows,
                                     fy=(fys_covered[0] if len(fys_covered) == 1 else None))
    annualwb.write_related_party(wb.create_sheet("Related-Party Alerts"), annual_data)
    annualwb.write_top_counterparties(wb.create_sheet("Top Counterparties"), annual_data)

    fy_tag = (fys_covered[0] if len(fys_covered) == 1 else
              f"{fys_covered[0]}_to_{fys_covered[-1]}" if fys_covered else "UNKNOWN_FY")
    outfile = f"GST_MASTER_{res['self_gstin']}_FY{fy_tag}.xlsx"
    wb.save(outfile)
    print(f"\nSaved: {outfile}")
    print(f"Months covered: {months_covered}")
    print(f"Gaps within span: {months_gap}")
    print(f"Rectification pairs: {len(rect_pairs)}")
    print(f"Annual-source REVIEW flags: {annual_review_count}")
    print(f"Cancelled e-invoices found: {len(all_cancelled)}")
    nflag = sum(1 for f in hsn_findings if f.severity == "FLAG")
    nrev = sum(1 for f in hsn_findings if f.severity == "REVIEW")
    print(f"HSN & Fraud Pattern Checks: {len(hsn_findings)} total ({nflag} FLAG, {nrev} REVIEW)")
    print(f"Forensic checks (R13/R14): {[(f.ref, f.severity) for f in forensic_findings]}")
    return outfile


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else ".")
