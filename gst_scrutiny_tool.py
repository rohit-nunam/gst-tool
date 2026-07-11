#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST SCRUTINY COMPARISON TOOL
============================
Plain, raw, side-by-side comparison of a single tax period across:
  GSTR-1  |  GSTR-2B  |  GSTR-3B  |  E-Invoice  |  Cash / Credit / Liability Ledgers

NO analysis. NO interpretation. NO safety nets.
Just: LEFT value | RIGHT value | DIFFERENCE | MATCH? (highlighted if mismatch)

Output: one Excel workbook. Every mismatch row shaded RED.
An "Exceptions" sheet on top lists ONLY the mismatched rows.

USAGE:
    Edit the CONFIG block below (file paths + period), then:
        python gst_scrutiny_tool.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import csv, re, sys
import merged_period_utils as mpu

# ======================================================================
# CONFIG  --  MERGED FILES (whole FY, one workbook per document type)
# ======================================================================
# master_build.py is the single entry point. It classifies the run folder
# once (folder_classifier.py), then before each month's run it sets the
# globals below: the merged-workbook paths stay the SAME across every month
# (there is only one file per type); only PERIOD_LABEL changes, since that's
# what tells the parsers below which month's block to read out of each
# merged file.
#
# No auto-detection and no fallback values live in this module any more --
# if these are wrong, that must be visible immediately, not masked.

GSTR1_FILE   = None   # merged GSTR-1 workbook path (marker rows inside each sub-sheet)
GSTR3B_FILE  = None   # merged GSTR-3B workbook path (one sheet per month)
EINV_FILE    = None   # merged E-Invoice workbook path, or None if not supplied at all
GSTR2B_FILE  = None   # merged GSTR-2B workbook path (monthly or quarterly marker blocks)

SELF_GSTIN   = ""
COMPANY_NAME = ""

PERIOD_LABEL = None   # e.g. 'Jan-23' -- which month to read out of every merged file above

# Filing dates for analysis checks #8 (IRN-lag) and #10 (GSTR-1 vs 3B filing gap).
# Auto-extracted from the ARN date inside the files by the unified tool;
# leave None here (do NOT hardcode).
GSTR1_FILING_DATE  = None
GSTR3B_FILING_DATE = None

OUTPUT_FILE  = "GST_Scrutiny_Comparison.xlsx"

TOLERANCE    = 1.0   # rupee tolerance; abs(diff) <= this  => treated as MATCH

def get_gstr2b_values():
    """Return the GSTR-2B summary dict for PERIOD_LABEL, read straight out of
    the merged GSTR-2B workbook. Requires GSTR2B_FILE + PERIOD_LABEL to be
    set. No hardcoded/zero fallback: if 2B isn't available for this month,
    that must stop the run for this month, not silently zero out the ITC
    comparison rows."""
    import gstr2b_parser as _g2b
    return _g2b.summary_for_month(GSTR2B_FILE, PERIOD_LABEL)

# ======================================================================
# HELPERS
# ======================================================================
def num(v):
    """Convert any cell/string to float. '-', '', None -> 0.0 . Strips commas."""
    if v is None: return 0.0
    s = str(v).strip()
    if s in ("", "-", "–"): return 0.0
    s = s.replace(",", "").replace("₹", "").strip()
    try: return float(s)
    except: return 0.0

def load_xlsx(path):
    return openpyxl.load_workbook(path, data_only=True)

def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))

# ======================================================================
# PARSERS  -- each returns the numbers this tool needs from one source
# ======================================================================

def parse_gstr1(path, month):
    """Sum outward tax & taxable value from GSTR-1 sub-sheets, for ONE month
    out of the merged (whole-FY) workbook. `month` e.g. 'Jan-23'. Every
    sub-sheet in the merged file carries its own period-marker rows (see
    merged_period_utils.py); this reads ONLY the block matching `month` --
    it raises if that month has no marker at all in a given sub-sheet (that
    sub-sheet is simply skipped for scoring only when the SHEET itself is
    entirely absent from the workbook, not when the month is missing from it)."""
    wb = load_xlsx(path)
    out = {"taxable":0.0,"IGST":0.0,"CGST":0.0,"SGST":0.0,"CESS":0.0,
           "b2b_count":0,"b2b_no_irn":0,"lines":{},"blank_invno_lines":0,"blank_invno_taxable":0.0,"named_taxable":0.0,"named_IGST":0.0,"named_CGST":0.0,"named_SGST":0.0,
           "cn_taxable":0.0,"cn_IGST":0.0,"cn_CGST":0.0,"cn_SGST":0.0,"cn_CESS":0.0,
           "hsn_IGST":0.0,"hsn_CGST":0.0,"hsn_SGST":0.0,"hsn_CESS":0.0,"hsn_taxable":0.0,
           "nil_taxable":0.0,"exempt_taxable":0.0,"nongst_taxable":0.0,"nil_exempt_taxable":None}

    # --- b2b ---
    if "b2b, sez, de_inv" in wb.sheetnames:
        ws=wb["b2b, sez, de_inv"]
        rows=list(ws.iter_rows(values_only=True))
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            out["taxable"]+=num(r[H.get("Taxable Value")])
            out["IGST"]+=num(r[H.get("Integrated Tax")])
            out["CGST"]+=num(r[H.get("Central Tax")])
            out["SGST"]+=num(r[H.get("State/UT Tax")])
            out["CESS"]+=num(r[H.get("Cess Amount")])
            out["b2b_count"]+=1
            irn_i=H.get("IRN")
            if irn_i is None or not str(r[irn_i] if irn_i<len(r) else "").strip():
                out["b2b_no_irn"]+=1
            # line-level key for reconciliation
            invno=str(r[H.get("Invoice Number")]).strip() if H.get("Invoice Number") is not None else "None"
            rate=num(r[H.get("Rate")]) if H.get("Rate") is not None else 0.0
            k=(invno,rate)
            L=out["lines"].setdefault(k,[0.0,0.0])
            L[0]+=num(r[H.get("Taxable Value")]); L[1]+=num(r[H.get("Integrated Tax")])
            if not invno or invno.lower()=="none":
                out["blank_invno_lines"]+=1
                out["blank_invno_taxable"]+=num(r[H.get("Taxable Value")])
            else:
                out["named_taxable"]+=num(r[H.get("Taxable Value")])
                out["named_IGST"]+=num(r[H.get("Integrated Tax")])
                out["named_CGST"]+=num(r[H.get("Central Tax")])
                out["named_SGST"]+=num(r[H.get("State/UT Tax")])

    # --- b2cl (inter-state large, IGST only) ---
    for sn,cols in [("b2cl",("Taxable Value","Integrated Tax",None,None,"Cess Amount")),
                    ("exp",("Taxable Value","Integrated Tax",None,None,"Cess Amount"))]:
        if sn in wb.sheetnames:
            rows=list(wb[sn].iter_rows(values_only=True))
            hdr=[str(c).strip() if c else "" for c in rows[3]]
            H={h:i for i,h in enumerate(hdr)}
            for r in mpu.rows_for_month(rows, 3, month):
                if not any(r): continue
                out["taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
                out["IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
                out["CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0

    # --- b2cs (intra/inter small) ---
    if "b2cs" in wb.sheetnames:
        rows=list(wb["b2cs"].iter_rows(values_only=True))
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            out["taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
            out["IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
            out["CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
            out["SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
            out["CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0

    # --- credit notes (cdnr + cdnur) ---
    for sn in ("cdnr","cdnur"):
        if sn in wb.sheetnames:
            rows=list(wb[sn].iter_rows(values_only=True))
            hdr=[str(c).strip() if c else "" for c in rows[3]]
            H={h:i for i,h in enumerate(hdr)}
            for r in mpu.rows_for_month(rows, 3, month):
                if not any(r): continue
                out["cn_taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
                out["cn_IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
                out["cn_CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
                out["cn_SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
                out["cn_CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0

    # --- HSN summary (internal cross-check of GSTR-1) ---
    if "hsn" in wb.sheetnames:
        rows=list(wb["hsn"].iter_rows(values_only=True))
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            out["hsn_taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
            out["hsn_IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
            out["hsn_CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
            out["hsn_SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
            out["hsn_CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0

    # --- Table 8: Nil rated / Exempted / Non-GST outward supplies (sheet 'exemp') ---
    # Header row is found dynamically (its index varies), then scoped to `month`
    # the same way as every other sub-sheet above.
    if "exemp" in wb.sheetnames:
        rows=list(wb["exemp"].iter_rows(values_only=True))
        hdr=None; hi=None
        for i,r in enumerate(rows):
            j=" ".join(str(c) for c in r if c is not None)
            if "Nil Rated" in j and "Non-GST" in j:
                hdr=[str(c).strip() if c else "" for c in r]; hi=i; break
        if hdr is not None:
            H={h:i for i,h in enumerate(hdr)}
            def col(*names):
                for n in names:
                    for h,idx in H.items():
                        if n.lower() in h.lower():
                            return idx
                return None
            i_nil=col("Nil Rated"); i_exm=col("Exempted"); i_non=col("Non-GST")
            for r in mpu.rows_for_month(rows, hi, month):
                if not any(c not in (None,"") for c in r): continue
                if i_nil is not None and i_nil<len(r): out["nil_taxable"]+=num(r[i_nil])
                if i_exm is not None and i_exm<len(r): out["exempt_taxable"]+=num(r[i_exm])
                if i_non is not None and i_non<len(r): out["nongst_taxable"]+=num(r[i_non])
            # header present => figures are known (0 if no rows). Set the combined nil+exempt
            # (matches 3B 3.1(b zero-rated is reported separately; 3.1(c) = nil+exempt).
            out["nil_exempt_taxable"]=out["nil_taxable"]+out["exempt_taxable"]
    return out


def _gstr3b_sheet_month(ws):
    """Read a GSTR-3B sheet's own 'Year' + 'Tax Period' key/value rows and
    return the 'Mon-YY' label they represent. Content-based only -- the
    sheet's NAME (e.g. 'Jan_2022-23') is never trusted, per instruction."""
    fy = tp = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c not in (None, "")]
        if not cells:
            continue
        key = cells[0].upper()
        if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
            fy = cells[1]
        elif key == "TAX PERIOD" and len(cells) >= 2:
            tp = cells[1]
        if fy and tp:
            break
    if not (fy and tp):
        raise mpu.PeriodParseError(
            f"Sheet {ws.title!r} has no readable 'Year'/'Tax Period' fields.")
    labels = mpu.months_for_tax_period(fy, tp)
    return labels[0]  # GSTR-3B's Tax Period is always a single month, never a quarter


def parse_gstr3b(path, month):
    """Pull Table 3.1 and Table 4 values from GSTR-3B for ONE month, out of
    the merged workbook (one sheet per month). The sheet is located by its
    OWN in-sheet 'Year'/'Tax Period' content, not by its sheet name."""
    wb=load_xlsx(path)
    ws = None
    months_found = []
    for sn in wb.sheetnames:
        candidate = wb[sn]
        try:
            m = _gstr3b_sheet_month(candidate)
        except mpu.PeriodParseError:
            continue  # not a GSTR-3B data sheet (e.g. a stray blank/help sheet)
        months_found.append(m)
        if m == month:
            ws = candidate
            break
    if ws is None:
        raise mpu.PeriodParseError(
            f"Month {month!r} not found as a GSTR-3B sheet in {path!r}. "
            f"Months present: {sorted(months_found)}")
    rows=[[c.value for c in r] for r in ws.iter_rows()]
    g={}
    def find(label):
        for r in rows:
            joined=" ".join(str(c) for c in r if c is not None)
            if label in joined:
                vals=[num(c) for c in r if isinstance(c,(int,float)) or (isinstance(c,str) and re.match(r'^-?[\d,\.]+$',str(c).strip()))]
                return r
        return None
    def vals_after(rowlist, n=5):
        nums=[num(c) for c in rowlist if (isinstance(c,(int,float)) or (isinstance(c,str) and re.match(r'^-?[\d,\.]+$',str(c).replace(',','').strip())))]
        return nums

    # ---- Table 4(B) ITC Reversed: anchor to the section boundary ----
    # BUG FIX (confirmed against the real file): the literal label "(2) Others"
    # appears TWICE in every Apr/May/Jun/Jul-22 sheet -- once under "B. ITC
    # Reversed" (the real 4B(2) figure) and once under "(D) Ineligibe ITC"
    # (always 0, a completely different field). A loose "if '(2) Others' in j"
    # scan over every row (no anchoring) picks whichever occurs LAST, which is
    # always the D-section zero -- silently zeroing out 4B(2) for those 4
    # months. From Aug-22 onward the GSTR-3B Table 4 format itself changed
    # (Circular 170/02/2022-GST): the old "(D) Ineligibe ITC -> (1) As per
    # section 17(5)" row is gone, replaced by "(D) Other Details" with
    # unrelated sub-items -- so there's only one "(2) Others" match there and
    # the old loose scan happened to still work for those months by luck, not
    # by design. Fixed properly here: find the "B. ITC Reversed" header row
    # and the "C. Net ITC available" header row, then match "(1)"/"(2) Others"
    # ONLY within that bounded slice -- safe regardless of which Table-4
    # format the sheet uses, and regardless of row order in the file.
    b_start = b_end = None
    for i, r in enumerate(rows):
        j = " ".join(str(c) for c in r if c is not None).strip()
        if j.startswith("B. ITC Reversed"):
            b_start = i
        elif j.startswith("C. Net ITC available") and b_start is not None:
            b_end = i
            break
    if b_start is not None and b_end is not None:
        for r in rows[b_start + 1:b_end]:
            j = " ".join(str(c) for c in r if c is not None).strip()
            nums = vals_after(r)
            if not j or not nums:
                continue
            if j.startswith("(2) Others"):
                g["4B2"] = nums
            elif j.startswith("(1)"):
                g["4B1"] = nums   # Rules 42/43 (Apr-Jul) or Rules 38/42/43 + Sec 17(5) (Aug onward)
    g.setdefault("4B1", [0.0, 0.0, 0.0, 0.0])
    g.setdefault("4B2", [0.0, 0.0, 0.0, 0.0])

    for r in rows:
        j=" ".join(str(c) for c in r if c is not None)
        nums=vals_after(r)
        # 3.1(a) outward taxable
        if "Outward Taxable" in j and "other than zero" in j:
            g["3.1a"]=nums  # [taxable, IGST, CGST, SGST, CESS]
        # 3.1(b) zero-rated, 3.1(c) nil/exempt, 3.1(e) non-GST.
        # Guard against the 3.1(a) line (which also contains 'zero rated, nil rated and exempted'):
        # match the bracketed sub-label, not the words inside 3.1(a)'s 'other than ...' clause.
        if ("(b)" in j or "(zero rated )" in j) and "other than zero" not in j and "zero rated" in j.lower():
            g["3.1b"]=nums  # [taxable, IGST, ?, ?, CESS]
        if "(c)" in j and "Nil rated" in j and "Other Outward" in j:
            g["3.1c"]=nums  # [taxable]
        if "(e)" in j and "Non-GST" in j:
            g["3.1e"]=nums  # [taxable]
        if "Inward supplies (liable to reverse charge)" in j:
            g["3.1d"]=nums
        if "(5) All other ITC" in j:
            g["4A5"]=nums   # [IGST,CGST,SGST,CESS]
        if "(3) Inward supplies liable to reverse charge" in j:
            g["4A3"]=nums
        if "Net ITC available" in j:
            g["4C"]=nums
    return g


def parse_einv(path, month):
    """E-Invoice file totals (B2B), for ONE month out of the merged workbook.
    E-Invoice is legitimately OPTIONAL, at both the whole-file level (some
    taxpayers/periods genuinely have none) AND the per-month level within an
    existing file (e-invoicing coverage can start partway through a file).
    Both cases produce the SAME clearly-surfaced 'not available' state
    (available=False) -- callers already branch on this explicitly, so it is
    not hidden, just not a hard stop for what is a documented PARTIAL source."""
    import os
    out={"taxable":0.0,"IGST":0.0,"CGST":0.0,"SGST":0.0,"CESS":0.0,"count":0,"errors":0,"available":True,"lines":{},
         "cancel_col_found":False,"cancel_date_col_found":False,"cancelled":[]}
    if not path or not os.path.exists(path):
        print(f"[info] E-Invoice file not supplied -> EINV checks skipped for {month}")
        out["available"]=False; return out
    wb=load_xlsx(path)
    if "b2b, sez, de" in wb.sheetnames:
        rows=list(wb["b2b, sez, de"].iter_rows(values_only=True))
        if month not in mpu.months_present(rows, 3):
            print(f"[info] E-Invoice file does not cover {month} -> EINV checks skipped for {month}")
            out["available"]=False; return out
        hdr=[str(c).strip() if c else "" for c in rows[3]]
        H={h:i for i,h in enumerate(hdr)}
        # Cancelled-e-invoice detection: try every known real-world header variant for the
        # IRN status / cancel-date columns (GSTN's own e-invoice/GSTR-1-auto-populate export
        # has used different header text across portal versions). CASE-INSENSITIVE match (fixed:
        # a real export's exact header was 'E-invoice status' -- lowercase 'i'/'s' -- which the
        # original case-sensitive exact-dict-key lookup never matched, even though "E-Invoice
        # Status" was already in the candidate list). Content-based, never a fixed column index.
        # If NONE of these are found, cancel_col_found stays False and the 'Cancelled E-Invoices'
        # sheet says so explicitly rather than reporting zero cancellations as if verified.
        H_LOWER = {h.strip().lower(): i for h, i in H.items()}
        STATUS_HDRS = ["irn status", "status", "e-invoice status", "einvoice status",
                       "cancel status", "invoice status"]
        CANCELDATE_HDRS = ["cancel date", "irn cancel date", "cancelled date",
                           "date of cancellation", "cancellation date"]
        status_col = next((H_LOWER[h] for h in STATUS_HDRS if h in H_LOWER), None)
        canceldate_col = next((H_LOWER[h] for h in CANCELDATE_HDRS if h in H_LOWER), None)
        out["cancel_col_found"] = status_col is not None
        out["cancel_date_col_found"] = canceldate_col is not None
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r): continue
            invno=str(r[H.get("Invoice number")]).strip() if H.get("Invoice number") is not None else "None"
            rate=num(r[H.get("Rate")]) if H.get("Rate") is not None else 0.0

            # FIX (was Bug 1's second half): check cancellation status FIRST. A cancelled
            # e-invoice is correctly ABSENT from GSTR-1 (GSTR-1's own auto-population/deletion
            # status marks it Deleted), so including it in the totals/line-map used for every
            # E-Invoice-vs-GSTR-1 comparison manufactures a false gap on every such invoice --
            # confirmed on the real file: 6 cancelled invoices totalling Rs 44,12,291 taxable
            # were previously producing 6 false "LINE-LEVEL GAP" mismatches, one per month.
            # Cancelled rows are recorded in out['cancelled'] for the Cancelled-E-Invoices sheet
            # and cross-checks, but do NOT contribute to taxable/IGST/CGST/SGST/CESS/count/lines.
            is_cancelled = False
            if status_col is not None and status_col < len(r):
                status_val = str(r[status_col] or "").strip().upper()
                if status_val in ("CANCELLED", "CANCEL", "CANCELED"):
                    is_cancelled = True
                    out["cancelled"].append(dict(
                        invno=invno, rate=rate,
                        taxable=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0.0,
                        igst=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0.0,
                        irn=str(r[H.get("IRN")] or "").strip() if H.get("IRN") is not None else "",
                        cancel_date=(str(r[canceldate_col]).strip()
                                     if canceldate_col is not None and canceldate_col < len(r) else None),
                        month=month,
                    ))
            if is_cancelled:
                continue

            out["taxable"]+=num(r[H.get("Taxable Value")]) if "Taxable Value" in H else 0
            out["IGST"]+=num(r[H.get("Integrated Tax")]) if "Integrated Tax" in H else 0
            out["CGST"]+=num(r[H.get("Central Tax")]) if "Central Tax" in H else 0
            out["SGST"]+=num(r[H.get("State/UT Tax")]) if "State/UT Tax" in H else 0
            out["CESS"]+=num(r[H.get("Cess Amount")]) if "Cess Amount" in H else 0
            out["count"]+=1
            k=(invno,rate)
            L=out["lines"].setdefault(k,[0.0,0.0])
            L[0]+=num(r[H.get("Taxable Value")]); L[1]+=num(r[H.get("Integrated Tax")])
            ei=H.get("Error in auto-population/ deletion")
            if ei is not None and ei<len(r) and str(r[ei] or "").strip():
                out["errors"]+=1
    return out


def parse_ledger_csv(path, period_only=False):
    """Generic ledger reader: returns list of dict rows with flattened heads.
       We only extract what's needed: period-wise debit/credit per head + RCM/cash markers."""
    rows=read_csv_rows(path)
    return rows  # raw; consumed by specific extractors below


# ======================================================================
# BUILD COMPARISONS
# ======================================================================
def build_comparisons():
    if not PERIOD_LABEL:
        raise ValueError("PERIOD_LABEL is not set -- caller must set raw.PERIOD_LABEL "
                          "before calling build_comparisons().")
    g1   = parse_gstr1(GSTR1_FILE, PERIOD_LABEL)
    g3b  = parse_gstr3b(GSTR3B_FILE, PERIOD_LABEL)
    einv = parse_einv(EINV_FILE, PERIOD_LABEL)
    b2b  = get_gstr2b_values()

    # g3b lists: 3.1a = [taxable,IGST,CGST,SGST,CESS]; 4A5=[IGST,CGST,SGST,CESS]
    def gv(key,i,default=0.0):
        v=g3b.get(key)
        return v[i] if v and i<len(v) else default

    C=[]  # each row: (section, check, left_label, left_val, right_label, right_val, tag)
    def add(section,check,llabel,lval,rlabel,rval,tag=""):
        C.append((section,check,llabel,lval,rlabel,rval,tag))

    # ---- A. OUTWARD: GSTR-1 (NET of credit notes) vs GSTR-3B 3.1(a) ----
    # GSTR-3B 3.1(a) is reported NET of credit notes, so net off GSTR-1 CN too.
    C.append(("A. Outward Liability","Outward taxable value (net of CN)",
              "GSTR-1 net", g1["taxable"]-g1["cn_taxable"], "GSTR-3B 3.1(a)", gv("3.1a",0)))
    C.append(("A. Outward Liability","Outward IGST (net of CN)",
              "GSTR-1 net", g1["IGST"]-g1["cn_IGST"], "GSTR-3B 3.1(a)", gv("3.1a",1)))
    C.append(("A. Outward Liability","Outward CGST (net of CN)",
              "GSTR-1 net", g1["CGST"]-g1["cn_CGST"], "GSTR-3B 3.1(a)", gv("3.1a",2)))
    C.append(("A. Outward Liability","Outward SGST (net of CN)",
              "GSTR-1 net", g1["SGST"]-g1["cn_SGST"], "GSTR-3B 3.1(a)", gv("3.1a",3)))
    C.append(("A. Outward Liability","Outward CESS (net of CN)",
              "GSTR-1 net", g1["CESS"]-g1["cn_CESS"], "GSTR-3B 3.1(a)", gv("3.1a",4)))

    # ---- A1b. GROSS outward (before CN) for reference ----
    C.append(("A. Outward Liability","Outward taxable GROSS (before CN)",
              "GSTR-1 gross", g1["taxable"], "GSTR-1 HSN", g1["hsn_taxable"]))

    # ---- A2. GSTR-1 internal: invoice-level (net of CN) vs HSN summary ----
    C.append(("A2. GSTR-1 internal","Taxable value (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["taxable"]-g1["cn_taxable"], "GSTR-1 HSN", g1["hsn_taxable"]))
    C.append(("A2. GSTR-1 internal","IGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["IGST"]-g1["cn_IGST"], "GSTR-1 HSN", g1["hsn_IGST"]))
    C.append(("A2. GSTR-1 internal","CGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["CGST"]-g1["cn_CGST"], "GSTR-1 HSN", g1["hsn_CGST"]))
    C.append(("A2. GSTR-1 internal","SGST (invoices-net vs HSN)",
              "GSTR-1 inv net", g1["SGST"]-g1["cn_SGST"], "GSTR-1 HSN", g1["hsn_SGST"]))

    # ---- B. E-INVOICE vs GSTR-1 B2B ----
    # GSTR-1 side uses NAMED-invoice sum (blank-invoice-no lines excluded), so the
    # summary stays consistent with the line-level section B2. The orphan blank line
    # surfaces as the difference and is detailed in B2.
    if einv.get("available"):
        note_b2 = "Excludes GSTR-1 blank-invoice-no line(s); see section B2 for line-level detail." if g1.get("blank_invno_lines",0)>0 else ""
        C.append(("B. E-Invoice vs GSTR-1","B2B taxable value",
                  "E-Invoice", einv["taxable"], "GSTR-1 B2B (named)", g1["named_taxable"], note_b2))
        C.append(("B. E-Invoice vs GSTR-1","B2B IGST",
                  "E-Invoice", einv["IGST"], "GSTR-1 B2B (named)", g1["named_IGST"], note_b2))
        note_cs = "Gap in B2B taxable/IGST is an IGST-only (inter-state) line; CGST/SGST unaffected, hence MATCH here." if g1.get("blank_invno_lines",0)>0 else ""
        C.append(("B. E-Invoice vs GSTR-1","B2B CGST",
                  "E-Invoice", einv["CGST"], "GSTR-1 B2B (named)", g1["named_CGST"], note_cs))
        C.append(("B. E-Invoice vs GSTR-1","B2B SGST",
                  "E-Invoice", einv["SGST"], "GSTR-1 B2B (named)", g1["named_SGST"], note_cs))
        C.append(("B. E-Invoice vs GSTR-1","B2B unique invoice count",
                  "E-Invoice", len(einv.get("lines",{})) and len(set(k[0] for k in einv["lines"])), "GSTR-1 B2B named-inv count", g1["b2b_count"]-g1.get("blank_invno_lines",0)))
        C.append(("B. E-Invoice vs GSTR-1","B2B invoices WITHOUT IRN (should be 0)",
                  "Flag", g1["b2b_no_irn"], "Target", 0))

    # ---- B2. E-INVOICE vs GSTR-1 B2B  (LINE-LEVEL, catches total-match hiding line gaps) ----
    if einv.get("available"):
        g1L=g1.get("lines",{}); eiL=einv.get("lines",{})
        allk=set(g1L)|set(eiL)
        line_mismatch=0
        for k in sorted(allk, key=lambda x:(str(x[0]),x[1])):
            a=g1L.get(k,[0.0,0.0]); b=eiL.get(k,[0.0,0.0])
            if abs(a[0]-b[0])>TOLERANCE or abs(a[1]-b[1])>TOLERANCE:
                line_mismatch+=1
                inv,rate=k
                C.append(("B2. E-Inv vs GSTR-1 (line-level)",
                          f"Invoice {inv} @ {rate}% - taxable",
                          "GSTR-1 line", a[0], "E-Invoice line", b[0],
                          "LINE-LEVEL GAP - present in one source/rate-line not the other; verify invoice."))
        # blank invoice-number lines in GSTR-1 (orphan taxable lines)
        C.append(("B2. E-Inv vs GSTR-1 (line-level)",
                  "GSTR-1 taxable lines with BLANK invoice no (should be 0)",
                  "GSTR-1 blank-invno lines", g1.get("blank_invno_lines",0), "Target", 0,
                  "DATA INTEGRITY - taxable value sitting on a line with no invoice number." if g1.get("blank_invno_lines",0)>0 else ""))

    # ---- C. RCM: GSTR-3B 3.1(d) vs GSTR-2B available ----
    # GRACEFUL DEGRADATION (fixed): b2b.get('available') is False when GSTR-2B was not
    # supplied for this month (gstr2b_parser.summary_for_month() -- see its docstring).
    # Previously this section unconditionally indexed b2b['ITC_rcm_IGST'] etc, which either
    # raised (crashing the whole month) or, if 2B fields were zero-filled instead, would have
    # produced a wall of false MISMATCH rows (GSTR-3B's real RCM figure vs a fake zero) that
    # look like genuine scrutiny findings but are really just "no data was available to check".
    if not b2b.get("available"):
        C.append(("C. RCM", "RCM liability vs GSTR-2B", "GSTR-3B 3.1(d)", gv("3.1d", 1),
                  "GSTR-2B", None,
                  f"SKIPPED -- GSTR-2B not supplied for this month "
                  f"({b2b.get('_reason', 'no reason recorded')}). RCM/ITC checks C, D, D2 all "
                  "skipped for this month; this is a data-availability gap, not a mismatch."))
    else:
        C.append(("C. RCM","RCM liability IGST (3.1d vs 2B-avail)",
                  "GSTR-3B 3.1(d)", gv("3.1d",1), "GSTR-2B RCM", b2b["ITC_rcm_IGST"]))
        C.append(("C. RCM","RCM liability CGST",
                  "GSTR-3B 3.1(d)", gv("3.1d",2), "GSTR-2B RCM", b2b["ITC_rcm_CGST"],
                  "SCOPE DIFF - 3.1(d) includes unreg/import-of-service RCM; 2B shows only registered-supplier RCM. Expected."))
        C.append(("C. RCM","RCM liability SGST",
                  "GSTR-3B 3.1(d)", gv("3.1d",3), "GSTR-2B RCM", b2b["ITC_rcm_SGST"],
                  "SCOPE DIFF - see RCM CGST note. Expected."))
        C.append(("C. RCM","RCM ITC claimed IGST (4A3 vs 2B)",
                  "GSTR-3B 4(A)(3)", gv("4A3",0), "GSTR-2B RCM", b2b["ITC_rcm_IGST"]))

    # ---- D. ITC: GSTR-3B 4(A)(5) vs GSTR-2B (net of credit notes) ----
    if b2b.get("available"):
        net2b_igst = b2b["ITC_all_other_IGST"] - b2b["CN_IGST"]
        net2b_cgst = b2b["ITC_all_other_CGST"] - b2b["CN_CGST"]
        net2b_sgst = b2b["ITC_all_other_SGST"] - b2b["CN_SGST"]
        C.append(("D. ITC (All other)","ITC IGST (3B 4A5 vs 2B gross)",
                  "GSTR-3B 4(A)(5)", gv("4A5",0), "GSTR-2B (gross)", b2b["ITC_all_other_IGST"],
                  "TO BE EXPLAINED - gap may be prev-period carryforward/provisional ITC; not auto-ineligible."))
        C.append(("D. ITC (All other)","ITC CGST (3B 4A5 vs 2B gross)",
                  "GSTR-3B 4(A)(5)", gv("4A5",1), "GSTR-2B (gross)", b2b["ITC_all_other_CGST"]))
        C.append(("D. ITC (All other)","ITC SGST (3B 4A5 vs 2B gross)",
                  "GSTR-3B 4(A)(5)", gv("4A5",2), "GSTR-2B (gross)", b2b["ITC_all_other_SGST"]))
        C.append(("D. ITC (All other)","ITC IGST (3B 4A5 vs 2B NET of CN)",
                  "GSTR-3B 4(A)(5)", gv("4A5",0), "GSTR-2B (net CN)", net2b_igst))
        C.append(("D. ITC (All other)","ITC CGST (3B 4A5 vs 2B NET of CN)",
                  "GSTR-3B 4(A)(5)", gv("4A5",1), "GSTR-2B (net CN)", net2b_cgst))
        C.append(("D. ITC (All other)","ITC SGST (3B 4A5 vs 2B NET of CN)",
                  "GSTR-3B 4(A)(5)", gv("4A5",2), "GSTR-2B (net CN)", net2b_sgst))

        # ---- D2. ITC reversal: 3B 4(B)(2) vs 2B credit notes ----
        C.append(("D2. ITC Reversal","Reversal IGST (3B 4B2 vs 2B CN)",
                  "GSTR-3B 4(B)(2)", gv("4B2",0), "GSTR-2B CN", b2b["CN_IGST"]))
        C.append(("D2. ITC Reversal","Reversal CGST (3B 4B2 vs 2B CN)",
                  "GSTR-3B 4(B)(2)", gv("4B2",1), "GSTR-2B CN", b2b["CN_CGST"]))
        C.append(("D2. ITC Reversal","Reversal SGST (3B 4B2 vs 2B CN)",
                  "GSTR-3B 4(B)(2)", gv("4B2",2), "GSTR-2B CN", b2b["CN_SGST"]))
    else:
        C.append(("D. ITC (All other)", "ITC vs GSTR-2B", "GSTR-3B 4(A)(5)", gv("4A5", 0),
                  "GSTR-2B", None, "SKIPPED -- GSTR-2B not supplied for this month (see section C note above)."))
        C.append(("D2. ITC Reversal", "Reversal vs GSTR-2B CN", "GSTR-3B 4(B)(2)", gv("4B2", 0),
                  "GSTR-2B", None, "SKIPPED -- GSTR-2B not supplied for this month (see section C note above)."))

    return C, dict(g1=g1, g3b=g3b, einv=einv, b2b=b2b)


# ======================================================================
# WRITE EXCEL
# ======================================================================
RED   = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
HEAD  = PatternFill("solid", fgColor="1F3864")
SECT  = PatternFill("solid", fgColor="D9E1F2")
BORDER= Border(*[Side(style="thin", color="BFBFBF")]*4)

def style_header(ws, row, ncols):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row, column=c)
        cell.fill=HEAD; cell.font=Font(bold=True, color="FFFFFF", size=10)
        cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border=BORDER

def write_rows(ws, start, comparisons, only_mismatch=False):
    r=start
    cur_sect=None
    for row in comparisons:
        if len(row)==7:
            sect, check, llabel, lval, rlabel, rval, tag = row
        else:
            sect, check, llabel, lval, rlabel, rval = row; tag=""
        diff = round(num(lval)-num(rval), 2)
        is_match = abs(diff) <= TOLERANCE
        if only_mismatch and is_match:
            continue
        if sect != cur_sect and not only_mismatch:
            cur_sect=sect
            ws.cell(row=r, column=1, value=sect).font=Font(bold=True, size=11, color="1F3864")
            for c in range(1,10): ws.cell(row=r,column=c).fill=SECT
            r+=1
        ws.cell(row=r, column=1, value=sect if only_mismatch else "")
        ws.cell(row=r, column=2, value=check)
        ws.cell(row=r, column=3, value=llabel)
        ws.cell(row=r, column=4, value=round(num(lval),2))
        ws.cell(row=r, column=5, value=rlabel)
        ws.cell(row=r, column=6, value=round(num(rval),2))
        ws.cell(row=r, column=7, value=diff)
        ws.cell(row=r, column=8, value="MATCH" if is_match else "MISMATCH")
        ws.cell(row=r, column=9, value=tag)
        fill = GREEN if is_match else RED
        for c in range(1,10):
            cell=ws.cell(row=r,column=c)
            cell.border=BORDER
            cell.font=Font(size=10)
            if c in (4,6,7): cell.number_format='#,##0.00'
            if c==8:
                cell.fill=fill; cell.font=Font(bold=True, size=10)
                cell.alignment=Alignment(horizontal="center")
            elif not is_match and c>=2:
                cell.fill=RED
        r+=1
    return r

def main():
    comparisons, raw = build_comparisons()

    wb=openpyxl.Workbook()

    # ---- Sheet 1: EXCEPTIONS (mismatches only) ----
    ws=wb.active; ws.title="Exceptions"
    ws.cell(row=1,column=1,value=f"GST SCRUTINY  -  MISMATCHES ONLY  -  Period: {PERIOD_LABEL}").font=Font(bold=True,size=13,color="C00000")
    ws.cell(row=2,column=1,value=f"GSTIN {SELF_GSTIN}  |  {COMPANY_NAME or '(company auto-detected)'}  |  Tolerance: Rs {TOLERANCE}").font=Font(size=9,italic=True)
    hdr=["Section","Check","Left source","Left value","Right source","Right value","Difference","Result","Note / Tag"]
    for i,h in enumerate(hdr,1): ws.cell(row=4,column=i,value=h)
    style_header(ws,4,9)
    end=write_rows(ws,5,comparisons,only_mismatch=True)
    if end==5:
        ws.cell(row=5,column=1,value="No mismatches beyond tolerance.").font=Font(italic=True,color="006100")
    widths=[26,46,16,15,16,15,14,11,55]
    for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w

    # ---- Sheet 2: FULL COMPARISON ----
    ws2=wb.create_sheet("Full Comparison")
    ws2.cell(row=1,column=1,value=f"GST SCRUTINY  -  FULL COMPARISON  -  Period: {PERIOD_LABEL}").font=Font(bold=True,size=13,color="1F3864")
    for i,h in enumerate(hdr,1): ws2.cell(row=3,column=i,value=h)
    style_header(ws2,3,9)
    write_rows(ws2,4,comparisons,only_mismatch=False)
    for i,w in enumerate(widths,1): ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws2.freeze_panes="A4"

    # ---- Sheet 3: RAW EXTRACTED VALUES (audit trail) ----
    ws3=wb.create_sheet("Raw Values")
    ws3.cell(row=1,column=1,value="RAW EXTRACTED VALUES (audit trail - what the tool read from each file)").font=Font(bold=True,size=11)
    r=3
    for src,d in [("GSTR-1",raw["g1"]),("E-Invoice",raw["einv"]),("GSTR-2B (manual from PDF)",raw["b2b"])]:
        ws3.cell(row=r,column=1,value=src).font=Font(bold=True,color="1F3864"); r+=1
        for k,v in d.items():
            if isinstance(v,dict):
                # 'lines' dict -> show line count + unique invoice count
                uniq=len(set(kk[0] for kk in v.keys())) if v else 0
                ws3.cell(row=r,column=2,value=k+" (line count)")
                ws3.cell(row=r,column=3,value=len(v))
                r+=1
                ws3.cell(row=r,column=2,value=k+" (unique invoices)")
                ws3.cell(row=r,column=3,value=uniq)
                r+=1
                continue
            if isinstance(v,bool):
                ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=str(v)); r+=1
                continue
            ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=round(num(v),2))
            ws3.cell(row=r,column=3).number_format='#,##0.00'; r+=1
        r+=1
    ws3.cell(row=r,column=1,value="GSTR-3B (parsed tables)").font=Font(bold=True,color="1F3864"); r+=1
    for k,v in raw["g3b"].items():
        ws3.cell(row=r,column=2,value=k); ws3.cell(row=r,column=3,value=str(v)); r+=1
    ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=34; ws3.column_dimensions["C"].width=20

    wb.save(OUTPUT_FILE)
    n_mismatch=sum(1 for row in comparisons if abs(num(row[3])-num(row[5]))>TOLERANCE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Total checks: {len(comparisons)}  |  Mismatches: {n_mismatch}")

if __name__=="__main__":
    main()
