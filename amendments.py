#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AMENDMENTS + DOC-SERIES  (new checks enabled by fields the earlier single-
month tool never read: GSTR-1's b2ba/cdnra/b2csa/expa amendment sheets, and
the 'docs' sheet = Table 13, Summary of Documents Issued).

A GSTR-1 filed in month N can contain amendment rows that correct an invoice
originally reported in an EARLIER month M (via 'Original Invoice Number' /
'Original Invoice date'). This is the direct evidence for the "error in an
earlier month, corrected later" requirement -- read these sheets across every
month you have and match Original -> that original month's B2B line.
"""

import openpyxl
import merged_period_utils as mpu


def _num(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def _hdr_row_idx(rows, must_contain):
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in r]
        if all(any(m.lower() in c.lower() for c in cells) for m in must_contain):
            return i
    return None


def parse_b2ba(path, month):
    """9A amendment sheet: corrections to B2B invoices reported in an earlier
    period, scoped to ONE month's block out of the merged workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "b2ba" not in wb.sheetnames:
        return []
    rows = list(wb["b2ba"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Original Invoice Number", "Revised Invoice Number"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not r[H.get("Original Invoice Number", 0)]:
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        out.append(dict(
            gstin=str(g("GSTIN/UIN of Recipient") or "").strip(),
            recipient=str(g("Receiver Name") or "").strip(),
            orig_invno=str(g("Original Invoice Number") or "").strip(),
            orig_date=g("Original Invoice date"),
            revised_invno=str(g("Revised Invoice Number") or "").strip(),
            revised_date=g("Revised Invoice date"),
            invval=_num(g("Invoice Value")), pos=str(g("Place Of Supply") or "").strip(),
            rate=_num(g("Rate")), taxable=_num(g("Taxable Value")),
            igst=_num(g("Integrated Tax")), cgst=_num(g("Central Tax")),
            sgst=_num(g("State/UT Tax")), cess=_num(g("Cess Amount")),
        ))
    return out


def parse_cdnra(path, month):
    """9C amendment sheet: corrections to credit/debit notes reported earlier,
    scoped to ONE month's block out of the merged workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "cdnra" not in wb.sheetnames:
        return []
    rows = list(wb["cdnra"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Original Note Number", "Revised Note Number"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not r[H.get("Original Note Number", 0)]:
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        out.append(dict(
            gstin=str(g("GSTIN/UIN of Recipient") or "").strip(),
            orig_noteno=str(g("Original Note Number") or "").strip(),
            orig_date=g("Original Note Date"),
            revised_noteno=str(g("Revised Note Number") or "").strip(),
            revised_date=g("Revised Note Date"),
            note_type=str(g("Note Type") or "").strip(),
            taxable=_num(g("Taxable Value")), igst=_num(g("Integrated Tax")),
            cgst=_num(g("Central Tax")), sgst=_num(g("State/UT Tax")),
        ))
    return out


def parse_docs(path, month):
    """Table 13: Summary of Documents Issued, scoped to ONE month's block out
    of the merged workbook. Returns list of dicts for gap analysis against B2B."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "docs" not in wb.sheetnames:
        return []
    rows = list(wb["docs"].iter_rows(values_only=True))
    hi = _hdr_row_idx(rows, ["Sr. No. From", "Sr. No. To"])
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in mpu.rows_for_month(rows, hi, month):
        if not r or not r[H.get("Nature of Document", 0)]:
            continue
        g = lambda k: r[H[k]] if k in H and H[k] < len(r) else None
        out.append(dict(
            nature=str(g("Nature of Document") or "").strip(),
            sr_from=str(g("Sr. No. From") or "").strip(),
            sr_to=str(g("Sr. No. To") or "").strip(),
            total=_num(g("Total Number")), cancelled=_num(g("Cancelled")),
        ))
    return out


def _split_series(invno):
    """'MR22-23/509' -> ('MR22-23/', 509). Returns (None, None) if not numeric-suffixed."""
    import re
    m = re.match(r"^(.*?)(\d+)$", invno.strip())
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def _normalize_prefix(p):
    """Strip slashes/hyphens/spaces and uppercase, for punctuation-tolerant prefix matching.
    'JWI/22-23/' and 'JWI22-23/' both normalize to 'JWI2223' -- confirmed necessary against
    the real file, where Table 13's own printed range header and the actual invoice numbers
    for the SAME series don't always agree on punctuation (this is a real inconsistency in
    the source export, not a code bug to silently paper over -- see doc_series_gap_check)."""
    import re
    return re.sub(r"[^A-Za-z0-9]", "", p or "").upper()


def doc_series_gap_check(docs, actual_invoice_numbers):
    """For each 'Invoices for outward supply' Sr-No range in Table 13, find
    which serials in that range are MISSING from the actual B2B invoice list
    (and not accounted for by the 'Cancelled' count).

    Matching is two-tier, because the real source file is not internally
    consistent about series-prefix punctuation:
      Tier 1 -- exact prefix match (fast path, works for most series).
      Tier 2 -- punctuation-normalized prefix match (handles 'JWI/22-23/'
                declared vs 'JWI22-23/' actually used -- same series).
      Tier 3 -- fuzzy fallback: does any actual invoice's normalized text
                CONTAIN the declared series' normalized prefix, with the
                right trailing number? (handles 'MR/JWI/22-23/001' actually
                used for a series Table 13 declares as 'JWI/22-23/001' --
                confirmed real, in the April data.) Numbers resolved only at
                this tier are marked so the difference stays visible rather
                than silently treated as an exact match.

    A number still unresolved after all three tiers is genuinely reported as
    missing UNLESS the count of such numbers exactly equals Table 13's own
    declared 'Cancelled' count for that range, in which case it's reported as
    explained (not a real gap) -- still shown, not hidden, just not painted
    as an open question when the source document already accounts for it.
    """
    findings = []
    actual_by_prefix = {}      # exact prefix -> {numbers}
    actual_by_norm_prefix = {} # normalized prefix -> {numbers}
    actual_norm_full = []      # [(normalized full invno, number)] for the tier-3 fallback
    for inv in actual_invoice_numbers:
        prefix, num = _split_series(inv)
        if prefix is None:
            continue
        actual_by_prefix.setdefault(prefix, set()).add(num)
        actual_by_norm_prefix.setdefault(_normalize_prefix(prefix), set()).add(num)
        actual_norm_full.append((_normalize_prefix(inv), num))

    for d in docs:
        if d["nature"] != "Invoices for outward supply":
            continue
        p_from, n_from = _split_series(d["sr_from"])
        p_to, n_to = _split_series(d["sr_to"])
        if p_from is None or p_to is None or p_from != p_to:
            findings.append(dict(range=f"{d['sr_from']} - {d['sr_to']}", missing=[],
                                  note="Could not parse series prefix/number -- check manually"))
            continue
        expected = set(range(n_from, n_to + 1))
        norm_prefix = _normalize_prefix(p_from)

        exact_have = actual_by_prefix.get(p_from, set())
        norm_have = actual_by_norm_prefix.get(norm_prefix, set())
        still_missing = expected - exact_have - norm_have
        fuzzy_found = set()
        for norm_inv, num in actual_norm_full:
            if num in still_missing and norm_prefix in norm_inv:
                fuzzy_found.add(num)
        missing_nums = sorted(still_missing - fuzzy_found)
        found_via_fuzzy = sorted(fuzzy_found)

        cancelled = d["cancelled"]
        explained_by_cancellation = bool(missing_nums) and len(missing_nums) == cancelled
        findings.append(dict(
            range=f"{d['sr_from']} - {d['sr_to']}", prefix=p_from,
            table13_total=d["total"], table13_cancelled=cancelled,
            actual_count=len(expected) - len(missing_nums),
            missing=[f"{p_from}{n}" for n in missing_nums],
            found_via_fuzzy_match=[f"{p_from}{n}" for n in found_via_fuzzy],
            explained_by_declared_cancellation=explained_by_cancellation,
        ))
    return findings


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        raise SystemExit("Usage: python amendments.py <GSTR1_Merged.xlsx> <month, e.g. Jan-23>")
    p, month = sys.argv[1], sys.argv[2]
    b2ba = parse_b2ba(p, month)
    cdnra = parse_cdnra(p, month)
    docs = parse_docs(p, month)
    print("B2BA amendment rows:", len(b2ba))
    for x in b2ba[:5]:
        print(" ", x)
    print("CDNRA amendment rows:", len(cdnra))
    print("\nDocs (Table 13):")
    for d in docs:
        print(" ", d)
