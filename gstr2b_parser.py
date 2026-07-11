#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GSTR-2B PARSER  (Excel)  --  shared by the scrutiny / analysis / e-way-bill tools
=================================================================================
Now reads the MERGED, whole-FY GSTR-2B workbook (quarterly blocks stacked in
one file, per merged_period_utils.py's marker convention):
  Sheet 'ITC Available' : the Table-3 summary, scattered across several small
                           tables per quarter block -- located by QUARTER
                           marker (find_block_for_month), since this sheet has
                           no single line-level month tag.
  Sheet 'B2B'            : invoice-level inward supplies -- each row carries
                           its OWN exact month in the "GSTR-1/IFF/GSTR-1A/
                           GSTR-5 Period" column (e.g. "May'22"), which is
                           more precise than the quarter marker and is used
                           directly, in preference to the marker.
  Sheet 'B2B-CDNR'       : supplier credit/debit notes -- same per-row period
                           column approach as B2B.

KNOWN, DELIBERATELY-UNCHANGED LIMITATION in the summary extraction below:
the 'ITC Available' table's rows contain FOUR column-groups back to back
(Month1, Month2, Month3, Total-for-quarter), and the extraction here takes
the FIRST group of numbers found in a matching row -- i.e. for any month in
a quarter it reads Month-1-of-that-quarter's IGST/CGST/SGST/CESS, not the
Total or the specific requested month's own column-group. This was already
true before the merge (single-quarter files) and is being left exactly as-is
here per explicit instruction -- only the QUARTER-block scoping is new (so a
merged whole-year file no longer silently reads whichever quarter happens to
be scanned last).
"""

import os, glob, re
import openpyxl
import merged_period_utils as mpu

def find_2b_excel(path, search_dir="."):
    """Return a usable 2B Excel path.
    1) if `path` exists and is .xlsx/.xlsm -> use it;
    2) else scan `search_dir` for a GSTR-2B Excel by filename pattern (handles the portal's
       long auto-generated names like '..._GSTR2B_....xlsx'). Returns None if none found."""
    if path and os.path.exists(path) and path.lower().endswith((".xlsx", ".xlsm")):
        return path
    cands = []
    for f in glob.glob(os.path.join(search_dir, "*.xlsx")) + glob.glob(os.path.join(search_dir, "*.xlsm")):
        name = os.path.basename(f).upper()
        # must look like a 2B file, must NOT be one of our own outputs or other returns
        if re.search(r"GSTR\s*[-_ ]?2B|[_\- ]2B[_\- ]", name) and "SCRUTINY" not in name \
           and "EWAYBILL" not in name and "EWAY" not in name and "COMPARISON" not in name \
           and "ANALYSIS" not in name:
            cands.append(f)
    if not cands:
        return None
    # prefer the one whose name also contains the GSTIN/period if multiple; else newest
    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return cands[0]


def _num(v):
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", "").replace("₹", "")
    if s in ("", "-", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _data_rows(ws, first_data_row=6):
    rows = list(ws.iter_rows(values_only=True))
    return rows[first_data_row:]


def _normalize_2b_row_period(tag):
    """"May'22" -> 'May-22'. Raises PeriodParseError if the tag can't be read --
    every real B2B/B2B-CDNR data row carries this column filled in, so an
    unparseable value here means something about the file has changed."""
    s = str(tag or "").strip()
    m = re.match(r"^([A-Za-z]{3})'(\d{2})$", s)
    if not m:
        raise mpu.PeriodParseError(f"Unrecognised per-row GSTR-2B period tag: {tag!r}")
    return f"{m.group(1).title()}-{m.group(2)}"


def _summary_from_block(rows):
    """Run the (deliberately unchanged) label-match + first-numeric-group
    extraction over an already quarter-scoped slice of 'ITC Available' rows."""
    summary = dict(
        ITC_all_other_IGST=0.0, ITC_all_other_CGST=0.0, ITC_all_other_SGST=0.0, ITC_all_other_CESS=0.0,
        ITC_rcm_IGST=0.0, ITC_rcm_CGST=0.0, ITC_rcm_SGST=0.0, ITC_rcm_CESS=0.0,
        CN_IGST=0.0, CN_CGST=0.0, CN_SGST=0.0, CN_CESS=0.0,
    )
    for r in rows:
        joined = " ".join(str(c) for c in r if c is not None)
        cells = list(r)
        if "All other ITC" in joined and "reverse charge" in joined:
            n = [_num(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                summary["ITC_all_other_IGST"], summary["ITC_all_other_CGST"], \
                    summary["ITC_all_other_SGST"] = n[0], n[1], n[2]
                summary["ITC_all_other_CESS"] = n[3] if len(n) > 3 else 0.0
        elif "reverse charge" in joined and "3.1(d)" in joined and "Net input" in joined:
            n = [_num(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                summary["ITC_rcm_IGST"], summary["ITC_rcm_CGST"], summary["ITC_rcm_SGST"] = n[0], n[1], n[2]
                summary["ITC_rcm_CESS"] = n[3] if len(n) > 3 else 0.0
        elif joined.strip().startswith("I ") and "Others" in joined and "4(A)" in joined:
            n = [_num(c) for c in cells if isinstance(c, (int, float))]
            if len(n) >= 3:
                summary["CN_IGST"], summary["CN_CGST"], summary["CN_SGST"] = n[0], n[1], n[2]
                summary["CN_CESS"] = n[3] if len(n) > 3 else 0.0
    return summary


def parse_2b_excel(path, month):
    """Return dict(summary=..., b2b=[...], cdnr=[...], available=True) for ONE
    month out of the merged (whole-FY) GSTR-2B workbook."""
    if not path or not os.path.exists(path) or not path.lower().endswith((".xlsx", ".xlsm")):
        raise mpu.PeriodParseError(f"Not a GSTR-2B Excel file: {path!r}")
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---------- Summary (Table 3), quarter-block-scoped ----------
    if "ITC Available" not in wb.sheetnames:
        raise mpu.PeriodParseError(f"'ITC Available' sheet not found in {path!r}")
    ws = wb["ITC Available"]
    rows = list(ws.iter_rows(values_only=True))
    start, end = mpu.find_block_for_month(rows, month)
    summary = _summary_from_block(rows[start:end])

    # ---------- B2B invoice list, filtered by each row's OWN period column ----------
    b2b = []
    if "B2B" in wb.sheetnames:
        for r in _data_rows(wb["B2B"]):
            if not any(r) or not r[0] or mpu.is_marker_row(r):
                continue
            if _normalize_2b_row_period(r[14]) != month:
                continue
            b2b.append(dict(
                gstin=str(r[0]).strip(), supplier=str(r[1] or "").strip(),
                invno=str(r[2] or "").strip(), invtype=str(r[3] or "").strip(),
                date=str(r[4] or "").strip(), invval=_num(r[5]),
                pos=str(r[6] or "").strip(), rcm=str(r[7] or "").strip(),
                rate=_num(r[8]), taxable=_num(r[9]),
                igst=_num(r[10]), cgst=_num(r[11]), sgst=_num(r[12]), cess=_num(r[13]),
                itc_avail=str(r[16] or "").strip() if len(r) > 16 else "",
                itc_avail_reason=str(r[17] or "").strip() if len(r) > 17 else "",
            ))

    # ---------- B2B-CDNR (credit/debit notes), same per-row period filtering ----------
    cdnr = []
    if "B2B-CDNR" in wb.sheetnames:
        for r in _data_rows(wb["B2B-CDNR"]):
            if not any(r) or not r[0] or mpu.is_marker_row(r):
                continue
            if _normalize_2b_row_period(r[15]) != month:
                continue
            cdnr.append(dict(
                gstin=str(r[0]).strip(), supplier=str(r[1] or "").strip(),
                note=str(r[2] or "").strip(), ntype=str(r[3] or "").strip(),
                supplytype=str(r[4] or "").strip(), date=str(r[5] or "").strip(),
                noteval=_num(r[6]), pos=str(r[7] or "").strip(),
                rate=_num(r[9]), taxable=_num(r[10]),
                igst=_num(r[11]), cgst=_num(r[12]), sgst=_num(r[13]), cess=_num(r[14]),
            ))

    summary["available"] = True
    return dict(summary=summary, b2b=b2b, cdnr=cdnr, available=True)


_ZERO_SUMMARY_KEYS = (
    "ITC_all_other_IGST", "ITC_all_other_CGST", "ITC_all_other_SGST", "ITC_all_other_CESS",
    "ITC_rcm_IGST", "ITC_rcm_CGST", "ITC_rcm_SGST", "ITC_rcm_CESS",
    "CN_IGST", "CN_CGST", "CN_SGST", "CN_CESS",
)


def summary_for_month(path, month):
    """Return the 2B summary dict for `month`, read directly from the merged
    GSTR-2B workbook.

    GRACEFUL DEGRADATION (fixed -- previously this raised PeriodParseError
    with NOTHING catching it anywhere in run_monthly_pipeline.py or
    master_build.py's per-month loop, so a taxpayer/month with no GSTR-2B
    supplied crashed the ENTIRE run, not just that month's 2B-dependent
    checks. GSTR-2B is auto-generated by GSTN and usually present, but a
    small/new taxpayer's early months, or a partial upload, can genuinely
    lack it -- exactly the 'limited data must not error out' requirement.

    Now: if `path` is missing, or doesn't cover `month`, returns a summary
    with available=False and every numeric field as None (NOT zero -- zero
    would look like a real, verified nil balance and get diffed against
    GSTR-3B as if it were data, producing a wall of false MISMATCH rows;
    None makes 'not available' visually and programmatically distinct from
    'available and nil'). Callers (gst_scrutiny_tool.build_comparisons(),
    gst_eway_recon.run()) must check summary.get('available') before using
    any numeric field -- both have been updated to do so; see their own
    docstrings for the resulting INFO/SKIP behaviour instead of a false
    numeric compare."""
    try:
        parsed = parse_2b_excel(path, month)
    except mpu.PeriodParseError as ex:
        s = {k: None for k in _ZERO_SUMMARY_KEYS}
        s["available"] = False
        s["_reason"] = str(ex)
        s["_source"] = "unavailable"
        s["_file"] = os.path.basename(path) if path else None
        s["_lines"] = None
        return s
    s = dict(parsed["summary"])
    s["_source"] = "excel"
    s["_file"] = os.path.basename(path)
    s["_lines"] = parsed
    return s


if __name__ == "__main__":
    import sys
    p = sys.argv[1] if len(sys.argv) > 1 else "GSTR2B_05AAECM6380J1ZA_012023.xlsx"
    out = parse_2b_excel(p)
    if not out:
        print("Not an Excel 2B:", p); raise SystemExit
    s = out["summary"]
    print("2B SUMMARY (from Excel):")
    for k, v in s.items():
        if k != "available":
            print(f"  {k}: {v:,.2f}" if isinstance(v, float) else f"  {k}: {v}")
    print(f"\nB2B invoice lines: {len(out['b2b'])}  |  CDNR notes: {len(out['cdnr'])}")
    print(f"B2B taxable sum: {sum(x['taxable'] for x in out['b2b']):,.2f}")
    print("CDNR notes:")
    for c in out["cdnr"]:
        print(f"  {c['note']:18} {c['ntype']:12} {c['gstin']} {c['taxable']:>11,.2f} / "
              f"{c['igst']+c['cgst']+c['sgst']:>10,.2f}")
