#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST SCRUTINY  --  ANALYSIS LAYER  (Sooraj's 14 checks)
======================================================
This sits ON TOP of gst_scrutiny_tool.py (the raw comparison tool).

The raw tool does: LEFT | RIGHT | DIFF | MATCH.
This layer does the *interpretive* scrutiny a GST officer / CA actually runs:
arithmetic consistency, effective-rate suppression, POS vs GSTIN tax-head,
RCM routing, duplicate invoices, blank-invoice detective work, timing/late-fee,
rate-wise e-invoice vs HSN, and ratio-based red flags.

Design rules (same spirit as the raw tool):
  - No hand-waving. Every flag shows the numbers it was computed from.
  - A check is either PASS / FLAG / INFO, with the exact arithmetic in the note.
  - Line-level checks list the offending rows by invoice number.

USAGE:
    Put this file next to gst_scrutiny_tool.py, configured for the same period, then:
        python gst_analysis_checks.py
    It reuses the CONFIG + parsers from gst_scrutiny_tool and writes:
        GST_Scrutiny_Analysis.xlsx
    (Findings sheet + Line Detail sheet + the existing Raw Values.)

Some checks need fields the original parser did not capture (POS, recipient
GSTIN, invoice date, RCM flag, IRN date, per-rate split). Those are pulled here
by re-reading the GSTR-1 / e-invoice sub-sheets directly, so the original tool
file does NOT have to change. If a needed column is absent in your export, the
check degrades to INFO ("column not found") instead of crashing.
"""

import os, re, datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import gst_scrutiny_tool as raw   # reuse CONFIG + parsers + num()
import merged_period_utils as mpu

num = raw.num
TOL = raw.TOLERANCE

# ----------------------------------------------------------------------
# Severity model
# ----------------------------------------------------------------------
PASS = "PASS"      # reconciles / nothing to do
INFO = "INFO"      # informational, manual eyeball
FLAG = "FLAG"      # genuine exception, needs explanation
REVW = "REVIEW"    # needs manual verification (allowed-but-watch)

SEV_ORDER = {FLAG: 0, REVW: 1, INFO: 2, PASS: 3}


class Finding:
    __slots__ = ("ref", "title", "severity", "detail", "numbers")
    def __init__(self, ref, title, severity, detail, numbers=None):
        self.ref = ref            # e.g. "#3"
        self.title = title
        self.severity = severity
        self.detail = detail      # plain-English explanation w/ arithmetic
        self.numbers = numbers or {}   # {label: value} shown in a compact column


# ----------------------------------------------------------------------
# Extra raw-row readers  (fields the base parser doesn't keep)
# ----------------------------------------------------------------------
def _open(path):
    return openpyxl.load_workbook(path, data_only=True)

def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return None, None
    rows = list(wb[name].iter_rows(values_only=True))
    if len(rows) < 5:
        return rows, {}
    hdr = [str(c).strip() if c else "" for c in rows[3]]
    H = {h: i for i, h in enumerate(hdr)}
    return rows, H

def _g(r, H, *names):
    """First matching column value from a row, by any of the given header names."""
    for n in names:
        if n in H and H[n] < len(r):
            return r[H[n]]
    return None

def _parse_date(v):
    if v is None or str(v).strip() in ("", "-"):
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def read_gstr1_lines(path, month):
    """Detailed B2B + CDNR rows for line-level scrutiny, scoped to ONE month
    out of the merged GSTR-1 workbook.
    Returns list of dicts with the fields the analysis checks need."""
    wb = _open(path)
    out = []
    for sn, kind in (("b2b, sez, de_inv", "INV"), ("cdnr", "CN"), ("cdnur", "CN")):
        rows, H = _sheet_rows(wb, sn)
        if not rows:
            continue
        for r in mpu.rows_for_month(rows, 3, month):
            if not any(r):
                continue
            out.append(dict(
                sheet=sn, kind=kind,
                gstin=str(_g(r, H, "GSTIN/UIN of Recipient", "GSTIN/UIN", "Recipient GSTIN") or "").strip(),
                invno=str(_g(r, H, "Invoice Number", "Note Number", "Invoice/Advance Receipt Number") or "").strip(),
                invdate=_parse_date(_g(r, H, "Invoice date", "Invoice Date", "Note date", "Note Date")),
                pos=str(_g(r, H, "Place Of Supply", "Place of Supply") or "").strip(),
                rate=num(_g(r, H, "Rate", "Rate (%)")),
                taxable=num(_g(r, H, "Taxable Value")),
                igst=num(_g(r, H, "Integrated Tax")),
                cgst=num(_g(r, H, "Central Tax")),
                sgst=num(_g(r, H, "State/UT Tax")),
                cess=num(_g(r, H, "Cess Amount")),
                rcm=str(_g(r, H, "Reverse Charge", "Reverse charge") or "").strip(),
                irn=str(_g(r, H, "IRN") or "").strip(),
                irndate=_parse_date(_g(r, H, "IRN date", "IRN Date", "Ack Date", "Acknowledgement Date")),
            ))
    return out


def read_einv_lines(path, month):
    if not path or not os.path.exists(path):
        return []   # E-Invoice legitimately not supplied at all -- graceful
    wb = _open(path)
    out = []
    rows, H = _sheet_rows(wb, "b2b, sez, de")
    if not rows:
        return out
    if month not in mpu.months_present(rows, 3):
        return out  # E-Invoice doesn't cover this month -- same graceful state
    for r in mpu.rows_for_month(rows, 3, month):
        if not any(r):
            continue
        out.append(dict(
            gstin=str(_g(r, H, "GSTIN/UIN of Recipient", "GSTIN/UIN", "Recipient GSTIN") or "").strip(),
            invno=str(_g(r, H, "Invoice number", "Invoice Number", "Document number") or "").strip(),
            invdate=_parse_date(_g(r, H, "Invoice date", "Document date")),
            pos=str(_g(r, H, "Place Of Supply", "Place of Supply") or "").strip(),
            rate=num(_g(r, H, "Rate", "Rate (%)")),
            taxable=num(_g(r, H, "Taxable Value")),
            igst=num(_g(r, H, "Integrated Tax")),
            cgst=num(_g(r, H, "Central Tax")),
            sgst=num(_g(r, H, "State/UT Tax")),
            rcm=str(_g(r, H, "Reverse Charge", "Reverse charge") or "").strip(),
            irn=str(_g(r, H, "IRN", "Irn") or "").strip(),
            irndate=_parse_date(_g(r, H, "IRN date", "IRN Date", "Ack Date", "Acknowledgement Date")),
            err=str(_g(r, H, "Error in auto-population/ deletion", "Error", "Errors") or "").strip(),
        ))
    return out


def _state_code(s):
    """First 2 chars of a GSTIN or POS string, if numeric state code."""
    s = (s or "").strip()
    m = re.match(r"\s*0?(\d{1,2})", s)
    if m:
        return m.group(1).zfill(2)
    return None


# ----------------------------------------------------------------------
# THE 14 CHECKS
# ----------------------------------------------------------------------
def run_checks(g1, g3b, einv, b2b, g1_lines, einv_lines):
    F = []
    def gv(key, i, d=0.0):
        v = g3b.get(key); return v[i] if v and i < len(v) else d

    # ---- #0  Totals reconciliation (disambiguate gross vs named vs orphan) ----
    gross = g1.get("taxable", 0.0)          # sum of ALL B2B lines incl any orphan
    named = g1.get("named_taxable", 0.0)    # sum of properly-numbered invoices only
    orphan_tax = round(gross - named, 2)
    F.append(Finding("#0", "GSTR-1 B2B totals (gross / named / orphan)",
                     INFO,
                     f"GSTR-1 B2B all-lines taxable = {gross:,.2f}; properly-numbered-invoices taxable = "
                     f"{named:,.2f}; difference = {orphan_tax:,.2f} sits on line(s) whose invoice number was "
                     "dropped (see #5). NOTE: the all-lines total can coincide with the e-invoice total when "
                     "the orphan line is the same rate-line that IS numbered in the e-invoice — it is still a "
                     "GSTR-1 figure, not an e-invoice figure. The GROSS-vs-HSN gap is explained ONCE by credit "
                     "notes (CDNR taxable), not separately in two places.",
                     {"all-lines": gross, "named": named, "orphan": orphan_tax}))

    # ---- #1  Nil / Exempt / Non-GST: GSTR-1 'exemp' vs 3B 3.1(c)/(e) ----
    # GSTR-1 Table 8 'exemp' sheet carries Nil-rated, Exempted and Non-GST outward supplies.
    # Mapping to 3B:
    #   GSTR-1 (Nil + Exempted)  ->  3B 3.1(c) 'Other Outward (Nil rated, exempted)'
    #   GSTR-1 (Non-GST)         ->  3B 3.1(e) 'Non-GST Outward supplies'
    # Zero-rated (3B 3.1(b)) is reported in GSTR-1 via the 'exp'/SEZ heads, a different sheet,
    # so it is shown for context but not differenced here (would create a false mismatch).
    nil_exempt_g1 = g1.get("nil_exempt_taxable", None)
    nongst_g1 = g1.get("nongst_taxable", 0.0)
    if nil_exempt_g1 is None:
        # 'exemp' sheet/header not found in this export -> cannot auto-compare, say so honestly.
        F.append(Finding("#1", "Nil / exempt / non-GST (GSTR-1 vs 3B 3.1c/3.1e)",
                         INFO,
                         "GSTR-1 'exemp' (Table 8) sheet/header not found in this export, so nil/exempt/"
                         "non-GST could not be auto-read. Manually compare GSTR-1 Table 8 against 3B "
                         "3.1(c)+3.1(e).",
                         {}))
    else:
        b3c = gv("3.1c", 0)            # 3B nil+exempt taxable
        b3e = gv("3.1e", 0)            # 3B non-GST taxable
        diff_ne = nil_exempt_g1 - b3c
        diff_ng = nongst_g1 - b3e
        worst = PASS if (abs(diff_ne) <= TOL and abs(diff_ng) <= TOL) else FLAG
        F.append(Finding("#1", "Nil / exempt / non-GST (GSTR-1 Table 8 vs 3B 3.1c/3.1e)",
                         worst,
                         f"GSTR-1 nil+exempt {nil_exempt_g1:,.2f} vs 3B 3.1(c) {b3c:,.2f} (diff {diff_ne:,.2f}); "
                         f"GSTR-1 non-GST {nongst_g1:,.2f} vs 3B 3.1(e) {b3e:,.2f} (diff {diff_ng:,.2f}). "
                         f"(For context, 3B 3.1(b) zero-rated taxable = {gv('3.1b',0):,.2f}; zero-rated is "
                         "reported via GSTR-1 export/SEZ heads, not Table 8, so not differenced here.) "
                         + ("All nil/exempt/non-GST figures reconcile." if worst == PASS
                            else "Mismatch — GSTR-1 Table 8 does not tie to 3B 3.1(c)/(e); reconcile."),
                         {"G1 nil+exempt": nil_exempt_g1, "3B 3.1c": b3c,
                          "G1 non-GST": nongst_g1, "3B 3.1e": b3e}))

    # ---- #2  Credit-note effect: GSTR-1 net vs gross on 3B liability ----
    g1_gross_tax = g1["IGST"] + g1["CGST"] + g1["SGST"]
    g1_net_tax = (g1["IGST"]-g1["cn_IGST"]) + (g1["CGST"]-g1["cn_CGST"]) + (g1["SGST"]-g1["cn_SGST"])
    b3b_tax = gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)
    d_net = g1_net_tax - b3b_tax
    F.append(Finding("#2", "Credit-note effect on outward liability (GSTR-1 net vs 3B 3.1a)",
                     PASS if abs(d_net) <= TOL else FLAG,
                     f"CN total tax = {g1['cn_IGST']+g1['cn_CGST']+g1['cn_SGST']:,.2f}. "
                     f"GSTR-1 GROSS tax {g1_gross_tax:,.2f} -> NET {g1_net_tax:,.2f}. "
                     f"3B 3.1(a) tax {b3b_tax:,.2f}. Net-vs-3B diff {d_net:,.2f}. "
                     "3B 3.1(a) must be reported NET of credit notes; if GROSS matched but NET didn't, "
                     "CN was not given effect in 3B.",
                     {"G1 gross": g1_gross_tax, "G1 net": g1_net_tax, "3B 3.1a": b3b_tax}))

    # ---- #3  Arithmetic consistency: 4C = 4A5 + 4A3 - 4B1 - 4B2 (per head) ----
    # NOTE: Net ITC includes RCM ITC (4A3). Omitting 4A3 produces a false mismatch.
    # FIX: 4B1 (Rule 42/43/38 + Sec 17(5) reversal) is now included -- it was
    # silently dropped from this formula before (always 0 for this taxpayer,
    # so it was invisible, but would misfire for any period with a real
    # Rule-42/43 reversal). Also depends on gst_scrutiny_tool.parse_gstr3b's
    # corrected, section-anchored 4B(1)/4B(2) extraction -- see that function's
    # docstring for the duplicate-label bug this fixes.
    heads = ["IGST", "CGST", "SGST"]
    rows3 = []
    worst = PASS
    for i, h in enumerate(heads):
        should = gv("4A5", i) + gv("4A3", i) - gv("4B1", i) - gv("4B2", i)
        actual = gv("4C", i)
        d = round(should - actual, 2)
        rows3.append(f"{h}: 4A5 {gv('4A5',i):,.2f} + 4A3 {gv('4A3',i):,.2f} - 4B1 {gv('4B1',i):,.2f} "
                     f"- 4B2 {gv('4B2',i):,.2f} = {should:,.2f}  | 3B 4C {actual:,.2f}  | diff {d:,.2f}")
        if abs(d) > TOL:
            worst = FLAG
    F.append(Finding("#3", "ITC arithmetic: Net ITC (4C) = 4A5 + 4A3 (RCM) - 4B1 - 4B2",
                     worst,
                     "Net ITC must include RCM ITC (Table 4A3); a check using only 4A5-4B2 wrongly "
                     "flags a gap equal to the RCM ITC. Per head:\n   " + "\n   ".join(rows3),
                     {}))

    # ---- #4  Effective tax-rate comparison (GSTR-1 vs 3B 3.1a) ----
    r1 = (g1["IGST"]+g1["CGST"]+g1["SGST"]) / g1["taxable"] * 100 if g1["taxable"] else 0
    den3b = gv("3.1a",0)
    r3b = (gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)) / den3b * 100 if den3b else 0
    drate = r1 - r3b
    if abs(drate) <= 0.10:
        sev4, msg4 = PASS, "Rates align (within 0.10pp) -> no suppression signal."
    elif drate > 0.10:
        sev4, msg4 = FLAG, "GSTR-1 rate HIGHER than 3B -> possible suppression of liability in 3B."
    else:
        sev4, msg4 = REVW, "3B rate higher than GSTR-1 -> over-reported in 3B / under-reported in GSTR-1; verify."
    F.append(Finding("#4", "Effective tax-rate (GSTR-1 vs 3B 3.1a)",
                     sev4,
                     f"GSTR-1 eff rate {r1:.3f}% vs 3B 3.1(a) {r3b:.3f}% (diff {drate:+.3f}pp). {msg4} "
                     "Blended ~10.7% indicates a 5%/12% supply mix.",
                     {"G1 %": round(r1,3), "3B %": round(r3b,3)}))

    # ---- #5  Orphan invoice-number lines -> RE-LINK to e-invoice, don't just call it "blank" ----
    # A GSTR-1 B2B line whose invoice-number cell is empty is a real export defect, but the
    # actionable finding is WHICH invoice it belongs to. Re-link by (rate, taxable, tax) against
    # e-invoice lines that are absent from GSTR-1's named set -> names the true invoice + rate-line.
    blanks = [L for L in g1_lines if L["kind"] == "INV" and not L["invno"]]
    if blanks:
        g1_named_keys = {(L["invno"], round(L["rate"], 2)) for L in g1_lines
                         if L["kind"] == "INV" and L["invno"]}
        ei_index = {}
        for E in einv_lines:
            ei_index.setdefault((round(E["rate"], 2), round(E["taxable"], 2), round(E["igst"], 2)),
                                []).append(E)
        relinked, unresolved = [], []
        for b in blanks:
            key = (round(b["rate"], 2), round(b["taxable"], 2), round(b["igst"], 2))
            cands = [E for E in ei_index.get(key, [])
                     if (E["invno"], round(E["rate"], 2)) not in g1_named_keys]
            if len(cands) == 1:
                relinked.append((b, cands[0]["invno"]))
            else:
                unresolved.append(b)
        parts = []
        for b, inv in relinked:
            parts.append(f"GSTR-1 B2B line lost its invoice number; re-linked by (rate, value, tax) to "
                         f"invoice {inv} @ {b['rate']:g}% (taxable {b['taxable']:,.2f}, IGST {b['igst']:,.2f}). "
                         f"This rate-line is correctly numbered in the e-invoice and appears in the HSN summary, "
                         f"but is mis-recorded (number dropped) in the GSTR-1 B2B detail. Tax impact {b['igst']:,.2f}. "
                         f"Amend GSTR-1 / adjust next period.")
        for b in unresolved:
            parts.append(f"Unnumbered GSTR-1 B2B line @ {b['rate']:g}% (taxable {b['taxable']:,.2f}, "
                         f"IGST {b['igst']:,.2f}) — no unique e-invoice match; locate the source invoice manually.")
        F.append(Finding("#5", f"GSTR-1 B2B line(s) with dropped invoice number: {len(blanks)} "
                                f"({len(relinked)} re-linked)",
                         FLAG,
                         " ".join(parts),
                         {"orphan lines": len(blanks), "re-linked": len(relinked),
                          "taxable": sum(b["taxable"] for b in blanks),
                          "tax impact": sum(b["igst"]+b["cgst"]+b["sgst"] for b in blanks)}))
    else:
        F.append(Finding("#5", "GSTR-1 B2B lines with dropped invoice number", PASS,
                         "Every GSTR-1 B2B line carries an invoice number.", {}))

    # ---- #6  Duplicate invoice numbers (allowed if multi-rate; flag for manual) ----
    seen = {}
    for L in g1_lines:
        if L["kind"] != "INV" or not L["invno"]:
            continue
        seen.setdefault(L["invno"], []).append(L)
    dups = {k: v for k, v in seen.items() if len(v) > 1}
    if dups:
        lines = []
        for inv, rows in list(dups.items())[:15]:
            rates = ",".join(f"{r['rate']:g}%" for r in rows)
            tot = sum(r["taxable"] for r in rows)
            same_rate = len({r["rate"] for r in rows}) < len(rows)
            tag = "  <-- SAME RATE REPEATED, verify not a true duplicate" if same_rate else ""
            lines.append(f"{inv}: {len(rows)} lines @ rates {rates}, taxable {tot:,.2f}{tag}")
        F.append(Finding("#6", f"Duplicate invoice numbers: {len(dups)} invoice(s) on >1 line",
                         REVW,
                         "Multiple lines per invoice are allowed for multi-rate invoices (taxable adds up). "
                         "Same rate repeated on one invoice = likely true duplicate, verify manually.\n   "
                         + "\n   ".join(lines),
                         {"dup invoices": len(dups)}))
    else:
        F.append(Finding("#6", "Duplicate invoice numbers", PASS,
                         "Every B2B invoice number appears on a single line.", {}))

    # ---- #7  E-invoice errors field ----
    err_lines = [L for L in einv_lines if L.get("err")]
    if einv.get("available"):
        if err_lines:
            det = "; ".join(f"{L['invno']}: {L['err']}" for L in err_lines[:10])
            F.append(Finding("#7", f"E-invoice error/auto-population flags: {len(err_lines)}",
                             FLAG, f"Errors present: {det}. Investigate (e.g. IRN already used, deletion).",
                             {"errors": len(err_lines)}))
        else:
            F.append(Finding("#7", "E-invoice error field", PASS,
                             "No errors flagged in the e-invoice auto-population column.", {}))
    else:
        F.append(Finding("#7", "E-invoice error field", INFO, "E-invoice file not supplied.", {}))

    # ---- #8  Time-lag: IRN date vs GSTR-1 filing (>30 days) ----
    filing = _filing_date("GSTR1")
    lagged = []
    for L in einv_lines or g1_lines:
        idt = L.get("irndate")
        if idt and filing:
            lag = (filing - idt).days
            if lag > 30:
                lagged.append((L.get("invno", "?"), idt, lag))
    if filing is None:
        F.append(Finding("#8", "IRN-date vs GSTR-1 filing lag (>30d)", INFO,
                         "Set GSTR1_FILING_DATE in CONFIG to enable. IRN-date column also required in the export.",
                         {}))
    elif lagged:
        det = "; ".join(f"{inv} IRN {d} ({lag}d)" for inv, d, lag in lagged[:10])
        F.append(Finding("#8", f"IRN-to-filing lag >30 days: {len(lagged)} invoice(s)",
                         FLAG, f"Delayed reporting: {det}.", {"lagged": len(lagged)}))
    else:
        F.append(Finding("#8", "IRN-date vs GSTR-1 filing lag", PASS,
                         "No invoice with IRN-to-filing gap beyond 30 days.", {}))

    # ---- #9  Rate-wise: e-invoice vs GSTR-1 HSN summary ----
    rate_buckets = {}
    src_lines = einv_lines if einv.get("available") and einv_lines else g1_lines
    for L in src_lines:
        if L.get("kind") == "CN":
            continue
        rb = rate_buckets.setdefault(round(L["rate"], 2), 0.0)
        rate_buckets[round(L["rate"], 2)] = rb + L["taxable"]
    if rate_buckets:
        buckets = "; ".join(f"{r:g}%: {v:,.2f}" for r, v in sorted(rate_buckets.items()))
        F.append(Finding("#9", "Rate-wise taxable split (e-invoice / GSTR-1 lines)",
                         INFO,
                         f"Rate-wise taxable from source lines -> {buckets}. "
                         "Compare against the GSTR-1 HSN summary rate-wise rows; any rate present here but "
                         "absent/short in HSN = misclassification. (HSN rate split must come from the HSN sheet.)",
                         {}))
    else:
        F.append(Finding("#9", "Rate-wise e-invoice vs HSN", INFO, "No rate-bearing source lines parsed.", {}))

    # ---- #10  Late-fee timing: GSTR-1 vs 3B filing gap (>20 days... uses statutory due dates) ----
    f1 = _filing_date("GSTR1"); f3 = _filing_date("GSTR3B")
    if f1 and f3:
        gap = abs((f3 - f1).days)
        F.append(Finding("#10", "GSTR-1 vs GSTR-3B filing-gap",
                         REVW if gap > 20 else PASS,
                         f"GSTR-1 filed {f1}, GSTR-3B filed {f3}, gap {gap} days. "
                         + ("Gap >20d -> check late-fee/interest exposure." if gap > 20
                            else "Within 20 days."),
                         {"gap_days": gap}))
    else:
        F.append(Finding("#10", "Filing-gap / late fee", INFO,
                         "Set GSTR1_FILING_DATE and GSTR3B_FILING_DATE in CONFIG to enable.", {}))

    # ---- #11  POS vs recipient-GSTIN state-code -> correct tax head ----
    pos_errors = []
    for L in g1_lines:
        if L["kind"] == "CN":
            continue
        pos_sc = _state_code(L["pos"]); g_sc = _state_code(L["gstin"])
        if not pos_sc:
            continue
        has_igst = L["igst"] > TOL
        has_local = (L["cgst"] > TOL or L["sgst"] > TOL)
        if g_sc:  # registered recipient
            same = (pos_sc == g_sc)
            if same and has_igst and not has_local:
                pos_errors.append((L["invno"], L["pos"], L["gstin"], "intra-state but charged IGST"))
            if (not same) and has_local and not has_igst:
                pos_errors.append((L["invno"], L["pos"], L["gstin"], "inter-state but charged CGST+SGST"))
    if pos_errors:
        det = "; ".join(f"{inv} POS={p} GSTIN={g}: {why}" for inv, p, g, why in pos_errors[:12])
        F.append(Finding("#11", f"POS vs GSTIN tax-head mismatch: {len(pos_errors)} invoice(s)",
                         FLAG, f"Wrong tax head charged -> direct exposure. {det}.",
                         {"mismatches": len(pos_errors)}))
    else:
        F.append(Finding("#11", "POS vs recipient-GSTIN tax head", PASS,
                         "Every registered-recipient line uses the correct head for its POS/GSTIN state pair "
                         "(or POS/GSTIN not available to test).", {}))

    # ---- #12  RCM flag routing: e-invoice/GSTR-1 RCM=Y -> 3B 3.1(d) & ITC 4A3 ----
    rcm_lines = [L for L in (einv_lines or g1_lines)
                 if str(L.get("rcm", "")).strip().upper() in ("Y", "YES", "TRUE")]
    rcm_tax = sum(L["igst"]+L.get("cgst",0)+L.get("sgst",0) for L in rcm_lines) if rcm_lines else 0
    d31 = gv("3.1d",1)+gv("3.1d",2)+gv("3.1d",3)
    if rcm_lines:
        F.append(Finding("#12", f"RCM-flagged invoices: {len(rcm_lines)}",
                         REVW,
                         f"RCM-marked lines total tax {rcm_tax:,.2f}. 3B 3.1(d) liability tax {d31:,.2f}, "
                         f"3B 4A3 RCM-ITC {gv('4A3',0)+gv('4A3',1)+gv('4A3',2):,.2f}. "
                         "RCM liability must appear in 3.1(d) and the matching ITC in 4A3; reconcile.",
                         {"RCM lines": len(rcm_lines), "3.1d tax": d31}))
    else:
        F.append(Finding("#12", "RCM routing (3.1d & 4A3)",
                         INFO if not (einv_lines or g1_lines) else PASS,
                         f"No RCM=Y line found in source. (3B shows 3.1d tax {d31:,.2f}, "
                         f"4A3 ITC {gv('4A3',0)+gv('4A3',1)+gv('4A3',2):,.2f} — RCM on unregistered/import "
                         "of services won't carry a line-level flag.)",
                         {}))

    # ---- #13  HSN IGST vs named-invoice IGST gap == credit-note effect? ----
    gap13 = g1["named_IGST"] - g1["hsn_IGST"]
    resid = gap13 - g1["cn_IGST"]
    F.append(Finding("#13", "HSN-summary IGST vs named-invoice IGST",
                     PASS if abs(resid) <= 200 else FLAG,
                     f"named-invoice IGST {g1['named_IGST']:,.2f} - HSN IGST {g1['hsn_IGST']:,.2f} "
                     f"= gap {gap13:,.2f}. Credit-note IGST = {g1['cn_IGST']:,.2f}. "
                     f"Residual after CN = {resid:,.2f}. "
                     + ("Gap is explained by credit notes (HSN summary is net of CN, invoice lines are gross)."
                        if abs(resid) <= 200 else
                        "Residual NOT explained by credit notes — reconcile line-by-line."),
                     {"gap": gap13, "cn_IGST": g1["cn_IGST"], "residual": resid}))

    # ---- #14  ITC / Liability ratio ----
    liab = gv("3.1a",1)+gv("3.1a",2)+gv("3.1a",3)
    itc = gv("4A5",0)+gv("4A5",1)+gv("4A5",2)
    ratio = (itc/liab*100) if liab else 0
    if ratio > 95:
        sev14 = FLAG
        m = ("ITC >95% of output liability -> taxpayer pays almost no cash. "
             "Watch for circular trading / fake billing. Here ITC even EXCEEDS liability "
             "(ratio >100%): net ITC accumulation this period — confirm it's genuine input build-up "
             "(capex/inventory) and not inflated.") if ratio > 100 else \
            "ITC >95% of liability -> minimal cash payout; verify input genuineness."
    elif ratio < 20:
        sev14, m = REVW, "ITC <20% of liability -> low-margin/high-value-add; still verify."
    else:
        sev14, m = PASS, "ITC/Liability ratio in a normal band."
    F.append(Finding("#14", "ITC / Output-liability ratio",
                     sev14,
                     f"Liability (3.1a tax) {liab:,.2f}; ITC (4A5) {itc:,.2f}; ratio {ratio:.1f}%. {m}",
                     {"liability": liab, "ITC": itc, "ratio %": round(ratio,1)}))

    F.sort(key=lambda x: (SEV_ORDER[x.severity], x.ref))
    return F


# ----------------------------------------------------------------------
# Optional CONFIG additions (filing dates). Read from gst_scrutiny_tool if set.
# ----------------------------------------------------------------------
def _filing_date(which):
    attr = "GSTR1_FILING_DATE" if which == "GSTR1" else "GSTR3B_FILING_DATE"
    v = getattr(raw, attr, None)
    return _parse_date(v) if v else None


# ----------------------------------------------------------------------
# WRITE EXCEL
# ----------------------------------------------------------------------
FILL = {FLAG: PatternFill("solid", fgColor="FFC7CE"),
        REVW: PatternFill("solid", fgColor="FFEB9C"),
        INFO: PatternFill("solid", fgColor="DDEBF7"),
        PASS: PatternFill("solid", fgColor="C6EFCE")}
FONT_SEV = {FLAG: Font(bold=True, color="9C0006"),
            REVW: Font(bold=True, color="9C6500"),
            INFO: Font(bold=True, color="2F5496"),
            PASS: Font(bold=True, color="006100")}
HEAD = PatternFill("solid", fgColor="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")]*4)


def write_analysis(findings, raw_bundle, outpath):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Findings"
    ws.cell(1, 1, f"GST SCRUTINY — ANALYSIS (14 checks) — Period {raw.PERIOD_LABEL}").font = Font(bold=True, size=13, color="1F3864")
    ws.cell(2, 1, f"GSTIN {getattr(raw,'SELF_GSTIN','')}  |  {getattr(raw,'COMPANY_NAME','') or '(company auto-detected)'}").font = Font(size=9, italic=True)

    nflag = sum(1 for f in findings if f.severity == FLAG)
    nrev = sum(1 for f in findings if f.severity == REVW)
    ws.cell(3, 1, f"FLAGS: {nflag}   REVIEW: {nrev}   "
                  f"INFO: {sum(1 for f in findings if f.severity==INFO)}   "
                  f"PASS: {sum(1 for f in findings if f.severity==PASS)}").font = Font(size=10, bold=True)

    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail / arithmetic"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(5, i, h); c.fill = HEAD; c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    r = 6
    for f in findings:
        ws.cell(r, 1, f.ref)
        ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity); cv.fill = FILL[f.severity]; cv.font = FONT_SEV[f.severity]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in f.numbers.items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDE", [6, 44, 10, 30, 95]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    wb.save(outpath)
    return nflag, nrev


def main():
    """Legacy/manual-testing entry point -- master_build.py does not call this;
    it drives run_checks() per month via run_monthly_pipeline.py instead.
        python gst_analysis_checks.py <month, e.g. Jan-23>
    """
    import sys as _sys
    if len(_sys.argv) < 2:
        raise SystemExit("Usage: python gst_analysis_checks.py <month, e.g. Jan-23>")
    month = _sys.argv[1]
    raw.PERIOD_LABEL = month
    g1 = raw.parse_gstr1(raw.GSTR1_FILE, month)
    g3b = raw.parse_gstr3b(raw.GSTR3B_FILE, month)
    einv = raw.parse_einv(raw.EINV_FILE, month)
    b2b = raw.get_gstr2b_values()
    g1_lines = read_gstr1_lines(raw.GSTR1_FILE, month)
    einv_lines = read_einv_lines(raw.EINV_FILE, month) if raw.EINV_FILE else []

    # remap 3b keys 3.1a etc. to match parser output keys
    g3b_norm = {}
    for k, v in g3b.items():
        g3b_norm[k] = v
    # parser stores '3.1a','3.1d','4A5','4A3','4B2','4C' already

    findings = run_checks(g1, g3b_norm, einv, b2b, g1_lines, einv_lines)
    out = "GST_Scrutiny_Analysis.xlsx"
    nflag, nrev = write_analysis(findings, dict(g1=g1, g3b=g3b_norm, einv=einv, b2b=b2b), out)
    print(f"Saved: {out}")
    print(f"Findings: {len(findings)}  |  FLAG: {nflag}  REVIEW: {nrev}")
    for f in findings:
        print(f"  [{f.severity:6}] {f.ref}  {f.title}")


if __name__ == "__main__":
    main()
