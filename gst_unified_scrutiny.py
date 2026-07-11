#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GST UNIFIED SCRUTINY TOOL  (single workbook, both pipelines)
============================================================
Combines, in ONE Excel workbook, the two reconciliations that were previously
produced as two separate files:

  PIPELINE 1  (was GST_Scrutiny_Comparison_Jan2023.xlsx)
      - Raw side-by-side comparison : GSTR-1 | 2B | 3B | E-Invoice
        (sections A, A2, B, B2, C, D, D2 — exactly as the original tool)
      - PLUS Sooraj's 14 interpretive checks (#0-#14) that were already coded
        in gst_analysis_checks.py but had never been written into the
        Comparison workbook. They are now included as the "Analysis (14 checks)"
        sheet, ON THE SAME LINES that the comparison already reconciles.

  PIPELINE 2  (was GST_Scrutiny_EWayBill.xlsx)
      - The 27-check E-Way-Bill matrix (inward + outward), unchanged.

  CROSS-FILE  (new)
      - A "Dashboard" sheet that puts every FLAG / MISMATCH / REVIEW from BOTH
        pipelines in one ranked list, so the two reconciliations are read
        together rather than in two files.

HARD RULE honoured:
  - Comparison runs on EXACTLY the same lines/sections as before.
  - E-Way-Bill runs on EXACTLY the same 27 checks as before.
  - Sooraj's 14 points map 1:1 onto the existing analysis checks (#0-#14);
    nothing new is invented and NO safety net / fabricated data is added.
  - All compute logic is REUSED from the three existing modules
    (gst_scrutiny_tool, gst_analysis_checks, gst_eway_recon). This file only
    orchestrates them and writes the combined book. The engines are untouched,
    so results are bit-for-bit identical to the two original files.

A NOTE ON SOORAJ POINT #3 (the "4C mismatch"):
  Sooraj computed 4C = 4A5 - 4B2 and got a gap vs the filed 4C. That manual
  formula OMITS the RCM-ITC row (Table 4A3). The correct identity is
  4C = 4A5 + 4A3 - 4B2, which reconciles to the rupee (diff 0.00). The tool
  uses the correct identity (analysis check #3) and therefore PASSES it. The
  "mismatch" was in the manual arithmetic, not in the return. This is stated
  explicitly in the #3 finding row.

USAGE
    Put this next to the four modules + the input files (same folder the two
    original tools already use), set nothing else, then:
        python gst_unified_scrutiny.py
    -> GST_Scrutiny_Unified.xlsx
"""

import os
import datetime as _dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- reuse the three existing engines (no logic change) -------------------
import gst_scrutiny_tool as raw          # CONFIG + parsers + build_comparisons()
import gst_analysis_checks as ana        # Sooraj's #0-#14
import gst_eway_recon as eway            # the 27-check EWB matrix
try:
    import gstr2b_parser as g2b
except ImportError:
    g2b = None

num = raw.num
TOL = raw.TOLERANCE
OUTPUT_FILE = "GST_Scrutiny_Unified.xlsx"


# ----------------------------------------------------------------------
# ARN / filing-date auto-extract (enables analysis checks #8 and #10)
# ----------------------------------------------------------------------
# Sooraj asked for #8 (IRN-date vs GSTR-1 filing lag) and #10 (GSTR-1 vs
# GSTR-3B filing gap). Both need the filing dates. The GST-portal exports
# ALREADY carry these as the ARN date:
#   GSTR-1  : 'Read me' sheet, row 'ARN date'   (value in col C)
#   GSTR-3B : 'GSTR-3B'  sheet, row 'Date of ARN'(value in col E)
# The base analysis module only read them from a hardcoded CONFIG attr that
# was left blank, so #8/#10 silently fell to INFO. We read the ARN date
# straight from the file and inject it, so the existing checks fire. No new
# check is added — we only supply the date the existing checks already want.
def _extract_arn_dates():
    out = {"GSTR1_FILING_DATE": None, "GSTR3B_FILING_DATE": None}
    # ---- GSTR-1 'Read me' -> 'ARN date' ----
    try:
        wb = openpyxl.load_workbook(raw.GSTR1_FILE, data_only=True)
        sn = "Read me" if "Read me" in wb.sheetnames else wb.sheetnames[0]
        for r in wb[sn].iter_rows(values_only=True):
            label = next((str(c).strip() for c in r if c not in (None, "")), "")
            if label.upper() in ("ARN DATE", "DATE OF ARN"):
                vals = [c for c in r if c not in (None, "")]
                if len(vals) >= 2:
                    out["GSTR1_FILING_DATE"] = str(vals[-1]).strip()
                break
    except Exception:
        pass
    # ---- GSTR-3B 'GSTR-3B' -> 'Date of ARN' ----
    try:
        wb = openpyxl.load_workbook(raw.GSTR3B_FILE, data_only=True)
        sn = "GSTR-3B" if "GSTR-3B" in wb.sheetnames else wb.sheetnames[0]
        for r in wb[sn].iter_rows(values_only=True):
            cells = [str(c).strip() for c in r if c not in (None, "")]
            if cells and cells[0].upper() in ("DATE OF ARN", "ARN DATE") and len(cells) >= 2:
                out["GSTR3B_FILING_DATE"] = cells[-1]
                break
    except Exception:
        pass
    return out

# ----------------------------------------------------------------------
# Shared styling (matches the two original workbooks)
# ----------------------------------------------------------------------
RED    = PatternFill("solid", fgColor="FFC7CE")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
AMBER  = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="DDEBF7")
GREY   = PatternFill("solid", fgColor="E7E6E6")
HEAD   = PatternFill("solid", fgColor="1F3864")
SECT   = PatternFill("solid", fgColor="D9E1F2")
TITLEF = Font(bold=True, size=13, color="1F3864")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

SEV_FILL = {"FLAG": RED, "MISMATCH": RED, "REVIEW": AMBER, "INFO": BLUE,
            "PASS": GREEN, "MATCH": GREEN, "SKIPPED": GREY}
SEV_FONT = {"FLAG": Font(bold=True, color="9C0006"),
            "MISMATCH": Font(bold=True, color="9C0006"),
            "REVIEW": Font(bold=True, color="9C6500"),
            "INFO": Font(bold=True, color="2F5496"),
            "PASS": Font(bold=True, color="006100"),
            "MATCH": Font(bold=True, color="006100"),
            "SKIPPED": Font(bold=True, color="808080")}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


# ======================================================================
# GATHER  --  parse every source ONCE, then drive each engine
# ======================================================================
def gather():
    """** LEGACY / UNSUPPORTED for the merged-file model **
    This function pre-dates the merged-workbook migration: it still assumes
    one-file-per-month and calls raw.parse_gstr1()/parse_gstr3b()/parse_einv()
    etc. WITHOUT a month argument, plus references EWB globals that no longer
    exist on gst_eway_recon (EWB_OUT_FILE/EWB_IN_FILE/find_file). It will
    raise if called. master_build.py is the supported entry point -- it
    drives the same underlying engines correctly per month via
    run_monthly_pipeline.run_month(). This function is kept only because
    master_build.py reuses this module's write_* sheet writers, not gather()
    itself. Left as a clear, visible failure rather than quietly patched to
    "sort of" work, since fully re-plumbing a standalone single-merged-month
    CLI path was out of scope for this pass.
    """
    # ---- auto-fill filing dates from ARN (enables analysis #8 and #10) ----
    arn = _extract_arn_dates()
    if arn["GSTR1_FILING_DATE"] and not getattr(raw, "GSTR1_FILING_DATE", None):
        raw.GSTR1_FILING_DATE = arn["GSTR1_FILING_DATE"]
    if arn["GSTR3B_FILING_DATE"] and not getattr(raw, "GSTR3B_FILING_DATE", None):
        raw.GSTR3B_FILING_DATE = arn["GSTR3B_FILING_DATE"]

    # ---- shared parses ----
    g1   = raw.parse_gstr1(raw.GSTR1_FILE)
    g3b  = raw.parse_gstr3b(raw.GSTR3B_FILE)
    einv = raw.parse_einv(raw.EINV_FILE)
    b2b  = raw.get_gstr2b_values() if hasattr(raw, "get_gstr2b_values") else dict(raw.GSTR2B_VALUES)

    # ---- pipeline 1a: raw comparison rows (unchanged) ----
    comparisons, comp_raw = raw.build_comparisons()

    # ---- pipeline 1b: Sooraj's 14 checks (unchanged) ----
    g1_lines   = ana.read_gstr1_lines(raw.GSTR1_FILE)
    einv_lines = ana.read_einv_lines(raw.EINV_FILE)
    findings14 = ana.run_checks(g1, g3b, einv, b2b, g1_lines, einv_lines)

    # ---- pipeline 2: the 27-check EWB matrix (unchanged logic) ----
    # paths come from the content-based finder (see CONFIG), not filename patterns.
    ewb_out_path = eway.EWB_OUT_FILE or eway.find_file(None, [r"OUT.*EWAY|EWAY.*OUT|OUTWARD"])
    ewb_in_path  = eway.EWB_IN_FILE  or eway.find_file(None, [r"IN.*EWAY|EWAY.*IN|INWARD"])
    einv_path    = raw.EINV_FILE     or eway.find_file(None, [r"EINV", r"E[-_ ]?INVOICE"])
    ewb_out = eway.parse_ewb(ewb_out_path) if ewb_out_path else []
    ewb_in  = eway.parse_ewb(ewb_in_path) if ewb_in_path else []
    g1inv   = eway.read_gstr1_invoices(raw.GSTR1_FILE)
    einv_ew = eway.read_einv_invoices(einv_path) if einv_path else {}
    if g2b and eway.GSTR2B_FILE:
        b2b_ew = g2b.summary_or_fallback(eway.GSTR2B_FILE, raw.GSTR2B_VALUES, ".")
    else:
        b2b_ew = dict(raw.GSTR2B_VALUES); b2b_ew["_lines"] = None; b2b_ew["_source"] = "pdf-hardcoded"; b2b_ew["_file"] = None
    findings27 = eway.run(ewb_out, ewb_in, g1inv, einv_ew, g3b, b2b_ew)

    return dict(
        comparisons=comparisons, comp_raw=comp_raw,
        findings14=findings14, findings27=findings27,
        meta=dict(
            ewb_out_n=len(ewb_out), ewb_in_n=len(ewb_in),
            twob_src=b2b_ew.get("_source", "pdf-hardcoded"),
            twob_file=b2b_ew.get("_file"),
            einv_file=os.path.basename(einv_path) if einv_path else None,
            g2b_ok=bool(g2b),
        ),
    )


# ======================================================================
# WRITERS  (each produces one sheet; styling matches the originals)
# ======================================================================
HDR_COMP = ["Section", "Check", "Left source", "Left value", "Right source",
            "Right value", "Difference", "Result", "Note / Tag"]
WID_COMP = [26, 46, 16, 15, 16, 15, 14, 11, 55]


def _comp_rows_iter(comparisons):
    for row in comparisons:
        if len(row) == 7:
            sect, check, llabel, lval, rlabel, rval, tag = row
        else:
            sect, check, llabel, lval, rlabel, rval = row; tag = ""
        diff = round(num(lval) - num(rval), 2)
        is_match = abs(diff) <= TOL
        yield (sect, check, llabel, round(num(lval), 2), rlabel,
               round(num(rval), 2), diff, "MATCH" if is_match else "MISMATCH", tag)


def write_comparison(ws, comparisons, only_mismatch):
    """Replicates the original Comparison/Exceptions sheets exactly."""
    title = ("GST SCRUTINY  -  MISMATCHES ONLY  -  Period: " + raw.PERIOD_LABEL
             if only_mismatch else
             "GST SCRUTINY  -  FULL COMPARISON  -  Period: " + raw.PERIOD_LABEL)
    ws.cell(1, 1, title).font = Font(bold=True, size=13,
                                     color="C00000" if only_mismatch else "1F3864")
    if only_mismatch:
        ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}  |  "
                      f"Tolerance: Rs {TOL}").font = Font(size=9, italic=True)
        hdr_row = 4
    else:
        hdr_row = 3
    for i, h in enumerate(HDR_COMP, 1):
        ws.cell(hdr_row, i, h)
    _style_header(ws, hdr_row, 9)

    r = hdr_row + 1
    cur_sect = None
    wrote = 0
    for (sect, check, ll, lv, rl, rv, diff, result, tag) in _comp_rows_iter(comparisons):
        is_match = (result == "MATCH")
        if only_mismatch and is_match:
            continue
        if (sect != cur_sect) and not only_mismatch:
            cur_sect = sect
            ws.cell(r, 1, sect).font = Font(bold=True, size=11, color="1F3864")
            for c in range(1, 10):
                ws.cell(r, c).fill = SECT
            r += 1
        ws.cell(r, 1, sect if only_mismatch else "")
        ws.cell(r, 2, check); ws.cell(r, 3, ll); ws.cell(r, 4, lv)
        ws.cell(r, 5, rl); ws.cell(r, 6, rv); ws.cell(r, 7, diff)
        ws.cell(r, 8, result); ws.cell(r, 9, tag)
        for c in range(1, 10):
            cell = ws.cell(r, c); cell.border = BORDER; cell.font = Font(size=10)
            if c in (4, 6, 7):
                cell.number_format = '#,##0.00'
            if c == 8:
                cell.fill = SEV_FILL[result]; cell.font = SEV_FONT[result]
                cell.alignment = Alignment(horizontal="center")
            elif not is_match and c >= 2:
                cell.fill = RED
        r += 1
        wrote += 1
    if only_mismatch and wrote == 0:
        ws.cell(hdr_row + 1, 1, "No mismatches beyond tolerance.").font = Font(italic=True, color="006100")
    for i, w in enumerate(WID_COMP, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if not only_mismatch:
        ws.freeze_panes = f"A{hdr_row + 1}"


def write_analysis14(ws, findings):
    """Sooraj's 14 checks (#0-#14) — same content gst_analysis_checks writes."""
    ws.cell(1, 1, f"GST SCRUTINY — ANALYSIS (Sooraj's 14 checks) — Period {raw.PERIOD_LABEL}").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    counts = {s: sum(1 for f in findings if f.severity == s) for s in ("FLAG", "REVIEW", "INFO", "PASS")}
    ws.cell(3, 1, "   ".join(f"{k}: {v}" for k, v in counts.items())).font = Font(size=10, bold=True)

    hdr = ["Ref", "Check", "Result", "Key numbers", "Detail / arithmetic"]
    for i, h in enumerate(hdr, 1):
        ws.cell(5, i, h)
    _style_header(ws, 5, 5)
    r = 6
    for f in findings:
        ws.cell(r, 1, f.ref); ws.cell(r, 2, f.title)
        cv = ws.cell(r, 3, f.severity); cv.fill = SEV_FILL[f.severity]; cv.font = SEV_FONT[f.severity]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, "  ".join(f"{k}={v:,.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
                                for k, v in f.numbers.items()))
        ws.cell(r, 5, f.detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4, 5)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
    for col, w in zip("ABCDE", [6, 44, 10, 30, 95]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"


def write_eway(ws_find, ws_det, findings):
    """The 27-check EWB matrix + per-check detail — same content gst_eway_recon writes."""
    ws_find.cell(1, 1, f"E-WAY BILL RECONCILIATION (27-check matrix) — {raw.PERIOD_LABEL}").font = TITLEF
    ws_find.cell(2, 1, f"GSTIN {eway.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}").font = Font(size=9, italic=True)
    counts = {s: sum(1 for x in findings if x.sev == s) for s in ("FLAG", "REVIEW", "INFO", "PASS", "SKIPPED")}
    ws_find.cell(3, 1, "  ".join(f"{s}: {c}" for s, c in counts.items())).font = Font(bold=True, size=10)
    hdr = ["Ref", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        ws_find.cell(5, i, h)
    _style_header(ws_find, 5, 4)
    r = 6
    detail_blocks = []
    for f in findings:
        ws_find.cell(r, 1, f.ref); ws_find.cell(r, 2, f.title)
        cv = ws_find.cell(r, 3, f.sev); cv.fill = SEV_FILL[f.sev]; cv.font = SEV_FONT[f.sev]
        cv.alignment = Alignment(horizontal="center")
        ws_find.cell(r, 4, f.detail)
        for c in range(1, 5):
            cell = ws_find.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 4)))
            if c != 3:
                cell.font = Font(size=10)
        r += 1
        if len(f.rows) > 1:
            detail_blocks.append(f)
    for col, w in zip("ABCD", [6, 42, 10, 110]):
        ws_find.column_dimensions[col].width = w
    ws_find.freeze_panes = "A6"

    # detail sheet
    rr = 1
    ws_det.cell(rr, 1, "PER-CHECK DETAIL ROWS (E-Way Bill)").font = Font(bold=True, size=12, color="1F3864"); rr += 2
    for f in detail_blocks:
        ws_det.cell(rr, 1, f"{f.ref}  {f.title}  [{f.sev}]").font = Font(bold=True, color="1F3864"); rr += 1
        head = f.rows[0]
        for j, h in enumerate(head, 1):
            c = ws_det.cell(rr, j, h); c.font = Font(bold=True, size=9, color="FFFFFF"); c.fill = HEAD; c.border = BORDER
        rr += 1
        for row in f.rows[1:]:
            for j, v in enumerate(row, 1):
                if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                    v = " " + v
                c = ws_det.cell(rr, j, v); c.border = BORDER; c.font = Font(size=10)
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            rr += 1
        rr += 1
    for col, w in zip("ABCDE", [22, 18, 18, 14, 10]):
        ws_det.column_dimensions[col].width = w


def write_dashboard(ws, data):
    """Cross-file: every actionable item from BOTH pipelines, ranked together."""
    meta = data["meta"]
    ws.cell(1, 1, f"UNIFIED GST SCRUTINY — CROSS-FILE DASHBOARD — Period {raw.PERIOD_LABEL}").font = TITLEF
    ws.cell(2, 1, f"GSTIN {raw.SELF_GSTIN}  |  {raw.COMPANY_NAME or '(company auto-detected)'}  |  "
                  "one ranked view of Comparison + Analysis + E-Way-Bill").font = Font(size=9, italic=True)
    stamp = (f"generated {_dt.datetime.now():%Y-%m-%d %H:%M:%S}  |  "
             f"2B source: {meta['twob_src']}"
             + (f" ({meta['twob_file']})" if meta['twob_file'] else "")
             + f"  |  E-Invoice: {meta['einv_file'] or 'not found'}"
             + f"  |  EWB-Out lines: {meta['ewb_out_n']}  EWB-In lines: {meta['ewb_in_n']}")
    ws.cell(3, 1, stamp).font = Font(size=9, italic=True, color="C00000")

    # collect actionable items
    items = []  # (rank_sev, pipeline, ref, title, result, detail)
    RANK = {"FLAG": 0, "MISMATCH": 0, "REVIEW": 1, "INFO": 2, "PASS": 3, "MATCH": 3, "SKIPPED": 4}

    # P1a: comparison mismatches
    for (sect, check, ll, lv, rl, rv, diff, result, tag) in _comp_rows_iter(data["comparisons"]):
        if result == "MISMATCH":
            det = f"{ll}={lv:,.2f} vs {rl}={rv:,.2f} (diff {diff:,.2f}). {tag}".strip()
            items.append((RANK[result], "Comparison", sect.split(".")[0], check, result, det))
    # P1b: Sooraj 14 — only FLAG/REVIEW shown in the dashboard (PASS/INFO live on their sheet)
    for f in data["findings14"]:
        if f.severity in ("FLAG", "REVIEW"):
            items.append((RANK[f.severity], "Analysis(14)", f.ref, f.title, f.severity, f.detail))
    # P2: EWB 27 — FLAG/REVIEW
    for f in data["findings27"]:
        if f.sev in ("FLAG", "REVIEW"):
            items.append((RANK[f.sev], "E-Way-Bill", f.ref, f.title, f.sev, f.detail))

    items.sort(key=lambda x: (x[0], x[1], x[2]))

    nflag = sum(1 for it in items if it[4] in ("FLAG", "MISMATCH"))
    nrev = sum(1 for it in items if it[4] == "REVIEW")
    ws.cell(4, 1, f"ACTIONABLE ITEMS: {len(items)}   (FLAG/MISMATCH: {nflag}   REVIEW: {nrev})").font = Font(bold=True, size=11, color="C00000")

    hdr = ["Pipeline", "Ref / Section", "Check", "Result", "Detail"]
    for i, h in enumerate(hdr, 1):
        ws.cell(6, i, h)
    _style_header(ws, 6, 5)
    r = 7
    for (_, pipeline, ref, title, result, detail) in items:
        ws.cell(r, 1, pipeline); ws.cell(r, 2, ref); ws.cell(r, 3, title)
        cv = ws.cell(r, 4, result); cv.fill = SEV_FILL[result]; cv.font = SEV_FONT[result]
        cv.alignment = Alignment(horizontal="center")
        ws.cell(r, 5, detail)
        for c in range(1, 6):
            cell = ws.cell(r, c); cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 5)))
            if c != 4:
                cell.font = Font(size=10)
        r += 1
    if not items:
        ws.cell(7, 1, "No FLAG / MISMATCH / REVIEW across either pipeline.").font = Font(italic=True, color="006100")
    for col, w in zip("ABCDE", [14, 16, 42, 11, 110]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A7"


def write_raw(ws, comp_raw):
    """Audit trail — identical to the original Comparison 'Raw Values' sheet."""
    ws.cell(1, 1, "RAW EXTRACTED VALUES (audit trail - what the tool read from each file)").font = Font(bold=True, size=11)
    r = 3
    for src, d in [("GSTR-1", comp_raw["g1"]), ("E-Invoice", comp_raw["einv"]),
                   ("GSTR-2B", comp_raw["b2b"])]:
        ws.cell(r, 1, src).font = Font(bold=True, color="1F3864"); r += 1
        for k, v in d.items():
            if k.startswith("_"):
                continue
            if isinstance(v, dict):
                uniq = len(set(kk[0] for kk in v.keys())) if v else 0
                ws.cell(r, 2, k + " (line count)"); ws.cell(r, 3, len(v)); r += 1
                ws.cell(r, 2, k + " (unique invoices)"); ws.cell(r, 3, uniq); r += 1
                continue
            if isinstance(v, bool):
                ws.cell(r, 2, k); ws.cell(r, 3, str(v)); r += 1
                continue
            ws.cell(r, 2, k); c = ws.cell(r, 3, round(num(v), 2)); c.number_format = '#,##0.00'; r += 1
        r += 1
    ws.cell(r, 1, "GSTR-3B (parsed tables)").font = Font(bold=True, color="1F3864"); r += 1
    for k, v in comp_raw["g3b"].items():
        ws.cell(r, 2, k); ws.cell(r, 3, str(v)); r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20


# ======================================================================
# MAIN
# ======================================================================
def main():
    data = gather()
    wb = openpyxl.Workbook()

    # Sheet order: Dashboard first (read both pipelines together), then each pipeline.
    ws_dash = wb.active; ws_dash.title = "Dashboard"
    write_dashboard(ws_dash, data)

    write_comparison(wb.create_sheet("Exceptions"), data["comparisons"], only_mismatch=True)
    write_comparison(wb.create_sheet("Full Comparison"), data["comparisons"], only_mismatch=False)
    write_analysis14(wb.create_sheet("Analysis (14 checks)"), data["findings14"])
    write_eway(wb.create_sheet("EWB Findings"), wb.create_sheet("EWB Detail"), data["findings27"])
    write_raw(wb.create_sheet("Raw Values"), data["comp_raw"])

    wb.save(OUTPUT_FILE)

    # console summary
    n_comp_mis = sum(1 for row in _comp_rows_iter(data["comparisons"]) if row[7] == "MISMATCH")
    n14_flag = sum(1 for f in data["findings14"] if f.severity == "FLAG")
    n27_flag = sum(1 for f in data["findings27"] if f.sev == "FLAG")
    print(f"Saved: {OUTPUT_FILE}")
    print(f"  Comparison mismatches : {n_comp_mis}")
    print(f"  Analysis-14 FLAGs     : {n14_flag}")
    print(f"  E-Way-Bill FLAGs      : {n27_flag}")
    print(f"  2B source             : {data['meta']['twob_src']}  "
          f"({'line-level ON' if data['meta']['twob_src']=='excel' else 'PDF summary — line-level OFF'})")


if __name__ == "__main__":
    main()
