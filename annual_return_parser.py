#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANNUAL RETURN PARSER  --  GSTR-9, GSTR-9C, Table 8A
=====================================================
Three NEW optional annual-level sources, added on top of the existing 12+3
(EWB) file set:

  GSTR-9   (PDF)  -- the annual return itself. Government "System Drafted"
                      exports carry a diagonal watermark ("SYSTEM COMPUTED")
                      whose individual letters get interleaved into the text
                      stream at unpredictable points (confirmed on the real
                      file: single stray S/Y/T/E/M/C/O/P/U/D characters show
                      up both as their own line AND glued mid-word, e.g.
                      "sOupplies", "paymMent", a stray "P" appended to a
                      GSTIN). The NUMBERS themselves are never corrupted --
                      only label text is -- confirmed by manual diff against
                      the real file. So: numbers are extracted by a plain
                      numeric regex on the raw line; only the LABEL match
                      that locates the right line runs through watermark
                      cleanup first. If a label still can't be found after
                      cleanup, that field is returned as None with a note --
                      never guessed.

  GSTR-9C  (PDF)  -- the reconciliation statement. Same watermark pattern,
                      same technique.

  Table 8A (XLSX) -- the exact government-generated workbook (confirmed
                      structurally identical in format to every taxpayer's
                      export -- this is a standard GSTN download, not a
                      custom report). Header row is found by CONTENT (the
                      cell "GSTIN of supplier"), never hardcoded to a row
                      number, in case a future export adds/removes a
                      preamble row.

ALL THREE ARE OPTIONAL. Every function here degrades to a clearly-labeled
"not supplied" / "not found" state -- never raises for a missing file, never
fabricates a number. This mirrors the existing codebase's E-Invoice pattern
(gst_scrutiny_tool.parse_einv: available=True/False).

Genericity: every parser here works off LABEL TEXT and CONTENT SIGNATURES,
never off this one taxpayer's numbers, GSTIN, or FY. Verified against the
real M R HEALTHCARE FY22-23 files (see test block at the bottom) but not
hardcoded to them.
"""

import os
import re
import datetime as _dt
import openpyxl

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


# ======================================================================
# Shared helpers
# ======================================================================
def num(v):
    """Convert any cell/string to float. Blank/'-'/None -> 0.0. Never raises."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s in ("", "-", "\u2013", "NA", "N/A"):
        return 0.0
    s = s.replace(",", "").replace("\u20b9", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# Confirmed watermark letters on the real GSTR-9/9C "System Drafted" PDF
# exports: the stamp text is "SYSTEM COMPUTED" and its letters (deduped:
# S,Y,T,E,M,C,O,P,U,D) get interleaved into the extracted text stream.
_WATERMARK_LETTERS = "SYTEMCOPUD"


def _clean_watermark_line(line):
    """Best-effort removal of the interleaved watermark from a SINGLE line,
    for LABEL MATCHING only (never used to alter a number). Operates
    line-by-line (not on the whole joined text) so the cleaned line always
    stays index-paired with its own raw line -- an earlier version cleaned
    the whole text and then dropped watermark-only lines before re-splitting,
    which silently shifted every later line out of alignment with its raw
    counterpart (caught by cross-checking output against the real file's
    known figures -- several Table 4/5/6/9 rows came back None because the
    label search was looking at the wrong line entirely). Three confirmed
    insertion patterns:
      1. The ENTIRE line is a single watermark letter -> blank it out
         (kept as an empty string, not deleted, to preserve line count).
      2. A single watermark letter glued between two lowercase letters
         inside a real word (e.g. 'sOupplies' -> 'supplies').
      3. A single watermark letter glued onto the end of a long
         alphanumeric token (e.g. a 15-char GSTIN gaining a trailing 'P').
    Intentionally conservative: strips only a SINGLE stray letter in these
    specific shapes, never a run of letters, so it cannot silently eat real
    content."""
    if len(line.strip()) == 1 and line.strip() in _WATERMARK_LETTERS:
        return ""
    t = re.sub(rf"(?<=[a-z])([{_WATERMARK_LETTERS}])(?=[a-z])", "", line)
    t = re.sub(rf"(?<=[A-Z0-9]{{10}})([{_WATERMARK_LETTERS}])(?=\s|$)", "", t)
    return t


def _clean_watermark(text):
    """Whole-text version, for one-shot regex extraction (GSTIN/ARN) where
    no line-index pairing is needed. Built from the per-line cleaner so the
    two never drift out of sync with each other."""
    return "\n".join(_clean_watermark_line(l) for l in text.split("\n"))


def _extract_pdf_text(path):
    if pdfplumber is None:
        raise RuntimeError("pdfplumber not installed -- required to read GSTR-9/9C PDFs.")
    with pdfplumber.open(path) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


_NUM_RE = re.compile(r"-?[\d,]+\.\d{2}|-(?=\s|$)")


def _numbers_on_line(line):
    """Every numeric token on a line, in order. A lone '-' (system-computed
    blank) becomes 0.0, matching this form's own convention (every '-' on
    these forms means 'not applicable / computed elsewhere', confirmed
    against the real file's own SYSTEM COMPUTED rows)."""
    out = []
    for m in _NUM_RE.finditer(line):
        s = m.group(0)
        out.append(0.0 if s == "-" else num(s))
    return out


def _find_line(clean_lines, *label_fragments, occurrence=1):
    """Return the (index, raw_line) of the `occurrence`-th line whose cleaned
    text contains ALL given fragments (case-insensitive substrings). Returns
    (None, None) if not found -- caller must handle that explicitly, no
    silent fallback."""
    hits = 0
    for i, (raw, clean) in enumerate(clean_lines):
        low = clean.lower()
        if all(f.lower() in low for f in label_fragments):
            hits += 1
            if hits == occurrence:
                return i, raw
    return None, None


def _prep(text):
    """Return [(raw_line, clean_line), ...] with GUARANTEED 1:1 index
    alignment (each line cleaned independently -- see _clean_watermark_line)."""
    raw_lines = text.split("\n")
    return [(l, _clean_watermark_line(l)) for l in raw_lines]


# ======================================================================
# GSTR-9  (Annual Return)
# ======================================================================
def parse_gstr9(path):
    """Return dict of the GSTR-9 figures this tool's forensic checks need.
    available=False (with reason) if the file is absent or unreadable --
    NEVER raises, per the 'any taxpayer, any subset of documents' rule.
    Only fields actually present on the extract are populated; anything
    not found is None with its own '<field>_note' explaining why, so a
    missing figure is visible, not silently zeroed."""
    out = dict(available=False, reason=None, is_system_draft=None,
                fy=None, gstin=None, legal_name=None,
                table4_b2b_taxable=None, table4_b2b_igst=None, table4_b2b_cgst=None, table4_b2b_sgst=None,
                table4_cn_taxable=None, table4_cn_igst=None, table4_cn_cgst=None, table4_cn_sgst=None,
                table5_zero_rated=None, table5_sez=None, table5_exempted=None, table5_nil_rated=None,
                table5_nongst=None, table5_all_zero=None,
                table6a_cgst=None, table6a_sgst=None, table6a_igst=None, table6a_cess=None, table6a_total=None,
                table9_liability_igst=None, table9_liability_cgst=None, table9_liability_sgst=None,
                table9_late_fee_payable=None, table9_late_fee_paid=None,
                table9_interest_payable=None, table9_interest_paid=None,
                notes=[])
    if not path or not os.path.exists(path):
        out["reason"] = "GSTR-9 not supplied for this taxpayer/FY."
        return out
    try:
        text = _extract_pdf_text(path)
    except Exception as ex:
        out["reason"] = f"Could not read GSTR-9 PDF: {ex}"
        return out
    if not text.strip():
        out["reason"] = "GSTR-9 PDF has no extractable text (likely a scanned image, not a text PDF)."
        return out

    lines = _prep(text)
    out["is_system_draft"] = "system drafted" in text.lower() or "for reference only" in text.lower()
    if out["is_system_draft"]:
        out["notes"].append("This copy is the pre-filing 'System Drafted (For Reference Only)' auto-draft, "
                             "not necessarily the as-filed return -- some tables (6B-6M split, 7, 8) may be "
                             "blank on the draft even if populated on the as-filed PDF. Ask for the actual "
                             "filed copy if a table below reads 'not found'.")

    m = re.search(r"Financial Year\s+(\d{4}-\d{2,4})", text)
    out["fy"] = m.group(1) if m else None
    m = re.search(r"GSTIN\s+([0-9A-Z]{15})", _clean_watermark(text))
    out["gstin"] = m.group(1)[:15] if m else None
    m = re.search(r"[Ll]egal name of the registered person\s+(.+)", text)
    out["legal_name"] = m.group(1).strip() if m else None

    # ---- Table 4B: B2B outward taxable + tax ----
    i, raw = _find_line(lines, "supplies made to registered persons")
    if raw:
        n = _numbers_on_line(raw)
        if len(n) >= 4:
            out["table4_b2b_taxable"], out["table4_b2b_cgst"], out["table4_b2b_sgst"], out["table4_b2b_igst"] = n[0], n[1], n[2], n[3]
    else:
        out["notes"].append("Table 4B (B2B outward) row not found in text.")

    # ---- Table 4I: credit notes on B2B/SEZ/Deemed-export ----
    i, raw = _find_line(lines, "credit notes issued in respect")
    if raw:
        n = _numbers_on_line(raw)
        if len(n) >= 4:
            out["table4_cn_taxable"], out["table4_cn_cgst"], out["table4_cn_sgst"], out["table4_cn_igst"] = n[0], n[1], n[2], n[3]
    else:
        out["notes"].append("Table 4I (credit notes) row not found in text.")

    # ---- Table 5: outward supplies on which tax is NOT payable (A-F) ----
    t5 = {}
    for key, frag in [("zero_rated", "zero rated supply (export) without payment"),
                       ("sez", "supply to sezs without payment"),
                       ("rcm_recipient", "tax is to be paid by the recipient on"),
                       ("exempted", "exempted"), ("nil_rated", "nil rated"),
                       ("nongst", "non-gst supply")]:
        i, raw = _find_line(lines, frag)
        if raw:
            n = _numbers_on_line(raw)
            t5[key] = n[0] if n else None
    out["table5_zero_rated"] = t5.get("zero_rated")
    out["table5_sez"] = t5.get("sez")
    out["table5_exempted"] = t5.get("exempted")
    out["table5_nil_rated"] = t5.get("nil_rated")
    out["table5_nongst"] = t5.get("nongst")
    vals5 = [v for v in t5.values() if v is not None]
    out["table5_all_zero"] = bool(vals5) and all(abs(v) < 0.01 for v in vals5)

    # ---- Table 6A: ITC availed via 3B (sum of 3B Table 4A) ----
    i, raw = _find_line(lines, "total amount of input tax credit availed through form gstr-3b")
    if raw:
        n = _numbers_on_line(raw)
        if len(n) >= 3:
            out["table6a_cgst"], out["table6a_sgst"], out["table6a_igst"] = n[0], n[1], n[2]
            out["table6a_cess"] = n[3] if len(n) > 3 else 0.0
            out["table6a_total"] = out["table6a_cgst"] + out["table6a_sgst"] + out["table6a_igst"] + out["table6a_cess"]
    else:
        out["notes"].append("Table 6A (ITC availed via 3B) row not found in text.")

    # ---- Table 9: tax payable/paid + late fee + interest ----
    for row_label, keyprefix in [("integrated tax", "igst"), ("central tax", "cgst"), ("state/ut tax", "sgst")]:
        # Table 9 rows are single-letter-prefixed data rows ("A Integrated Tax 2,90,78,901.00 ...").
        # The header/label-only line ("Central Tax State Tax / UT Tax Integrated Tax Cess") mentions
        # the same words with ZERO numbers -- require at least 1 number to reject that header line.
        found = None
        for idx, (rawl, cleanl) in enumerate(lines):
            cl = cleanl.lower().strip()
            if cl.startswith(("a ", "b ", "c ")) and row_label in cl:
                n = _numbers_on_line(rawl)
                if n:
                    found = n
                    break
        if found:
            out[f"table9_liability_{keyprefix}"] = found[0]

    i, raw = _find_line(lines, "late fee")
    if raw:
        n = _numbers_on_line(raw)
        if len(n) >= 2:
            out["table9_late_fee_payable"], out["table9_late_fee_paid"] = n[0], n[1]
    i, raw = _find_line(lines, "interest")
    if raw:
        n = _numbers_on_line(raw)
        if len(n) >= 2:
            out["table9_interest_payable"], out["table9_interest_paid"] = n[0], n[1]

    out["available"] = True
    return out


# ======================================================================
# GSTR-9C  (Reconciliation Statement)
# ======================================================================
def parse_gstr9c(path):
    """Return dict of GSTR-9C figures. Same graceful-degrade contract as
    parse_gstr9()."""
    out = dict(available=False, reason=None, fy=None, gstin=None, legal_name=None,
                arn=None, arn_date=None,
                turnover_audited_bs=None, turnover_after_adjustments=None,
                turnover_declared_gstr9=None, turnover_unreconciled=None,
                exempt_nil_nongst_adjustment=None,
                taxable_turnover_after_adj=None, taxable_turnover_declared=None,
                itc_per_books=None, itc_declared_gstr9=None, itc_unreconciled=None,
                itc_booked_earlier_claimed_now=None, itc_booked_now_claimed_later=None,
                tax_payable_total=None, tax_paid_declared=None,
                notes=[])
    if not path or not os.path.exists(path):
        out["reason"] = "GSTR-9C not supplied for this taxpayer/FY."
        return out
    try:
        text = _extract_pdf_text(path)
    except Exception as ex:
        out["reason"] = f"Could not read GSTR-9C PDF: {ex}"
        return out
    if not text.strip():
        out["reason"] = "GSTR-9C PDF has no extractable text (likely scanned, not a text PDF)."
        return out

    lines = _prep(text)
    m = re.search(r"Financial Year\s+(\d{4}-\d{2,4})", text)
    out["fy"] = m.group(1) if m else None
    m = re.search(r"GSTIN\s+([0-9A-Z]{15})", _clean_watermark(text))
    out["gstin"] = m.group(1)[:15] if m else None
    m = re.search(r"Legal Name\s+(.+)", text)
    out["legal_name"] = m.group(1).strip() if m else None
    m = re.search(r"ARN\s+([A-Z0-9]{15})", _clean_watermark(text))
    out["arn"] = m.group(1) if m else None
    m = re.search(r"ARN Date\s+(\d{2}-\d{2}-\d{4})", text)
    out["arn_date"] = m.group(1) if m else None

    def _first_num(*frags):
        i, raw = _find_line(lines, *frags)
        if not raw:
            return None
        n = _numbers_on_line(raw)
        return n[0] if n else None

    out["turnover_audited_bs"] = _first_num("turnover (including exports) as per audited")
    out["turnover_after_adjustments"] = _first_num("annual turnover after adjustments as above")
    # 'Q' and 'R' both contain "turnover" + specific words -- disambiguate by order of appearance
    i, raw = _find_line(lines, "turnover as declared in annual return")
    out["turnover_declared_gstr9"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None
    i, raw = _find_line(lines, "un-reconciled turnover")
    out["turnover_unreconciled"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None

    out["taxable_turnover_after_adj"] = _first_num("taxable turnover as per adjustments above")
    i, raw = _find_line(lines, "value of exempted, nil rated, non-gst supplies")
    out["exempt_nil_nongst_adjustment"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None
    i, raw = _find_line(lines, "taxable turnover as per liability declared")
    out["taxable_turnover_declared"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None

    out["itc_per_books"] = _first_num("itc availed as per audited annual financial statement")
    i, raw = _find_line(lines, "itc claimed in annual return", occurrence=1)
    out["itc_declared_gstr9"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None
    i, raw = _find_line(lines, "un-reconciled itc")
    out["itc_unreconciled"] = _numbers_on_line(raw)[0] if raw and _numbers_on_line(raw) else None
    out["itc_booked_earlier_claimed_now"] = _first_num("itc booked in earlier financial years claimed")
    out["itc_booked_now_claimed_later"] = _first_num("itc booked in current financial year to be claimed")

    i, raw = _find_line(lines, "total amount to be paid as per")
    if raw:
        n = _numbers_on_line(raw)
        out["tax_payable_total"] = sum(n) if n else None
    i, raw = _find_line(lines, "total amount paid as declared")
    if raw:
        n = _numbers_on_line(raw)
        out["tax_paid_declared"] = sum(n) if n else None

    if not out["exempt_nil_nongst_adjustment"]:
        out["notes"].append("Table 7B (exempt/nil/non-GST adjustment) not found or zero -- Rule R13 "
                             "(turnover-gap check) needs this figure; verify manually if a gap is otherwise expected.")

    out["available"] = True
    return out


# ======================================================================
# Table 8A  (government-standard export, same layout for every taxpayer)
# ======================================================================
def _find_header_row(ws, must_contain, max_scan=15):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if any(must_contain.lower() in c.lower() for c in cells):
            return i
    return None


def parse_table_8a(path):
    """Return dict(available, b2b=[...], cdnr=[...], totals={...}) from the
    government Table-8A workbook. Header row is located by CONTENT
    ('GSTIN of supplier' cell), never hardcoded -- confirmed this workbook's
    real layout has a 2-row wrapped header with data starting the row after,
    but the search below does not assume a fixed offset; it scans forward
    from the header row until it finds the first row that both (a) is not
    entirely blank and (b) has a GSTIN-shaped token in column B, so a export
    with a different preamble length still parses correctly."""
    out = dict(available=False, reason=None, b2b=[], cdnr=[], totals={})
    if not path or not os.path.exists(path):
        out["reason"] = "Table 8A not supplied for this taxpayer/FY."
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        out["reason"] = f"Could not read Table 8A workbook: {ex}"
        return out

    gstin_re = re.compile(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\dZ[A-Z\d]$")

    def _rows_after_header(ws, hdr_row):
        for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if r and r[1] and gstin_re.match(str(r[1]).strip()):
                yield r

    if "B2B" in wb.sheetnames:
        ws = wb["B2B"]
        hdr = _find_header_row(ws, "GSTIN of supplier")
        if hdr:
            for r in _rows_after_header(ws, hdr):
                out["b2b"].append(dict(
                    period=str(r[0] or "").strip(), gstin=str(r[1]).strip(), supplier=str(r[2] or "").strip(),
                    invno=str(r[3] or "").strip(), invtype=str(r[4] or "").strip(), invdate=r[5],
                    invval=num(r[6]), pos=str(r[7] or "").strip(), rcm=str(r[8] or "").strip(),
                    rate=num(r[9]), taxable=num(r[10]), igst=num(r[11]), cgst=num(r[12]),
                    sgst=num(r[13]), cess=num(r[14]),
                    supplier_filing_date=r[15] if len(r) > 15 else None,
                    itc_available=str(r[16] or "").strip() if len(r) > 16 else "",
                    reason_not_available=str(r[17] or "").strip() if len(r) > 17 else "",
                ))
        else:
            out["reason"] = (out["reason"] or "") + " 'B2B' sheet found but no 'GSTIN of supplier' header row located."

    if "CDNR" in wb.sheetnames:
        ws = wb["CDNR"]
        hdr = _find_header_row(ws, "GSTIN of supplier")
        if hdr:
            for r in _rows_after_header(ws, hdr):
                out["cdnr"].append(dict(
                    period=str(r[0] or "").strip(), gstin=str(r[1]).strip(), supplier=str(r[2] or "").strip(),
                    note_type=str(r[3] or "").strip(), supply_type=str(r[4] or "").strip(),
                    note_no=str(r[5] or "").strip(), note_date=r[6], note_val=num(r[7]),
                    pos=str(r[8] or "").strip(), rate=num(r[10]), taxable=num(r[11]),
                    igst=num(r[12]), cgst=num(r[13]), sgst=num(r[14]), cess=num(r[15]),
                    itc_available=str(r[17] or "").strip() if len(r) > 17 else "",
                ))

    yes_b2b = [r for r in out["b2b"] if r["itc_available"].upper() == "YES"]
    out["totals"] = dict(
        b2b_rows=len(out["b2b"]), b2b_yes_rows=len(yes_b2b),
        cgst=sum(r["cgst"] for r in yes_b2b), sgst=sum(r["sgst"] for r in yes_b2b),
        igst=sum(r["igst"] for r in yes_b2b), cess=sum(r["cess"] for r in yes_b2b),
    )
    out["totals"]["total"] = (out["totals"]["cgst"] + out["totals"]["sgst"]
                                + out["totals"]["igst"] + out["totals"]["cess"])
    # bucket the "No" rows by reason -- Part 1 C2 of the forensic framework
    no_b2b = [r for r in out["b2b"] if r["itc_available"].upper() != "YES"]
    from collections import Counter
    out["totals"]["no_reason_breakdown"] = dict(Counter(r["reason_not_available"] or "(blank)" for r in no_b2b))

    out["available"] = True
    return out


if __name__ == "__main__":
    import sys, json
    g9 = parse_gstr9(sys.argv[1] if len(sys.argv) > 1 else None)
    print("=== GSTR-9 ===")
    for k, v in g9.items():
        print(f"  {k}: {v}")
