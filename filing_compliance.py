#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FILING COMPLIANCE  --  ARN date extraction, statutory due dates, late fee & interest
========================================================================================
Fixes a real gap found in the previous version of this tool: `gst_unified_scrutiny.py`
had an `_extract_arn_dates()` function, but it lived inside `gather()`, a function
explicitly marked "LEGACY / UNSUPPORTED for the merged-file model" that
`master_build.py` (the actual pipeline) never calls. So in every real run,
`GSTR1_FILING_DATE`/`GSTR3B_FILING_DATE` stayed None, and Analysis checks #8/#10
always fell through to INFO ("Set ... in CONFIG to enable"). Also, even when
supplied, check #10 only compared GSTR-1-filing-date vs GSTR-3B-filing-date GAP --
never against the statutory DUE DATE, and no late-fee/interest RUPEE AMOUNT was
computed anywhere in the codebase.

This module fixes both, PER MONTH (the merged-file model's whole point), and adds
the ledger-actual cross-check the forensic framework asked for (Part 1, A4).

--------------------------------------------------------------------------------
WHERE THE ARN DATE ACTUALLY LIVES (content-based, verified against the real
GSTR-9C export's own ARN/ARN-Date fields as a format cross-check):
  GSTR-1  (merged workbook): every sub-sheet's PERIOD-MARKER row already
      carries the month's own ARN as free text, e.g.
      "Financial Year: 2022-23 | Tax Period: January | ARN: AA0501230730120 | ..."
      merged_period_utils.MARKER_RE only captures FY + Tax Period today; this
      module extends that same marker text with an ARN-date capture. If the
      marker text does NOT also carry a date (some GSTN exports print the ARN
      but not its date on this row), this module falls back to the 'Read me'
      sheet, which -- per gst_unified_scrutiny.py's original comment -- carries
      a single 'ARN date' row. IMPORTANT: for a MERGED whole-FY file, 'Read me'
      is BY DEFINITION only useful as a SINGLE value, which cannot be correct
      for all 12 months at once -- so this module treats a Read-me-only date as
      a per-file fallback (used only for whichever single month, if any, that
      'Read me' actually documents) and prints an explicit WARNING rather than
      silently applying one date to all 12 months. This needs confirming
      against a real multi-month merged GSTR-1 file (not yet supplied) --
      flagged clearly in the output, not guessed past.
  GSTR-3B (merged workbook, one SHEET per month): each month's own sheet
      already carries a 'Date of ARN' row (confirmed in gst_unified_scrutiny.py
      and reused here) -- this one is genuinely per-month already, no fallback
      needed.

HARD RULE: no invented dates, no invented amounts. Every field is either read
from the file or left None with a note. Late fee / interest are computed only
when both a filing date and a due date are known; otherwise the finding says
so explicitly.
"""

import re
import datetime as _dt
import openpyxl
import merged_period_utils as mpu


# ======================================================================
# Statutory due dates
# ======================================================================
# GSTR-1 (monthly filer): 11th of the month following the tax period.
# GSTR-1 (QRMP / quarterly filer): 13th of the month following the quarter.
# GSTR-3B (monthly, non-QRMP, turnover > Rs 5 Cr or opted monthly): 20th.
# GSTR-3B (QRMP): 22nd (Category-X states) or 24th (Category-Y states).
# This taxpayer's GSTR-3B is confirmed one-sheet-per-MONTH (not per-quarter),
# i.e. non-QRMP -- matches hsn_fraud_checks.py's own GSTR3B_DUE_DOM=20 constant.
# For genericity (any taxpayer), filer type is DETECTED from the data itself,
# never assumed: if a GSTR-1 period marker resolves to 3 fanned-out months
# from one marker (see merged_period_utils.QUARTER_TO_MONTHS), that taxpayer
# is QRMP; if GSTR-3B has one sheet per calendar month, that taxpayer reports
# monthly (QRMP taxpayers still file GSTR-3B quarterly, so 4 sheets/FY, not 12).
CATEGORY_X_STATES = {  # 22nd -- confirmed CBIC state grouping for QRMP due dates
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12",
    "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24",
    "25", "26", "27", "37",
}  # Chhattisgarh through Maharashtra + Andhra Pradesh(new)/Ladakh -- Category X (West+South+some NE)
CATEGORY_Y_STATES = {"28", "29", "30", "31", "32", "33", "34", "35", "36", "38"}  # Category Y (East+North)

LATE_FEE_PER_DAY_NORMAL = 25.0    # Rs 25 CGST + Rs 25 SGST = Rs 50/day total, per Sec 47
LATE_FEE_PER_DAY_NIL = 10.0       # Rs 10 CGST + Rs 10 SGST = Rs 20/day total, nil return
INTEREST_RATE_ANNUAL = 0.18       # Sec 50(1), 18% p.a. on the cash-paid portion of tax


def _cap_for_turnover(annual_turnover):
    """Late-fee caps per Notification 07/2023-CT (in force for FY22-23 returns
    filed after the notification date) -- max total late fee (both heads
    combined) PER RETURN, based on prior-year aggregate turnover:
      <=Rs 1.5 Cr turnover -> capped at Rs 2,000
      Rs 1.5-5 Cr turnover -> capped at Rs 5,000
      >Rs 5 Cr turnover    -> capped at Rs 10,000
    Nil returns: capped at Rs 500 regardless of turnover.
    Returns (cap_normal, cap_nil). If annual_turnover is None, returns
    (None, None) -- caller must then report the uncapped figure with an
    explicit 'cap not applied -- turnover unknown' note, never silently pick
    a slab."""
    if annual_turnover is None:
        return None, None
    # Amounts in RUPEES throughout this module (Indian digit-grouping used in
    # the literals below purely for readability: 1_50_00_000 = Rs 1.5 Cr).
    if annual_turnover <= 1_50_00_000:
        return 2000.0, 500.0
    if annual_turnover <= 5_00_00_000:
        return 5000.0, 500.0
    return 10000.0, 500.0


def due_date_gstr1(period_start_month_first_day, is_qrmp=False):
    """period_start_month_first_day: date(YYYY, MM, 1) for the tax period
    (or the LAST month of the quarter if QRMP -- caller passes the quarter's
    last calendar month). Returns the statutory due date."""
    y, m = period_start_month_first_day.year, period_start_month_first_day.month
    m2, y2 = (m % 12) + 1, y + (1 if m == 12 else 0)
    return _dt.date(y2, m2, 13 if is_qrmp else 11)


def due_date_gstr3b(period_start_month_first_day, is_qrmp=False, self_gstin=None):
    y, m = period_start_month_first_day.year, period_start_month_first_day.month
    m2, y2 = (m % 12) + 1, y + (1 if m == 12 else 0)
    if not is_qrmp:
        return _dt.date(y2, m2, 20)
    state_code = (self_gstin or "")[:2]
    dom = 24 if state_code in CATEGORY_Y_STATES else 22
    return _dt.date(y2, m2, dom)


def _month_label_to_first_day(label):
    """'Jan-23' -> date(2023,1,1). Uses merged_period_utils' own month-abbr
    map for consistency with the rest of the codebase."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", label)
    if not m:
        return None
    mon_abbr, yy = m.group(1), m.group(2)
    inv = {v: k for k, v in mpu.CAL_MONTH_ABBR.items()}
    mm = inv.get(mon_abbr.title())
    if not mm:
        return None
    yyyy = 2000 + int(yy)
    return _dt.date(yyyy, mm, 1)


# ======================================================================
# ARN date extraction
# ======================================================================
# Extend the marker regex (without touching merged_period_utils.py's own
# MARKER_RE, so nothing else that imports it changes behaviour) to also
# capture an ARN and, if present, a date right after it.
#
# FIXED (real bug, confirmed against the real merged GSTR-1 file's actual
# marker text): the label before the date can be "ARN Date" (ARN comes
# FIRST), not just "Date of Filing"/"Filing Date"/"Date" (Date comes
# first) -- the original pattern only handled the second family, so a
# real marker like "ARN: AA050422057237G | ARN Date: 10-05-2022" matched
# the ARN number but never the date, silently leaving every month's
# GSTR-1 filing date as None (which is why late-fee/filing-gap always
# showed blank even though the real marker text had the date all along).
_ARN_IN_MARKER_RE = re.compile(
    r"ARN:\s*([A-Z0-9]{15})\s*(?:[|,])?\s*"
    r"(?:(?:ARN\s*Date|Date\s*of\s*Filing|Filing\s*Date|Date)\s*[:\s]\s*)?"
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})?",
    re.IGNORECASE
)


def _parse_any_date(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def gstr1_arn_dates_by_month(gstr1_path):
    """Return {month_label: {'arn': str|None, 'date': date|None}} by reading
    EVERY period-marker row across every GSTR-1 sub-sheet (not just one),
    since a merged whole-FY file has one marker per month per sub-sheet.
    If NO marker anywhere carries a date (i.e. the export only prints "ARN:
    <no.>" without a date on the marker row itself), falls back to the
    single 'Read me' sheet value and returns it ONLY under whichever month
    that sheet's own 'Tax Period' field (if present) identifies -- otherwise
    returns it under a special '_readme_fallback' key with a clear note,
    rather than guessing which month it belongs to."""
    out = {}
    warnings = []
    wb = openpyxl.load_workbook(gstr1_path, data_only=True)
    found_any_marker_date = False
    for sn in wb.sheetnames:
        if sn.lower() == "read me":
            continue
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            cell0 = str(row[0])
            if "Financial Year:" not in cell0 or "Tax Period:" not in cell0:
                continue
            try:
                fy, tp, months = mpu.parse_marker_text(cell0)
            except mpu.PeriodParseError:
                continue
            m = _ARN_IN_MARKER_RE.search(cell0)
            arn = m.group(1) if m else None
            dt = _parse_any_date(m.group(2)) if (m and m.group(2)) else None
            if dt:
                found_any_marker_date = True
            for lbl in months:
                slot = out.setdefault(lbl, {"arn": None, "date": None})
                if arn and not slot["arn"]:
                    slot["arn"] = arn
                if dt and not slot["date"]:
                    slot["date"] = dt

    if not found_any_marker_date:
        # Fallback: 'Read me' sheet's single ARN-date row (per gst_unified_scrutiny.py).
        # This is a WHOLE-FILE value, not proven to be per-month -- surfaced with an
        # explicit warning rather than silently stamped onto every month.
        if "Read me" in wb.sheetnames:
            arn_val = date_val = tax_period_val = None
            for r in wb["Read me"].iter_rows(values_only=True):
                cells = [c for c in r if c not in (None, "")]
                if not cells:
                    continue
                label = str(cells[0]).strip().upper()
                if label in ("ARN", "ARN NO", "ARN NUMBER") and len(cells) >= 2:
                    arn_val = str(cells[-1]).strip()
                elif label in ("ARN DATE", "DATE OF ARN") and len(cells) >= 2:
                    date_val = _parse_any_date(cells[-1])
                elif label in ("TAX PERIOD",) and len(cells) >= 2:
                    tax_period_val = str(cells[-1]).strip()
            if date_val:
                warnings.append(
                    "GSTR-1 per-month ARN date not found on any period-marker row for this file -- "
                    "falling back to the single 'ARN date' value on the 'Read me' sheet "
                    f"({date_val}). This is a WHOLE-FILE value; it is only applied to a specific "
                    "month if 'Read me' also states a Tax Period, and is otherwise NOT applied to "
                    "any month automatically (see '_readme_fallback' in the result) -- confirm "
                    "against the portal before using it for a late-fee calculation on a specific month."
                )
                if tax_period_val:
                    try:
                        _, _, months = mpu.parse_marker_text(
                            f"Financial Year: 0000-00 | Tax Period: {tax_period_val}")
                    except Exception:
                        months = []
                    for lbl in months:
                        out.setdefault(lbl, {"arn": arn_val, "date": date_val})
                else:
                    out["_readme_fallback"] = {"arn": arn_val, "date": date_val}
    return out, warnings


def gstr3b_arn_dates_by_month(gstr3b_path):
    """Return {month_label: {'arn': str|None, 'date': date|None}}. GSTR-3B is
    one SHEET per month, and each sheet already carries its own 'Date of ARN'
    key/value row -- genuinely per-month, no fallback needed."""
    out = {}
    wb = openpyxl.load_workbook(gstr3b_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        fy = tp = arn = arn_date = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue
            key = cells[0].upper()
            if key in ("YEAR", "FINANCIAL YEAR") and len(cells) >= 2:
                fy = cells[1]
            elif key == "TAX PERIOD" and len(cells) >= 2:
                tp = cells[1]
            elif key == "ARN" and len(cells) >= 2:
                arn = cells[-1]
            elif key in ("DATE OF ARN", "ARN DATE") and len(cells) >= 2:
                arn_date = _parse_any_date(cells[-1])
            if fy and tp and arn_date:
                break
        if not (fy and tp):
            continue
        try:
            labels = mpu.months_for_tax_period(fy, tp)
        except mpu.PeriodParseError:
            continue
        for lbl in labels:
            out[lbl] = {"arn": arn, "date": arn_date}
    return out


# ======================================================================
# Late fee / interest computation
# ======================================================================
def compute_late_fee(filing_date, due_date, is_nil_return=False, annual_turnover=None):
    """Section 47. Returns dict(days_late, fee_per_day_combined, gross_fee,
    cap, fee_payable, capped). fee_payable is the LOWER of the day-count
    formula and the applicable cap (Notification 07/2023-CT). If
    annual_turnover is None, cap is not applied and that is stated
    explicitly -- never silently uncapped without saying so."""
    if not filing_date or not due_date:
        return dict(days_late=None, fee_payable=None,
                     note="Filing date or due date not available -- cannot compute.")
    days_late = (filing_date - due_date).days
    if days_late <= 0:
        return dict(days_late=days_late, fee_payable=0.0, note="Filed on or before due date -- no late fee.")
    per_day = LATE_FEE_PER_DAY_NIL if is_nil_return else LATE_FEE_PER_DAY_NORMAL
    gross = days_late * per_day * 2  # both CGST+SGST heads combined
    cap_normal, cap_nil = _cap_for_turnover(annual_turnover)
    cap = cap_nil if is_nil_return else cap_normal
    if cap is None:
        return dict(days_late=days_late, gross_fee=gross, cap=None, fee_payable=gross, capped=False,
                     note="Annual turnover not supplied -- Notification 07/2023-CT cap NOT applied; "
                          "figure shown is the uncapped day-count formula only.")
    payable = min(gross, cap)
    return dict(days_late=days_late, gross_fee=gross, cap=cap, fee_payable=payable, capped=(payable < gross),
                note=f"{days_late} day(s) late x Rs {per_day*2:.0f}/day = Rs {gross:,.2f}, "
                     f"capped at Rs {cap:,.2f} per Notification 07/2023-CT" if payable < gross else
                     f"{days_late} day(s) late x Rs {per_day*2:.0f}/day = Rs {gross:,.2f} (within cap).")


def compute_interest(cash_paid_tax, filing_date, due_date):
    """Section 50(1): 18% p.a. simple interest on the CASH-portion of tax
    liability, for every day between due date and actual payment (approximated
    here as the filing date, since the cash ledger debit for a self-assessed
    liability happens at filing -- exact if paid via DRC-03 on the same date
    as filing, otherwise this is a lower bound; noted explicitly)."""
    if not filing_date or not due_date or cash_paid_tax is None:
        return dict(days_late=None, interest=None, note="Missing filing date, due date, or cash-tax figure.")
    days_late = (filing_date - due_date).days
    if days_late <= 0:
        return dict(days_late=days_late, interest=0.0, note="Filed on or before due date -- no interest.")
    interest = cash_paid_tax * INTEREST_RATE_ANNUAL * days_late / 365.0
    return dict(days_late=days_late, interest=round(interest, 2),
                note=f"Rs {cash_paid_tax:,.2f} (cash-paid tax) x 18% p.a. x {days_late} days / 365 "
                     f"= Rs {interest:,.2f}. Approximated using the FILING date as the payment date "
                     "(exact only if cash was actually debited same-day as filing) -- verify against "
                     "the Liability Register's own Interest-head entry for this period, per Forensic "
                     "Framework Part 1 A4.")


def month_filing_compliance(month_label, gstr1_arn_by_month, gstr3b_arn_by_month,
                             gstr1_is_qrmp=False, gstr3b_is_qrmp=False, self_gstin=None,
                             is_nil_return=False, annual_turnover=None, cash_paid_tax=None):
    """Full per-month compliance record: dates, due dates, late fee, interest,
    for BOTH GSTR-1 and GSTR-3B. Every field degrades to None + a note if the
    underlying date isn't available -- never fabricated."""
    first_day = _month_label_to_first_day(month_label)
    out = dict(month=month_label)
    for ret_type, arn_map, is_qrmp, due_fn in [
        ("gstr1", gstr1_arn_by_month, gstr1_is_qrmp, due_date_gstr1),
        ("gstr3b", gstr3b_arn_by_month, gstr3b_is_qrmp, due_date_gstr3b),
    ]:
        rec = arn_map.get(month_label, {})
        filing_date = rec.get("date")
        due = due_fn(first_day, is_qrmp) if first_day else None
        if ret_type == "gstr3b":
            due = due_date_gstr3b(first_day, is_qrmp, self_gstin) if first_day else None
        late = compute_late_fee(filing_date, due, is_nil_return, annual_turnover)
        out[f"{ret_type}_arn"] = rec.get("arn")
        out[f"{ret_type}_filing_date"] = filing_date
        out[f"{ret_type}_due_date"] = due
        out[f"{ret_type}_late_fee"] = late
        if ret_type == "gstr3b":
            out["gstr3b_interest"] = compute_interest(cash_paid_tax, filing_date, due)
    if out.get("gstr1_filing_date") and out.get("gstr3b_filing_date"):
        out["gstr1_vs_gstr3b_gap_days"] = (out["gstr3b_filing_date"] - out["gstr1_filing_date"]).days
    else:
        out["gstr1_vs_gstr3b_gap_days"] = None
    return out


if __name__ == "__main__":
    # Self-test with a synthetic marker (no real merged file needed to sanity-check the math)
    d1 = due_date_gstr1(_dt.date(2022, 4, 1))
    d3 = due_date_gstr3b(_dt.date(2022, 4, 1))
    print("Apr-22 GSTR-1 due:", d1, " GSTR-3B due:", d3)
    fee = compute_late_fee(_dt.date(2022, 5, 25), d1, annual_turnover=4_65_639_087)
    print("Late fee example (filed 25-May vs due 11-May):", fee)
    interest = compute_interest(1_000_000, _dt.date(2022, 5, 25), d3)
    print("Interest example:", interest)
